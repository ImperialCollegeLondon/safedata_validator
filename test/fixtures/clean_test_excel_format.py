import os

import openpyxl

"""
There were some oddities in the original test files that caused warnings,
which I think were to do with external data sources. This code simply copies
the data from one file into another, which should tidy up the contents.
"""

os.chdir("test/fixtures")
path = "Test_format_good.xlsx"

# Open the original and create a new file
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
new = openpyxl.Workbook()

# Loop the sheets and rows - this will lose custom formats (e.g. dates)
for sh in wb:
    new_ws = new.create_sheet(sh.title)
    for rw in sh.values:
        new_ws.append(rw)

# new was created with a single sheet - remove that
del new["Sheet"]

# Save it
new.save("Test_format_good_new.xlsx")
