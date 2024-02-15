"""Collection of fixtures to assist the testing scripts."""

import os
import sys
from collections import OrderedDict

import appdirs
import certifi
import openpyxl
import pytest
from dotmap import DotMap

from safedata_validator.field import Dataset, DataWorksheet
from safedata_validator.locations import Locations
from safedata_validator.resources import Resources
from safedata_validator.taxa import GBIFTaxa, GBIFValidator, NCBIValidator


def fixture_files():
    """Function to direct test functions to the fixture files.

    Whenever testing is run, this conftest file is loaded and the path of the
    file can be used to provide paths to the location of all other testing and
    fixture files on a given test system. This helps make sure that absolute
    paths are maintained so that testing is not sensitive to the directory where
    tests are run.

    The Dotmap contains real files (.rf) that will be copied into the fake
    filesystem, along with key paths to virtual files (.vf) created within the
    same system. A missing file (.mf) is also provided to test responses to
    missing files.
    """

    fixture_dir = os.path.join(os.path.dirname(__file__), "fixtures")

    real_files = [
        ("gaz_file", "gazetteer_simple.geojson"),
        ("localias_file", "location_aliases_simple.csv"),
        ("empty_localias_file", "location_aliases_empty.csv"),
        ("project_database_file", "project_database_simple.csv"),
        ("project_database_file_bad", "project_database_bad.csv"),
        ("gbif_file", "gbif_backbone_truncated.sqlite"),
        ("ncbi_file", "ncbi_database_truncated.sqlite"),
        ("json_not_locations", "notalocationsjson.json"),
        ("bad_excel_file", "Test_format_bad.xlsx"),
        ("good_excel_file", "Test_format_good.xlsx"),
        ("bad_ncbi_file", "Test_format_bad_NCBI.xlsx"),
        ("weird_ncbi_file", "Test_format_weird_NCBI.xlsx"),
        ("good_ncbi_file", "Test_format_good_NCBI.xlsx"),
        ("good_ncbi_file_dataset_json", "Test_format_good_NCBI.json"),
        ("good_ncbi_file_zenodo_json", "zenodo_27557.json"),
    ]

    real_files = {ky: os.path.join(fixture_dir, vl) for ky, vl in real_files}

    # Need to provide the path to the certifi CA bundle or requests breaks!
    real_files["certifi"] = certifi.where()

    # Virtual file paths for the locations of config files.
    virtual_files = {
        "user_config": os.path.join(
            appdirs.user_config_dir(), "safedata_validator", "safedata_validator.cfg"
        ),
        "site_config": os.path.join(
            appdirs.site_config_dir(), "safedata_validator", "safedata_validator.cfg"
        ),
        "fix_config": os.path.join(fixture_dir, "safedata_validator.cfg"),
        "fix_config_no_projects": os.path.join(
            fixture_dir, "safedata_validator_no_proj.cfg"
        ),
    }

    return DotMap(
        dict(
            rf=real_files,
            vf=virtual_files,
            mf=os.path.join(fixture_dir, "thisfiledoesnotexist"),
        )
    )


FIXTURE_FILES = fixture_files()
"""DotMap: This global variable contains a dotmap of the paths of various
key testing files. This could be a fixture, but these paths are used with
parameterisation as well as within tests, and using fixture values in
parameterisation is clumsy and complex.
"""

# Populate the virtual filesystem along with clones that provide user and site config
# files


@pytest.fixture()
def config_filesystem(fs):
    """Create a config and resources file system for testing.

    Testing requires access to the configuration files for the package resources
    and the paths to these files are going to differ across test system. This
    fixture uses the pyfakefs plugin for pytest to create a fake file system
    containing config files and versions of the resources, cut down to save
    filespace.

    See also: https://gist.github.com/peterhurford/09f7dcda0ab04b95c026c60fa49c2a68

    Args:
        fs: The pyfakefs plugin object

    Returns:
        A fake filesystem containing copies of key actual fixture files in their
        correct location, along with virtual config files containing the correct
        paths to those fixtures files.
    """

    # Point to real locations of test fixture example resources
    for ky, val in FIXTURE_FILES.rf.items():
        fs.add_real_file(val)

    # The config files _cannot_ be real files because they need to
    # contain absolute paths to the resources available on the specific
    # test machine, and so need to be created with those paths

    config_contents = [
        "gazetteer = ",
        "location_aliases = ",
        "gbif_database = ",
        "ncbi_database = ",
        "project_database = ",
        "[extents]",
        "temporal_soft_extent = 2002-02-02, 2030-01-31",
        "temporal_hard_extent = 2002-02-01, 2030-02-01",
        "latitudinal_hard_extent = -90, 90",
        "latitudinal_soft_extent = -4, 8",
        "longitudinal_hard_extent = -180, 180",
        "longitudinal_soft_extent = 110, 120",
        "[zenodo]",
        "community_name = safe",
        "use_sandbox = True",
        "zenodo_sandbox_api = https://sandbox.zenodo.org",
        "zenodo_sandbox_token = xyz",
        "zenodo_api = https://api.zenodo.org",
        "zenodo_token = xyz",
    ]

    # Create config with local paths in the fixture directory
    config_contents[0] += FIXTURE_FILES.rf.gaz_file
    config_contents[1] += FIXTURE_FILES.rf.localias_file
    config_contents[2] += FIXTURE_FILES.rf.gbif_file
    config_contents[3] += FIXTURE_FILES.rf.ncbi_file
    fs.create_file(
        FIXTURE_FILES.vf.fix_config_no_projects, contents="\n".join(config_contents)
    )

    config_contents[4] += FIXTURE_FILES.rf.project_database_file
    fs.create_file(FIXTURE_FILES.vf.fix_config, contents="\n".join(config_contents))

    yield fs


@pytest.fixture()
def user_config_file(config_filesystem):
    """Creates local user config.

    Duplicate of the existing config in the fixture directory
    """

    with open(FIXTURE_FILES.vf.fix_config) as infile:
        config_filesystem.create_file(
            FIXTURE_FILES.vf.user_config, contents="".join(infile.readlines())
        )

    yield config_filesystem


@pytest.fixture()
def site_config_file(config_filesystem):
    """Creates local site config.

    Duplicate of the existing config in the fixture directory
    """
    with open(FIXTURE_FILES.vf.fix_config) as infile:
        config_filesystem.create_file(
            FIXTURE_FILES.vf.site_config, contents="".join(infile.readlines())
        )

    yield config_filesystem


# Helper function for validation of log contents


def log_check(caplog, expected_log):
    """Helper function to check that the captured log is as expected.

    Arguments:
        caplog: An instance of the caplog fixture
        expected_log: An iterable of 2-tuples containing the
            log level and message.
    """

    assert len(expected_log) == len(caplog.records)

    level_correct = [
        exp[0] == rec.levelno for exp, rec in zip(expected_log, caplog.records)
    ]

    message_correct = [
        exp[1] in rec.message for exp, rec in zip(expected_log, caplog.records)
    ]

    if not all(level_correct):
        failed_records = (
            (exp, obs)
            for passed, exp, obs in zip(level_correct, expected_log, caplog.records)
            if not passed
        )
        for obs, exp in failed_records:
            sys.stderr.write(f"Log level mismatch: {obs}, {exp.levelno, exp.message}")

        assert False

    if not all(message_correct):
        failed_records = (
            (exp, obs)
            for passed, exp, obs in zip(message_correct, expected_log, caplog.records)
            if not passed
        )
        for obs, exp in failed_records:
            sys.stderr.write(f"Log message mismatch: {obs}, {exp.levelno, exp.message}")

        assert False


# ------------------------------------------
# Resources: fixtures containing local Resources instances
# ------------------------------------------


@pytest.fixture()
def fixture_resources(config_filesystem):
    """Creates a Resource object configured to use local NCBI and GBIF.

    Returns:
        A safedata_validator.resources.Resources instance
    """

    return Resources(config=FIXTURE_FILES.vf.fix_config)


# ------------------------------------------
# Other fixtures
# ------------------------------------------


@pytest.fixture()
def example_excel_files(config_filesystem, request):
    """Fixture providing example excel files for testing.

    This uses indirect parameterisation, to allow the shared fixture to be paired with
    request specific expectations rather than all pair combinations:

    https://stackoverflow.com/questions/70379640
    """
    if request.param == "good":
        wb = openpyxl.load_workbook(FIXTURE_FILES.rf.good_excel_file, read_only=True)
        return wb
    elif request.param == "bad":
        wb = openpyxl.load_workbook(FIXTURE_FILES.rf.bad_excel_file, read_only=True)
        return wb


@pytest.fixture()
def example_ncbi_files(config_filesystem, request):
    """Fixture providing additional excel files with NCBI taxa details for testing.

    This uses indirect parameterisation, to allow the shared fixture to be paired with
    request specific expectations rather than all pair combinations. This is an
    equivalent function to example_excel_files but for the NCBI specific excel files.
    """
    if request.param == "good":
        wb = openpyxl.load_workbook(FIXTURE_FILES.rf.good_ncbi_file, read_only=True)
        return wb
    elif request.param == "bad":
        wb = openpyxl.load_workbook(FIXTURE_FILES.rf.bad_ncbi_file, read_only=True)
        return wb
    elif request.param == "weird":
        wb = openpyxl.load_workbook(FIXTURE_FILES.rf.weird_ncbi_file, read_only=True)
        return wb


@pytest.fixture()
def fixture_gbif_validator(fixture_resources):
    """Fixture to return GBIF taxon validators."""

    return GBIFValidator(fixture_resources)


@pytest.fixture()
def fixture_ncbi_validator(fixture_resources):
    """Fixture to NCBI validator."""

    return NCBIValidator(fixture_resources)


# Fixtures to provide Taxon, Locations, Dataset, Dataworksheet
# and field meta objects for testing


@pytest.fixture()
def fixture_taxa(fixture_resources):
    """Fixture to provide a taxon object with a couple of names.

    These examples need to be in the cutdown local GBIF testing database in fixtures.
    """

    taxa = GBIFTaxa(fixture_resources)

    test_taxa = [
        ("C_born", ["Crematogaster borneensis", "Species", None, None], None),
        ("V_salv", ["Varanus salvator", "Species", None, None], None),
    ]

    for tx in test_taxa:
        taxa.validate_and_add_taxon(tx)

    return taxa


@pytest.fixture()
def fixture_locations(fixture_resources):
    """Fixture to provide a taxon object with a couple of names.

    These examples need to be in the cutdown local GBIF testing database in fixtures.
    """

    locations = Locations(fixture_resources)

    test_locs = ["A_1", "A_2", 1, 2]

    locations.add_known_locations(test_locs)

    return locations


@pytest.fixture()
def fixture_dataset(fixture_resources):
    """Fixture to provide a dataset.

    This dataset has been prepopulated with some taxon and location names for field
    tests.
    """

    dataset = Dataset(fixture_resources)

    test_taxa = [
        ("C_born", ["Crematogaster borneensis", "Species", None, None], None),
        ("V_salv", ["Varanus salvator", "Species", None, None], None),
    ]

    for tx in test_taxa:
        dataset.taxa.gbif_taxa.validate_and_add_taxon(tx)

    test_locs = ["A_1", "A_2", 1, 2]

    dataset.locations.add_known_locations(test_locs)

    return dataset


@pytest.fixture(scope="module")
def fixture_field_meta():
    """field_meta object for use across tests."""

    return OrderedDict(
        field_type=["numeric", "numeric", "numeric"],
        description=["a", "b", "c"],
        units=["a", "b", "c"],
        method=["a", "b", "c"],
        field_name=["a", "b", "c"],
    )


@pytest.fixture()
def fixture_dataworksheet(fixture_dataset):
    """field_meta object for use across tests."""

    dws = DataWorksheet(
        {
            "name": "DF",
            "title": "My data table",
            "description": "This is a test data worksheet",
            "external": None,
        },
        dataset=fixture_dataset,
    )

    return dws


@pytest.fixture()
def mocked_requests(requests_mock):
    """Prototype requests mocker.

    This fixture can be switched from mocking the request URLs to using the actual URLs
    by setting the SDV_DO_NOT_MOCK_REQUESTS environment variable. This allows test to be
    run using the actual APIs to validate that the mocked and real APIs have not
    diverged. This will result in some long running tests and we will need to configure
    GitHub Action secrets and other environment variables to interact with Zenodo.
    """

    requests_mock.real_http = True

    if "SDV_DO_NOT_MOCK_REQUESTS" not in os.environ:
        requests_mock.get("http://google.com", text="data")
