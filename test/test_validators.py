import pytest
import os
import string
import openpyxl
from safedata_validator.validators import *

@pytest.fixture()
def dummy_ws_with_extra():

    wb = openpyxl.Workbook()
    ws = wb.active

    # Create a 11 by 38 simple block of data with A-K column headers
    # containing a multiplication table.
    for col in range(1, 12):
        ws.cell(column=col, row=1).value = openpyxl.utils.get_column_letter(col)
        for row in range(2, 40):
            ws.cell(column=col, row=row).value = (row - 1) * col

    # Put in an empty cell at 20, 50 which leads to the kind of extra rows
    # and columns that often occur with Excel
    ws.cell(column=20, row=50).value = None

    return ws


def test_naive_dims(dummy_ws_with_extra):
    """
    Just checks that a naive read does indeed give inflated frame size
    """
    assert dummy_ws_with_extra.max_column == 20
    assert dummy_ws_with_extra.max_row == 50


@pytest.mark.parametrize(
    'hd_vals,nonstr,padded,duped,blank',
    [(None, [], [], [], []),
     (['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K'],
      [], [], [], []),  # replacement works!
     (['A', 'B', 'C', 'D', 'E', '  \n', 'G', 'H', None, 'J', 'J'],
      [], [], [], ['  \n', None]),  # blanks
     (['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'J'],
      [], [], ['J'], []),  # dupe
     (['A', 'B', 3, 'D', 'E', 'F', False, 'H', 'I', 'J', 'K'],
      [3, False], [], [], []),  # nonstring
     (['A', ' B', 'C', 'D ', 'E', 'F', 'G', 'H', 'I', 'J', 'K'],
      [], [' B', 'D '], [], []),  # padded
     (['A', ' B', 3, 'D ', 'E', '  \n', False, 'H', None, 'J', 'J'],
      [3, False, None], [' B', 'D '], ['J'], ['  \n', None]),  # all - None caught twice
     ]
)
def test_getdataframe(dummy_ws_with_extra, hd_vals, nonstr, padded, duped, blank):
    """
    Does GetDataFrame work?
    """

    # Update headers for header checking
    if hd_vals:
        for col, val in enumerate(hd_vals):
            dummy_ws_with_extra.cell(1, col + 1).value = val

    dataframe = GetDataFrame(dummy_ws_with_extra)

    # Check dimensions
    assert len(dataframe.headers) == 11
    assert all([len(v) == 38 for v in dataframe.data_columns])

    # Use sum to check contents
    def _times_table(v, to):

        return [v * (r + 1) for r in range(to)]

    assert all([sum(v) == sum(_times_table(idx + 1, 38))
                for idx, v in enumerate(dataframe.data_columns)])

    # Header checking
    if nonstr:
        assert (('non_string' in dataframe.bad_headers)and
                (dataframe.bad_headers['non_string'] == nonstr))

    if padded:
        assert (('padded' in dataframe.bad_headers)and
                (dataframe.bad_headers['padded'] == padded))

    if duped:
        assert (('duplicated' in dataframe.bad_headers)and
                (dataframe.bad_headers['duplicated'] == duped))

    if blank:
        assert (('blank' in dataframe.bad_headers)and
                (dataframe.bad_headers['blank'] == blank))


@pytest.mark.parametrize(
    'input_vals,duplicated,bool_val',
    [(['a', 'b', 'c', 'd'], [], False),
     (['a', 'b', 'c', 'd', 'a'], ['a'], True),
     ]
)
def test_hasduplicates(input_vals, duplicated, bool_val):

    out = HasDuplicates(input_vals)

    assert out.duplicated == duplicated
    assert out.__bool__() is bool_val # Can't use out directly because class Filter != bool


@pytest.mark.parametrize(
    'input_vals,output_vals',
    [(['a', 'b', 'c', 'd'], ['a', 'b', 'c', 'd']),
     (['A', 'b', 'C', 'd'], ['a', 'b', 'c', 'd']),
     (['A', 2, 'C', 'd'], ['a', 2, 'c', 'd']),
     (['A', 'b', 'C', None], ['a', 'b', 'c', None]),
     ]
)
def test_islower(input_vals, output_vals):

    out = IsLower(input_vals)

    assert out.values == output_vals


# TODO - more unit tests on other validators.