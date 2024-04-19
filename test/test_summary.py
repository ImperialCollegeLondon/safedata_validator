"""Tests to check that the summary sheet functions work as intended."""

import datetime
from logging import ERROR, INFO, WARNING

import pytest
from openpyxl import Workbook

from safedata_validator.logger import LOGGER
from safedata_validator.resources import Resources
from safedata_validator.summary import Summary

from .conftest import FIXTURE_FILES, log_check


@pytest.fixture
def fixture_summary(fixture_resources):
    """Fixture providing summary instances for testing."""
    return Summary(fixture_resources)


@pytest.fixture()
def fixture_summary_projects(config_filesystem, request):
    """Fixture providing summary instances for testing.

    This uses indirect parameterisation to allow tests to pick between a summary with
    projects and without.
    """

    if not request.param:
        return Summary(Resources(FIXTURE_FILES.vf.fix_config_no_projects))

    return Summary(Resources(FIXTURE_FILES.vf.fix_config))


# NOTE - _read_block is being tested by repeated _read_`block` calls and
#        this could be simplified to reduce the code, but there are block
#        specific variations (types, mandatory etc) that make it easier
#        to test block by block


@pytest.mark.parametrize(
    "alterations,should_log_error,expected_log",
    [
        (dict(), False, "Metadata for Authors found"),  # no amendments
        (
            {"author name": (None,)},
            True,
            "Missing metadata in mandatory field author name",
        ),  # via _read_block
        (
            {"author name": ("   ",)},
            True,
            "Whitespace only cells in field author name",
        ),  # via _read_block
        (
            {"author name": (123456789,)},
            True,
            "Field author name contains values of wrong type",
        ),  # via _read_block
        (
            {"author name": ("David Orme",)},
            True,
            "Author names not formatted as last_name, first_names",
        ),
        (
            {"author affiliation": (None,)},
            False,  # missing affiliation acceptable
            "Metadata for Authors found",
        ),
        (
            {"author affiliation": ("   ",)},
            True,
            "Whitespace only cells in field author affiliation",
        ),  # via _read_block
        (
            {"author affiliation": (123456789,)},
            True,
            "Field author affiliation contains values of wrong type",
        ),  # via _read_block
        (
            {"author email": (None,)},
            False,  # missing email acceptable
            "Metadata for Authors found",
        ),
        (
            {"author email": ("   ",)},
            True,
            "Whitespace only cells in field author email",
        ),  # via _read_block
        (
            {"author email": (123456789,)},
            True,
            "Field author email contains values of wrong type",
        ),  # via _read_block
        (
            {"author email": ("d.ormeatimperial.ac.uk",)},
            True,
            "Author emails not properly formatted",
        ),
        (
            {"author orcid": (None,)},
            False,  # Blank ORCID is OK
            "Metadata for Authors found",
        ),
        (
            {"author orcid": ("   ",)},
            True,
            "Whitespace only cells in field author orcid",
        ),  # via _read_block
        (
            {"author orcid": (123456789,)},
            True,
            "Field author orcid contains values of wrong type",
        ),  # via _read_block
        (
            {"author orcid": ("123456789",)},
            True,
            "Author ORCIDs not properly formatted",
        ),
    ],
)
def test_authors(caplog, fixture_summary, alterations, should_log_error, expected_log):
    # Valid set of information
    input = {
        "author name": ("Orme, David",),
        "author email": ("d.orme@imperial.ac.uk",),
        "author affiliation": ("Imperial College London",),
        "author orcid": ("0000-0002-7005-1394",),
    }

    # Update valid to test error conditions and populate _rows
    # directly (bypassing .load() and need to pack in worksheet object
    input.update(alterations)
    fixture_summary._rows = input

    # Test the block load
    fixture_summary._load_authors()

    if should_log_error:
        assert "ERROR" in [r.levelname for r in caplog.records]

    assert expected_log in caplog.text


@pytest.mark.parametrize(
    "row_data,should_log_error,expected_log",
    [
        (
            {
                "access status": ("Open",),  # Open
                "embargo date": (None,),
                "access conditions": (None,),
            },
            False,
            "Metadata for Access details found",
        ),
        (
            {
                "access status": (123,),  # Bad access status type
                "embargo date": (None,),
                "access conditions": (None,),
            },
            True,
            "Field access status contains values of wrong type",
        ),
        (
            {
                "access status": ("Oeopn",),  # Bad access status value
                "embargo date": (None,),
                "access conditions": (None,),
            },
            True,
            "Access status must be Open, Embargo or Restricted",
        ),
        (
            {
                "access status": ("Embargo",),  # Embargoed correctly
                "embargo date": (
                    datetime.datetime.today() + datetime.timedelta(days=365),
                ),
                "access conditions": (None,),
            },
            False,
            "Metadata for Access details found",
        ),
        (
            {
                "access status": ("Embargo",),  # Bad embargo date type
                "embargo date": (123,),
                "access conditions": (None,),
            },
            True,
            "Field embargo date contains values of wrong type",
        ),
        (
            {
                "access status": ("Embargo",),  # Embargo date in the past
                "embargo date": (
                    datetime.datetime.today() - datetime.timedelta(days=365),
                ),
                "access conditions": (None,),
            },
            True,
            "Embargo date is in the past",
        ),
        (
            {
                "access status": ("Embargo",),  # Embargo date too far into the future
                "embargo date": (
                    datetime.datetime.today() + datetime.timedelta(days=5 * 365),
                ),
                "access conditions": (None,),
            },
            True,
            "Embargo date more than",
        ),
        (
            {
                "access status": ("Embargo",),  # Conditions set on embargoed data
                "embargo date": (
                    datetime.datetime.today() + datetime.timedelta(days=365),
                ),
                "access conditions": ("My precious, mine",),
            },
            True,
            "Access conditions cannot be set on embargoed data",
        ),
        (
            {
                "access status": ("Restricted",),  # Restricted
                "embargo date": (None,),
                "access conditions": ("My precious, mine",),
            },
            False,
            "Metadata for Access details found",
        ),
        (
            {
                "access status": ("Restricted",),  # Restricted but with no conditions
                "embargo date": (None,),
                "access conditions": (None,),
            },
            True,
            "Dataset restricted but no access conditions specified",
        ),
        (
            {
                "access status": ("Restricted",),  # Bad type on access conditions
                "embargo date": (None,),
                "access conditions": (123,),
            },
            True,
            "Field access conditions contains values of wrong type",
        ),
        (
            {
                "access status": ("Restricted",),  # Embargo date set on restricted data
                "embargo date": (
                    datetime.datetime.today() + datetime.timedelta(days=365),
                ),
                "access conditions": ("My precious, mine",),
            },
            True,
            "Do not set an embargo date with restricted datasets",
        ),
        (
            {
                "access status": ("Open", "Embargoed"),  # Not a singleton record
                "embargo date": (None, None),
                "access conditions": (None, None),
            },
            True,
            "Only a single record should be present",
        ),
    ],
)
def test_access(caplog, fixture_summary, row_data, should_log_error, expected_log):
    # Update valid to test error conditions and populate _rows
    # directly (bypassing .load() and need to pack in worksheet object
    fixture_summary._rows = row_data

    # Test the block load
    fixture_summary._load_access_details()

    if should_log_error:
        assert "ERROR" in [r.levelname for r in caplog.records]

    assert expected_log in caplog.text


@pytest.mark.parametrize(
    "alterations,should_log_error,expected_log",
    [
        (dict(), False, "Metadata for Keywords found"),  # no amendments,
        (
            {"keywords": (None, None, None)},
            True,
            "No Keywords metadata found",
        ),  # via _read_block
        (
            {"keywords": ("   ", "abc", "def")},
            True,
            "Whitespace only cells in field keywords",
        ),  # via _read_block
        (
            {"keywords": (123456789, 12.0)},
            True,
            "Field keywords contains values of wrong type",
        ),  # via _read_block
        (
            {"keywords": ("abc,def", "ghi;jkl")},
            True,
            "Put each keyword in a separate cell",
        ),
    ],
)
def test_keywords(caplog, fixture_summary, alterations, should_log_error, expected_log):
    # Valid set of information
    input = {"keywords": ("abc", "def")}

    # Update valid to test error conditions and populate _rows
    # directly (bypassing .load() and need to pack in worksheet object
    input.update(alterations)
    fixture_summary._rows = input

    # Test the block load
    fixture_summary._load_keywords()

    if should_log_error:
        assert "ERROR" in [r.levelname for r in caplog.records]

    assert expected_log in caplog.text


@pytest.mark.parametrize(
    "alterations,should_log_error,expected_log",
    [
        (dict(), False, "Metadata for Permits found"),  # no amendments,
        (
            {
                "permit type": (None,),
                "permit authority": (None,),
                "permit number": (None,),
            },
            False,  # Optional block
            "",
        ),
        (
            {"permit type": (None,)},
            True,
            "Missing metadata in mandatory field permit type",
        ),  # via _read_block
        (
            {"permit type": ("   ",)},
            True,
            "Whitespace only cells in field permit type",
        ),  # via _read_block
        (
            {"permit type": (123456789,)},
            True,
            "Field permit type contains values of wrong type",
        ),  # via _read_block
        (
            {"permit authority": (None,)},
            True,
            "Missing metadata in mandatory field permit authority",
        ),  # via _read_block
        (
            {"permit authority": ("   ",)},
            True,
            "Whitespace only cells in field permit authority",
        ),  # via _read_block
        (
            {"permit authority": (123456789,)},
            True,
            "Field permit authority contains values of wrong type",
        ),  # via _read_block
        (
            {"permit number": (None,)},
            True,
            "Missing metadata in mandatory field permit number",
        ),  # via _read_block
        (
            {"permit number": ("   ",)},
            True,
            "Whitespace only cells in field permit number",
        ),  # via _read_block
        (
            {"permit number": (123456789,)},
            False,  # Allow numeric numbers!
            "Metadata for Permits found",
        ),
        (
            {"permit number": (datetime.date(1901, 1, 1),)},
            True,
            "Field permit number contains values of wrong type",
        ),  # via _read_block
    ],
)
def test_permits(caplog, fixture_summary, alterations, should_log_error, expected_log):
    # Valid set of information
    input = {
        "permit type": ("research",),
        "permit authority": ("Sabah Biodiversity Centre",),
        "permit number": ("ABC-123-456",),
    }

    # Update valid to test error conditions and populate _rows
    # directly (bypassing .load() and need to pack in worksheet object
    input.update(alterations)
    fixture_summary._rows = input

    # Test the block load
    fixture_summary._load_permits()

    if should_log_error:
        assert "ERROR" in [r.levelname for r in caplog.records]

    assert expected_log in caplog.text


@pytest.mark.parametrize(
    "alterations,should_log_error,expected_log,do_val_doi",
    [
        (dict(), False, "Metadata for DOI found", False),  # no amendments,
        ({"publication doi": (None,)}, False, "", False),  # Optional block
        (
            {"publication doi": ("   ",)},
            True,
            "Whitespace only cells in field publication doi",
            False,
        ),  # via _read_block
        (
            {"publication doi": (123456789,)},
            True,
            "Field publication doi contains values of wrong type",
            False,
        ),  # via _read_block
        (
            {"publication doi": ("thisisnotaURL",)},
            True,
            "Publication DOIs not all in format",
            False,
        ),
        (
            {"publication doi": ("https://doi.org/this.does/not.exist",)},
            True,
            "DOI not found",
            True,
        ),
    ],
)
def test_doi(
    caplog, fixture_summary, alterations, should_log_error, expected_log, do_val_doi
):
    # Valid set of information
    input = {"publication doi": ("https://doi.org/10.1098/rstb.2011.0049",)}

    # Update valid to test error conditions and populate _rows
    # directly (bypassing .load() and need to pack in worksheet object
    input.update(alterations)
    fixture_summary._rows = input
    fixture_summary.validate_doi = do_val_doi

    # Test the block load
    fixture_summary._load_doi()

    if should_log_error:
        assert "ERROR" in [r.levelname for r in caplog.records]

    assert expected_log in caplog.text


@pytest.mark.parametrize(
    "alterations,should_log_error,expected_log",
    [
        (dict(), False, "Metadata for Funding Bodies found"),  # no amendments
        (
            {
                "funding body": (None,),
                "funding type": (None,),
                "funding reference": (None,),
                "funding link": (None,),
            },
            False,  # Not a mandatory block
            "",
        ),
        (
            {"funding body": (None,)},
            True,
            "Missing metadata in mandatory field funding body",
        ),  # via _read_block
        (
            {"funding body": ("   ",)},
            True,
            "Whitespace only cells in field funding body",
        ),  # via _read_block
        (
            {"funding body": (123456789,)},
            True,
            "Field funding body contains values of wrong type",
        ),  # via _read_block
        (
            {"funding type": (None,)},
            True,
            "Missing metadata in mandatory field funding type",
        ),  # via _read_block
        (
            {"funding type": ("   ",)},
            True,
            "Whitespace only cells in field funding type",
        ),  # via _read_block
        (
            {"funding type": (123456789,)},
            True,
            "Field funding type contains values of wrong type",
        ),  # via _read_block
        (
            {"funding reference": (None,)},
            False,
            "Metadata for Funding Bodies found",
        ),  # Not mandatory
        (
            {"funding reference": ("   ",)},
            True,
            "Whitespace only cells in field funding reference",
        ),  # via _read_block
        (
            {"funding reference": (123456789,)},
            False,
            "Metadata for Funding Bodies found",
        ),  # Could be numeric
        (
            {"funding reference": (datetime.date(2011, 1, 1),)},
            True,
            "Field funding reference contains values of wrong type",
        ),  # via _read_block
        (
            {"funding link": (None,)},
            False,
            "Metadata for Funding Bodies found",
        ),  # Not mandatory
        (
            {"funding link": ("   ",)},
            True,
            "Whitespace only cells in field funding link",
        ),  # via _read_block
        (
            {"funding link": (123456789,)},
            True,
            "Field funding link contains values of wrong type",
        ),  # via _read_block
    ],
)
def test_funders(caplog, fixture_summary, alterations, should_log_error, expected_log):
    # Valid set of information
    input = {
        "funding body": ("NERC",),
        "funding type": ("Standard grant",),
        "funding reference": ("NE/K006339/1",),
        "funding link": ("https://gtr.ukri.org/projects?ref=NE%2FK006339%2F1",),
    }

    # Update valid to test error conditions and populate _rows
    # directly (bypassing .load() and need to pack in worksheet object
    input.update(alterations)
    fixture_summary._rows = input

    # Test the block load
    fixture_summary._load_funders()

    if should_log_error:
        assert "ERROR" in [r.levelname for r in caplog.records]

    assert expected_log in caplog.text


@pytest.mark.parametrize(
    "alterations,should_log_error,expected_log",
    [
        (dict(), False, "Metadata for Date Extents found"),  # no amendments
        (
            {"start date": (None,), "end date": (None,)},
            False,  # Not a mandatory block
            "No Date Extents metadata found",
        ),
        (
            {"start date": (None,)},
            True,
            "Missing metadata in mandatory field start date",
        ),  # via _read_block
        (
            {"start date": ("   ",)},
            True,
            "Whitespace only cells in field start date",
        ),  # via _read_block
        (
            {"start date": ("2001-01-01",)},
            True,
            "Field start date contains values of wrong type",
        ),  # via _read_block
        (
            {"end date": (None,)},
            True,
            "Missing metadata in mandatory field end date",
        ),  # via _read_block
        (
            {"end date": ("   ",)},
            True,
            "Whitespace only cells in field end date",
        ),  # via _read_block
        (
            {"end date": ("2001-01-01",)},
            True,
            "Field end date contains values of wrong type",
        ),  # via _read_block
        (
            {
                "start date": (
                    datetime.datetime(2001, 1, 1),
                    datetime.datetime(2001, 1, 1),
                ),
                "end date": (
                    datetime.datetime(2011, 1, 1),
                    datetime.datetime(2011, 1, 1),
                ),
            },
            True,
            "Only a single record should be present",
        ),
        (
            {
                "start date": (datetime.datetime(2011, 1, 1),),
                "end date": (datetime.datetime(2001, 1, 1),),
            },
            True,
            "Start date is after end date",
        ),
    ],
)
def test_temporal_extent(
    caplog, fixture_summary, alterations, should_log_error, expected_log
):
    # Valid set of information - openpyxl loads dates as datetimes but
    # the validation checks that these values are dates
    # (have no time information)
    input = {
        "start date": (datetime.datetime(2001, 1, 1),),
        "end date": (datetime.datetime(2011, 1, 1),),
    }

    # Update valid to test error conditions and populate _rows
    # directly (bypassing .load() and need to pack in worksheet object
    input.update(alterations)
    fixture_summary._rows = input

    # Test the block load
    fixture_summary._load_temporal_extent()

    if should_log_error:
        assert "ERROR" in [r.levelname for r in caplog.records]
    else:
        # Check the extent is updated - note that datetime.date is enforced
        start = input["start date"][0]
        end = input["end date"][0]
        expected_dates = (
            start if start is None else start.date(),
            end if end is None else end.date(),
        )
        assert fixture_summary.temporal_extent.extent == expected_dates

    assert expected_log in caplog.text


@pytest.mark.parametrize(
    "alterations,should_log_error,expected_log",
    [
        (dict(), False, "Metadata for Geographic Extents found"),  # no amendments
        (
            {"south": (None,), "north": (None,), "east": (None,), "west": (None,)},
            False,  # Not a mandatory block
            "",
        ),
        (
            {"south": (None,)},
            True,
            "Missing metadata in mandatory field south",
        ),  # via _read_block
        (
            {"south": ("   ",)},
            True,
            "Whitespace only cells in field south",
        ),  # via _read_block
        (
            {"south": ("2001-01-01",)},
            True,
            "Field south contains values of wrong type",
        ),  # via _read_block
        (
            {"north": (None,)},
            True,
            "Missing metadata in mandatory field north",
        ),  # via _read_block
        (
            {"north": ("   ",)},
            True,
            "Whitespace only cells in field north",
        ),  # via _read_block
        (
            {"north": ("2001-01-01",)},
            True,
            "Field north contains values of wrong type",
        ),  # via _read_block
        (
            {"east": (None,)},
            True,
            "Missing metadata in mandatory field east",
        ),  # via _read_block
        (
            {"east": ("   ",)},
            True,
            "Whitespace only cells in field east",
        ),  # via _read_block
        (
            {"east": ("2001-01-01",)},
            True,
            "Field east contains values of wrong type",
        ),  # via _read_block
        (
            {"west": (None,)},
            True,
            "Missing metadata in mandatory field west",
        ),  # via _read_block
        (
            {"west": ("   ",)},
            True,
            "Whitespace only cells in field west",
        ),  # via _read_block
        (
            {"west": ("2001-01-01",)},
            True,
            "Field west contains values of wrong type",
        ),  # via _read_block
        (
            {
                "west": (116.75, 116.75),
                "east": (117.82, 117.82),
                "south": (4.50, 4.50),
                "north": (5.07, 5.07),
            },
            True,
            "Only a single record should be present",
        ),
        (
            {"west": (117.82,), "east": (116.75,)},
            True,
            "West limit is greater than east limit",
        ),
        (
            {"south": (5.07,), "north": (4.50,)},
            True,
            "South limit is greater than north limit",
        ),
    ],
)
def test_geographic_extent(
    caplog, fixture_summary, alterations, should_log_error, expected_log
):
    # Valid set of information
    input = {"west": (116.75,), "east": (117.82,), "south": (4.50,), "north": (5.07,)}

    # Update valid to test error conditions and populate _rows
    # directly (bypassing .load() and need to pack in worksheet object
    input.update(alterations)
    fixture_summary._rows = input

    # Test the block load
    fixture_summary._load_geographic_extent()

    if should_log_error:
        assert "ERROR" in [r.levelname for r in caplog.records]

    assert expected_log in caplog.text


@pytest.mark.parametrize(
    "alterations,should_log_error,expected_log",
    [
        (dict(), False, "Metadata for External Files found"),  # no amendments
        (
            {"external file": (None,), "external file description": (None,)},
            False,  # Not a mandatory block
            "",
        ),
        (
            {"external file": (123,)},
            True,  # Non string filenames
            "Field external file contains values of wrong type",
        ),
        (
            {"external file description": (456,)},
            True,  # Non string description
            "Field external file description contains values of wrong type",
        ),
        (
            {"external file": (None, None)},
            True,  # Must provide names
            "Missing metadata in mandatory field external file",
        ),
        (
            {"external file description": (None, None)},
            True,  # Must describe files
            "Missing metadata in mandatory field external file description",
        ),
        (
            {"external file": ("   ", "   \t")},
            True,  # Empty names
            "Missing metadata in mandatory field external file",
        ),
        (
            {"external file": (" BaitTrapImages.zip", "BaitTrapTransects.geojson ")},
            True,  # Whitespace in filenames (captures padding too)
            "External file names must not contain whitespace",
        ),
        (
            {"external file": ("Bait Trap Images.zip", "Bait Trap Transects.geojson")},
            True,  # Whitespace in filenames (captures padding too)
            "External file names must not contain whitespace",
        ),
    ],
)
def test_external_files(
    caplog, fixture_summary, alterations, should_log_error, expected_log
):
    # Valid set of information
    input = {
        "external file": ("BaitTrapImages.zip", "BaitTrapTransects.geojson"),
        "external file description": (
            "Zip file containing 5000 JPEG images of bait trap cards",
            "GeoJSON file containing polylines of the bait trap transects",
        ),
    }

    # Update valid to test error conditions and populate _rows
    # directly (bypassing .load() and need to pack in worksheet object
    input.update(alterations)
    fixture_summary._rows = input

    # Test the block load
    fixture_summary._load_external_files()

    if should_log_error:
        assert "ERROR" in [r.levelname for r in caplog.records]

    assert expected_log in caplog.text


@pytest.mark.parametrize(
    "alterations,alt_sheets,ext_alterations,should_log_error,expected_log",
    [
        (dict(), None, dict(), False, "Metadata for Worksheets found"),  # no amendments
        (
            {
                "worksheet title": (
                    "My shiny dataset",
                    None,
                    "Bait trap transect lines",
                )
            },  # Missing data
            None,
            dict(),
            True,
            "Missing metadata in mandatory field worksheet title",
        ),
        (
            {"worksheet name": ("DF", None, "Transects")},  # Missing data
            None,
            dict(),
            True,
            "Missing metadata in mandatory field worksheet name",
        ),
        (
            {
                "worksheet description": (
                    "This is a test dataset",
                    "A test dataset too",
                    None,
                )
            },  # Missing data
            None,
            dict(),
            True,
            "Missing metadata in mandatory field worksheet description",
        ),
        (
            {
                "worksheet external file": (None, None, None)
            },  # Missing data not a problem for external files
            None,
            dict(),
            False,
            "Metadata for Worksheets found",
        ),
        (
            {"worksheet title": (123, 123, 123)},  # Type errors
            None,
            dict(),
            True,
            "Field worksheet title contains values of wrong type",
        ),
        (
            {"worksheet name": (123, 123, 123)},
            None,
            dict(),
            True,
            "Field worksheet name contains values of wrong type",
        ),
        (
            {"worksheet description": (123, 123, 123)},
            None,
            dict(),
            True,
            "Field worksheet description contains values of wrong type",
        ),
        (
            {"worksheet external file": (None, None, 123)},
            None,
            dict(),
            True,
            "Field worksheet external file contains values of wrong type",
        ),
        (
            {"worksheet name": ("NotInTheSheets",)},  # Unknown worksheet
            None,
            dict(),
            True,
            "Data worksheet NotInTheSheets not found",
        ),
        (
            dict(),  # Unused worksheet
            {"DF", "Incidence", "Transects", "Summary", "Taxa", "Locations", "NotUsed"},
            dict(),
            True,
            " Undocumented sheets found in workbook",
        ),
        (
            {
                "worksheet external file": (
                    None,
                    None,
                    "Amissingfile.dat",
                )
            },  # Unknown external file.
            None,
            dict(),
            True,
            "Worksheet descriptions refer to unreported external files",
        ),
        (
            {
                "worksheet title": ("My Taxa",),
                "worksheet name": ("Taxa",),  # Taxa included in data worksheets
                "worksheet description": ("A list of taxa",),
                "worksheet external file": (None,),
            },
            None,
            dict(),
            True,
            "Do not include Taxa or Locations metadata sheets in Data worksheet "
            "details",
        ),
        (
            {
                "worksheet title": ("My Locs",),
                "worksheet name": (
                    "Locations",
                ),  # Locations included in data worksheets
                "worksheet description": ("A list of locations",),
                "worksheet external file": (None,),
            },
            None,
            dict(),
            True,
            "Do not include Taxa or Locations metadata sheets in Data worksheet"
            " details",
        ),
        (
            {
                "worksheet title": (None,),
                "worksheet name": (None,),  # No data or external files
                "worksheet description": (None,),
                "worksheet external file": (None,),
            },
            None,
            {"external file": (None,), "external file description": (None,)},
            True,
            "No data worksheets or external files provided - no data.",
        ),
        (
            {
                "worksheet title": (None,),
                "worksheet name": (None,),  # Only external files
                "worksheet description": (None,),
                "worksheet external file": (None,),
            },
            None,
            dict(),
            False,
            "Only external file descriptions provided",
        ),
    ],
)
def test_data_worksheets(
    caplog,
    fixture_summary,
    alterations,
    alt_sheets,
    ext_alterations,
    should_log_error,
    expected_log,
):
    """Test that sheetnames are properly detected and handled.

    This test suite is more complex as the data worksheets need to match to the
    existing sheetnames in the workbook and potentially to a set of external files.
    """

    sheetnames = alt_sheets or {
        "DF",
        "Incidence",
        "Transects",
        "Summary",
        "Taxa",
        "Locations",
    }

    # Valid set of information
    input = {
        "worksheet title": (
            "My shiny dataset",
            "My incidence matrix",
            "Bait trap transect lines",
        ),
        "worksheet name": ("DF", "Incidence", "Transects"),
        "worksheet description": (
            "This is a test dataset",
            "A test dataset too",
            "Attribute table for transect GIS",
        ),
        "worksheet external file": (None, None, "BaitTrapTransects.geojson"),
    }

    # Update valid to test error conditions and populate _rows
    # directly (bypassing .load() and need to pack in worksheet object
    input.update(alterations)
    fixture_summary._rows = input

    # Populate the external files
    external = {
        "external file": ("BaitTrapImages.zip", "BaitTrapTransects.geojson"),
        "external file description": (
            "Zip file containing 5000 JPEG images of bait trap cards",
            "GeoJSON file containing polylines of the bait trap transects",
        ),
    }

    external.update(ext_alterations)
    fixture_summary._rows.update(external)

    # Load the external file data
    fixture_summary._load_external_files()

    # Test the block load
    fixture_summary._load_data_worksheets(sheetnames)

    if should_log_error:
        assert "ERROR" in [r.levelname for r in caplog.records]

    assert expected_log in caplog.text


@pytest.mark.parametrize(
    "example_excel_files, n_errors",
    [("good", 0), ("bad", 16)],
    indirect=["example_excel_files"],  # take actual params from fixture
)
def test_summary_load(fixture_summary, example_excel_files, n_errors):
    """This tests the ensemble loading of a summary from a file.

    It uses indirect parameterisation to access the fixtures containing the sample excel
    files.
    """

    fixture_summary.load(
        example_excel_files["Summary"], set(example_excel_files.sheetnames)
    )

    assert fixture_summary.n_errors == n_errors


@pytest.mark.parametrize(
    argnames=["fixture_summary_projects", "rows", "expected_log_entries"],
    indirect=["fixture_summary_projects"],
    argvalues=[
        pytest.param(
            False,
            {
                "title": ["title", None],
                "description": ["desc", None],
                "access status": ["open", None],
                "author name": ["Author, Anne", None],
                "keywords": ["testing", "testing"],
            },
            (
                (
                    (INFO, "Loading project id metadata"),
                    (INFO, "No Project IDs metadata found"),
                    (INFO, "No project id data required or provided."),
                )
            ),
            id="not required or provided",
        ),
        pytest.param(
            False,
            {
                "project id": [1, 2],
                "title": ["title", None],
                "description": ["desc", None],
                "access status": ["open", None],
                "author name": ["Author, Anne", None],
                "keywords": ["testing", "testing"],
            },
            (
                (INFO, "Loading project id metadata"),
                (INFO, "Metadata for Project IDs found:"),
                (ERROR, "Project ids are not required but are provided"),
            ),
            id="not required but provided",
        ),
        pytest.param(
            True,
            {
                "title": ["title", None],
                "description": ["desc", None],
                "access status": ["open", None],
                "author name": ["Author, Anne", None],
                "keywords": ["testing", "testing"],
            },
            (
                (INFO, "Loading project id metadata"),
                (ERROR, "No Project IDs metadata found"),  # block is mandatory
                (ERROR, "Project ids are required but not provided"),
            ),
            id="required but not provided",
        ),
        pytest.param(
            True,
            {
                "safe project id": [1, 2],
                "project id": [1, 2],
                "title": ["title", None],
                "description": ["desc", None],
                "access status": ["open", None],
                "author name": ["Author, Anne", None],
                "keywords": ["testing", "testing"],
            },
            (
                (INFO, "Loading project id metadata"),
                (INFO, "Metadata for Project IDs found:"),
                (
                    ERROR,
                    "Both 'project id' and 'safe project id' provided: "
                    "use only 'project id'",
                ),
                (INFO, "Valid project ids provided:"),
            ),
            id="both project id keys provided",
        ),
        pytest.param(
            True,
            {
                "safe project id": [1, 2],
                "title": ["title", None],
                "description": ["desc", None],
                "access status": ["open", None],
                "author name": ["Author, Anne", None],
                "keywords": ["testing", "testing"],
            },
            (
                (INFO, "Loading project id metadata"),
                (INFO, "Metadata for Project IDs found:"),
                (
                    WARNING,
                    "Use 'project id' rather than the legacy 'safe project id' key.",
                ),
                (INFO, "Valid project ids provided:"),
            ),
            id="using legacy safe project id",
        ),
        pytest.param(
            True,
            {
                "project id": [1, 3000],
                "title": ["title", None],
                "description": ["desc", None],
                "access status": ["open", None],
                "author name": ["Author, Anne", None],
                "keywords": ["testing", "testing"],
            },
            (
                (INFO, "Loading project id metadata"),
                (INFO, "Metadata for Project IDs found:"),
                (ERROR, "Unknown project ids provided"),
                (INFO, "Valid project ids provided:"),
            ),
            id="unknown ids",
        ),
        pytest.param(
            True,
            {
                "project id": [1, 2],
                "title": ["title", None],
                "description": ["desc", None],
                "access status": ["open", None],
                "author name": ["Author, Anne", None],
                "keywords": ["testing", "testing"],
            },
            (
                (INFO, "Loading project id metadata"),
                (INFO, "Metadata for Project IDs found:"),
                (INFO, "Valid project ids provided:"),
            ),
            id="all good",
        ),
    ],
)
def test_project_ids(caplog, fixture_summary_projects, rows, expected_log_entries):
    """Tests that project IDs are validated correctly.

    The fixture_summary_indirect will be configured with or without projects, depending
    on the indirect boolean value provided with each test.
    """

    # Add rows to fixture and toggle project use
    fixture_summary_projects._rows = rows
    fixture_summary_projects._ncols = 3

    # Check that the mandatory fields are raised by key validation and that project id
    # validation works as expected.
    fixture_summary_projects._validate_keys()
    fixture_summary_projects._load_project_ids()

    log_check(caplog, expected_log_entries)


@pytest.mark.parametrize(
    argnames=["data", "expected_rows"],
    argvalues=[
        pytest.param(
            [["header"]],
            [("header",)],
            id="simple case",
        ),
        pytest.param(
            [["header_1", 1, 3], ["header_2", 2, 6], ["header_3", 3, 9]],
            [("header_1", 1, 3), ("header_2", 2, 6), ("header_3", 3, 9)],
            id="complex case",
        ),
    ],
)
def test_load_rows_from_worksheet(data, expected_rows):
    """Test that function to load rows from a worksheet works as intended."""
    from safedata_validator.summary import load_rows_from_worksheet

    # Make an empty worksheet
    workbook = Workbook()
    worksheet = workbook.active

    # Loop through the rows of the list of lists and add them to the worksheet
    for row in data:
        worksheet.append(row)

    rows = load_rows_from_worksheet(worksheet)

    # Test parsed rows match what was expected
    assert rows == expected_rows


@pytest.mark.parametrize(
    argnames=["rows", "expected_log_entries"],
    argvalues=[
        pytest.param(
            {
                "title": [],
                "description": [],
                "access status": [],
                "author name": [],
                "keywords": [],
            },
            (),
            id="good no padding",
        ),
        pytest.param(
            {
                "title ": [],
                "description": [],
                "access status": [],
                "author name": [],
                "keywords": [],
            },
            ((ERROR, "Whitespace padding in summary field names:"),),
            id="padding after",
        ),
        pytest.param(
            {
                "title": [],
                " description": [],
                " access status": [],
                "author name": [],
                "keywords": [],
            },
            ((ERROR, "Whitespace padding in summary field names:"),),
            id="padding before",
        ),
    ],
)
def test_check_for_whitespace(caplog, fixture_summary, rows, expected_log_entries):
    """Test that function to check for whitespace padding in summary fields works."""

    # Add rows to fixture
    fixture_summary._rows = rows

    # Check that row row headers contain mandatory fields
    fixture_summary._check_for_whitespace()

    log_check(caplog, expected_log_entries)


@pytest.mark.parametrize(
    argnames=["rows", "expected_log_entries"],
    argvalues=[
        pytest.param(
            {
                "title": [],
                "description": [],
                "access status": [],
                "author name": [],
                "keywords": [],
            },
            (),
            id="good only mandatory",
        ),
        pytest.param(
            {
                "project id": [],
                "title": [],
                "description": [],
                "access status": [],
                "author name": [],
                "keywords": [],
            },
            (),
            id="good with optional",
        ),
        pytest.param(
            {
                "description": [],
                "access status": [],
                "author name": [],
                "keywords": [],
            },
            ((ERROR, "Missing mandatory metadata fields:"),),
            id="missing title",
        ),
        pytest.param(
            {
                "sproject id": [],
                "title": [],
                "description": [],
                "access status": [],
                "author name": [],
                "keywords": [],
            },
            ((ERROR, "Unknown metadata fields:"),),
            id="unknown field",
        ),
    ],
)
def test_validate_keys(caplog, fixture_summary, rows, expected_log_entries):
    """Test that function to check mandatory fields works as expected."""

    # Add rows to fixture
    fixture_summary._rows = rows

    # Check that row row headers contain mandatory fields
    fixture_summary._validate_keys()

    log_check(caplog, expected_log_entries)


@pytest.mark.parametrize(
    argnames=["rows", "expected_log_entries"],
    argvalues=[
        pytest.param(
            {
                "title": ("Test data set",),
                "description": ("An example data set to test core loading",),
            },
            (
                (INFO, "Loading core metadata"),
                (INFO, "Metadata for Core fields found: 1 records"),
            ),
            id="valid core",
        ),
    ],
)
def test_core(
    caplog,
    fixture_summary,
    rows,
    expected_log_entries,
):
    """Test that function to load in core block works as expected."""

    # Update valid to test error conditions and populate _rows
    # directly (bypassing .load() and need to pack in worksheet object
    fixture_summary._rows = rows

    # Test the block load
    fixture_summary._load_core()

    # check that relevant attributes have be populated correctly
    assert fixture_summary.title == "Test data set"
    assert fixture_summary.description == "An example data set to test core loading"

    log_check(caplog, expected_log_entries)
