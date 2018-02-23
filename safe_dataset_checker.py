#!/usr/bin/python
# -*- coding: UTF8 -*-

"""
Module containing code to verify the format of a SAFE project Excel dataset.

The functions are written to extract as much information as possible from
checking a file so, rather than raising an error and halting at the first
fault, functions are written to check as much as they can on the inputs.

Metadata is stored in a Dataset object, which also keeps a record of any
information and warnings issued during processing. These are optionally
printed to screen during processing but can also be retrieved from the
Dataset object.

Taxon names and locations are both validated:
i)  By default, taxonomic names are validated against the GBIF backbone taxonomy
    using API queries over the internet. A user can download and build a local
    sqlite copy of the database to avoid the need for a web connection and to
    improve the speed of validation.

ii) Locations are validated against a list of valid locations. By default, this
    is downloaded from a web service provided from the SAFE website, but a local
    file can be provided for off line use.
"""

from __future__ import print_function
import os
import datetime
import argparse
import re
from collections import Counter
from StringIO import StringIO
import numbers
import logging
import sqlite3
import openpyxl
from openpyxl import utils
import requests
import simplejson

# define some regular expressions used to check validity
RE_ORCID = re.compile(r'[0-9]{4}-[0-9]{4}-[0-9]{4}-[0-9]{4}')
RE_EMAIL = re.compile(r'\S+@\S+\.\S+')
RE_NAME = re.compile(r'[^,]+,[ ]?[^,]+')
RE_WSPACE_ONLY = re.compile(r'^\s*$')
# note that logically the next one also catches whitespace at both ends
RE_WSPACE_AT_ENDS = re.compile(r'^\s+\w+|\w+\s+$')
RE_DMS = re.compile(r'[°\'"dms’”]+')

# Taxon levels used in GBIF taxonomy. We explicitly exclude form and variety
# because they cannot be matched into the backbone hierarchy without extra
# API calls
BACKBONE_TYPES = ['kingdom', 'phylum', 'order', 'class', 'family',
                  'genus', 'species', 'subspecies']

# Logger setup - setup the standard logger to provide
# i)   A handler providing a counter of the number of calls to each log level
# ii)  A formatter to provide user controlled indentation and to
#      code levels as symbols for compactness
# iii) A handler writing to a global log file written to a StringIO container
# iv)  Optionally a handler also writing the log to stdout can be added when
#      a Dataset instance is created.

# Note that the handlers are created when the module is loaded, so when running
# behind a web server, the content of the handlers persist between runs of the code.
# To avoid constant concatenation of outputs, Dataset.__init__() empties them.


class CounterHandler(logging.Handler):
    """
    Handler instance that maintains a count of calls at each log level
    """
    def __init__(self, *args, **kwargs):
        logging.Handler.__init__(self, *args, **kwargs)
        self.counters = {'DEBUG': 0, 'INFO': 0, 'WARNING': 0, 'ERROR': 0, 'CRITICAL': 0}
    
    def emit(self, rec):
        self.counters[rec.levelname] += 1
    
    def reset(self):
        self.counters = {'DEBUG': 0, 'INFO': 0, 'WARNING': 0, 'ERROR': 0, 'CRITICAL': 0}


class IndentFormatter(logging.Formatter):
    """
    A formatter that provides an indentation depth and encodes the logging
    levels as single character strings. The extra argument to logger messages
    can be used to provide a dictionary to set:
        - 'join': a list of entries to join as comma separated list on
          to the end of the message.
        - 'quote': a flag to set joined entries to be quoted to show
          whitespace around values.
        - 'indent_before' or 'indent_after': two options to set the indent
          depth of the formatter.
    """

    def __init__(self, fmt=None, datefmt=None):
        logging.Formatter.__init__(self, fmt, datefmt)
        self.depth = 0

    def format(self, rec):

        # adjust indent before
        if hasattr(rec, 'indent_before'):
            self.depth = int(rec.indent_before)

        # store so users don't have to provide both
        rec.indent_before = self.depth

        rec.indent = '    ' * self.depth
        # encode level
        codes = {'DEBUG': '>', 'INFO': '-', 'WARNING': '?', 'ERROR': '!', 'CRITICAL': '*'}
        rec.levelcode = codes[rec.levelname]
        # format message
        msg = logging.Formatter.format(self, rec)
        # add any joined values
        if hasattr(rec, 'join'):
            if hasattr(rec, 'quote') and rec.quote:
                # quote if requested and then avoid requoting if the
                # formatter is emitting to more than one stream handler
                rec.join = ["'" + unicode(o) + "'" for o in rec.join]
                del rec.quote
            msg += ', '.join(rec.join)

        # adjust indent after
        if hasattr(rec, 'indent_after'):
            self.depth = int(rec.indent_after)

        # store so users don't have to provide both
        rec.indent_after = self.depth

        return msg


# Setup the logging instance
LOGGER = logging.getLogger(__name__)
LOGGER.setLevel(logging.DEBUG)

# Add the counter handler
CH = CounterHandler()
CH.setLevel(logging.DEBUG)
LOGGER.addHandler(CH)

# Create the formatter
FORMATTER = IndentFormatter("%(indent)s%(levelcode)s %(message)s")

# Create a StringIO object to hold the log, set a stream handler to use that,
# attach the custom formatter and add it to the logger. LOG will then contain
# a complete record of logging messages.
LOG = StringIO()
LOG_SH = logging.StreamHandler(LOG)
LOG_SH.setFormatter(FORMATTER)
LOGGER.addHandler(LOG_SH)

"""
Some static functions
"""


def is_blank(value):
    """
    Helper function that checks if a value is None or contains only whitespace
    Args:
        value: A single value

    Returns:
        Boolean
    """

    return (value is None) or (RE_WSPACE_ONLY.match(unicode(value)) is not None)


def is_padded(value):
    """
    Helper function that checks if the value is padded with whitespace
    Args:
        value: Contents of a cell

    Returns:
        Boolean
    """

    return (value is not None) and (RE_WSPACE_AT_ENDS.match(str(value)))


def duplication(data):
    """
    Looks for duplication in a list of strings, returning the duplicated elements.

    Args:
        data: A list of strings

    Returns:
        A list of duplicated elements
    """

    return [ky for ky, vl in Counter(data).iteritems() if vl > 1]


def is_integer_string(txt):
    """
    Checks if a string value can represent an integer.
    Args:
        txt: A string

    Returns:
        A boolean.
    """
    try:
        int(txt)
        return True
    except ValueError:
        return False


def all_numeric(data):
    """
    Tests if all values in an iterable are numeric.
    Args:
        data: An iterable claiming to contain numbers

    Returns:
        A list containing a boolean indicating whether all of the
        values were numbers and then a list of the genuine numeric values
    """

    nums = [vl for vl in data if isinstance(vl, numbers.Number)]

    return [len(nums) == len(data), nums]


def web_gbif_validate(tax, rnk, gbif_id=None, backbone_types=BACKBONE_TYPES):

    """
    Validates a taxon name and rank against the GBIF web API.

    Args:
        tax: The (case insensitive) taxon name
        rnk: The taxon rank
        gbif_id: Optional GBIF taxon id to resolve ambiguous taxon/rank
            combinations. If gbif_id is not None, then tax and rank can be
            None and the GBIF response will be used to populate those values.
        backbone_types: A list of higher taxon rank names to be added
            to the index along with the target taxon. This should be
            in rank order from root to tip.

    Returns:
        A dictionary with the following possible keys:
            'status': the outcome of the lookup with one of the following values:
                found, no_match, validation_fail, unknown_id, id_mismatch
            'user': an index tuple of the user provided information for synonymous usages
            'canon': an index tuple of the canonical information
            'hier': a list of 2-tuples of rank and GBIF ID for the taxonomic hierarchy
            'note': a string of any extra information provided by the search

        An index tuple has the structure: (taxon ID, parent ID, name, rank, taxonomic status)
    """

    def _hierarchy_walk(json, tax_levels):
        """
        Takes the JSON response from the species/match endpoint and converts it into a
        dictionary, keyed by taxonomic level, of GBIF UsageKeys for the available
        hierarchy in a response.

        It robustly handles non-standard parental assignments: the hierarchy doesn't
        necessarily always go up to the next most nested level. Analysis of the local
        database shows that both accepted and doubtful taxa do always go up, but this
        isn't true for synonyms.

        Note that this code will not work if the taxon is a form or variety, but those
        are not recognized by the SAFE dataset checker

        Args:
            json: The JSON response from the species/match endpoint
            tax_levels: An ordered list (kingdom > species) list of permitted taxonomic levels

        Returns:
            A list of 2 tuples (rank, GBIF ID).
        """

        # Add a subspecies entry into the response if needed
        if json['rank'].lower() == 'subspecies':
            json['subspecies'] = json['canonicalName']
            json['subspeciesKey'] = json['usageKey']

        # Which backbone levels are found in this response (preserves order)
        found = [bk for bk in tax_levels if bk in json]

        # build up a list of taxon hierarchy and GBIF keys
        ranks = []

        for each_rank in found:
            ranks.append((each_rank, json[each_rank + 'Key']))

        return ranks

    # Use the GBIF API to validate
    if gbif_id is None:
        # If no ID is provided then use the species/match?name={} endpoint
        url = u"http://api.gbif.org/v1/species/match?name={}&rank={}&strict=true".format(tax, rnk)
        tax_gbif = requests.get(url)

        # failures here suggest some kind of gateway problem
        if tax_gbif.status_code != 200:
            return {'status': 'validation_fail'}

        resp = tax_gbif.json()

    else:
        # Otherwise use the species/{id} endpoint
        url = u"http://api.gbif.org/v1/species/{}".format(gbif_id)
        tax_gbif = requests.get(url)

        # unknown ID numbers return a 404 error
        if tax_gbif.status_code == 404:
            return {'status': 'unknown_id'}
        elif tax_gbif.status_code != 200:
            return {'status': 'validation_fail'}

        resp = tax_gbif.json()

        # Handle usage
        if tax is None and rnk is None:
            # If neither taxon and rank are provided then populate from the response
            tax = resp['canonicalName']
            rnk = resp['rank'].lower()
        elif tax != resp['canonicalName'] or rnk.lower() != resp['rank'].lower():
            # Otherwise check they are congruent with the ID
            return {'status': 'id_mismatch'}

        # Two values have different keys in the species/match endpoint
        resp['status'] = resp['taxonomicStatus']
        resp['usageKey'] = resp['key']
        # Accepted usage also has a different key and is only present for synonym usages
        resp['acceptedUsageKey'] = resp.get('acceptedKey')
        # The page loaded, so we found the taxon!
        resp['matchType'] = 'EXACT'

    # get the status
    if 'status' in resp:
        tax_status = resp['status'].lower()
    else:
        tax_status = None

    # check the response status
    if resp['matchType'] == u'NONE':

        # No match found - look for explanatory notes
        if 'note' in resp:
            return {'status': 'no_match', 'note': resp['note']}
        else:
            return {'status': 'no_match'}

    elif tax_status in ('accepted', 'doubtful'):

        # get the hierarchy
        hier = _hierarchy_walk(resp, backbone_types)

        # build the canonical taxon index
        parent_index = None if len(hier) == 1 else hier[-2][1]
        canon = (resp['usageKey'], parent_index, tax, rnk, tax_status)

        # return the details
        return {'status': 'found', 'canon': canon, 'hier': hier}

    else:
        # A non-accepted match, so will have acceptedUsageKey in the species/match endpoint
        # data, which can be used to get the accepted usage and taxonomy
        usage_url = u"http://api.gbif.org/v1/species/{}".format(resp['acceptedUsageKey'])
        accept = requests.get(usage_url)
        if accept.status_code != 200:
            return {'status': 'validation_fail'}
        else:
            acpt = accept.json()
            hier = _hierarchy_walk(acpt, backbone_types)

            # build the canonical and user taxon index entries, storing the usageKey
            # for the user provided name and the backbone (nub) key for the accepted
            # usage.
            # TODO - currently using acpt['key'] as acpt['nubKey'] sometimes missing
            #        I don't know if this is correct, question opened on GBIF
            # https://github.com/gbif/portal-feedback/issues/1199
            parent_index = None if len(hier) == 1 else hier[-2][1]
            user = (resp['usageKey'], parent_index, tax, rnk, tax_status)
            canon = (acpt['key'],  parent_index, acpt['canonicalName'],
                     acpt['rank'].lower(), acpt['taxonomicStatus'].lower())

            # return the details
            return {'status': 'found', 'canon': canon, 'user': user, 'hier': hier}


def local_gbif_validate(conn, tax, rnk, gbif_id=None, backbone_types=BACKBONE_TYPES):

    """
    Validates a taxon name and rank against a connection to a local GBIF database.

    Args:
        conn: An sqlite3 connection to the backbone database
        tax: The (case sensitive) taxon name
        rnk: The taxon rank
        gbif_id: Optional GBIF taxon id to resolve ambiguous taxon/rank
            combinations. If gbif_id is not None, then tax and rank can be
            None and the GBIF response will be used to populate those values.
        backbone_types: A list of higher taxon rank names to be added
            to the index along with the target taxon. This should be
            in rank order from root to tip.

    Returns:
        A dictionary with the following possible keys:
            'status': the outcome of the lookup with one of the following values:
                found, no_match, validation_fail, unknown_id, id_mismatch
            'user': an index tuple of the user provided information for synonymous usages
            'canon': an index tuple of the canonical information
            'hier': a list of 2-tuples of rank and GBIF ID for the taxonomic hierarchy
            'note': a string of any extra information provided by the search

        An index tuple has the structure: (taxon ID, parent ID, name, rank, taxonomic status)
    """

    def _row_to_index_tuple(row):
        """
        Turns a row from the local GBIF database into a tuple for the taxon index
        Args:
            row: A row from the GBIF backbone database as a dictionary

        Returns:
            A 5 tuple: taxon ID, parent ID, name, rank, status
        """

        # handle kingdoms with no parents
        if row['parentNameUsageID'] == '':
            parent = None
        else:
            parent = int(row['parentNameUsageID'])

        return (int(row['taxonID']), parent, row['canonicalName'],
                row['taxonRank'], row['taxonomicStatus'])

    # Make sure the connection is returning results as sqlite.Row objects
    if conn.row_factory != sqlite3.Row:
        conn.row_factory = sqlite3.Row

    if gbif_id is not None:
        # get the record associated with the provided ID
        tax_sql = ("select * from backbone where taxonID={}".format(gbif_id))
        tax_gbif = conn.execute(tax_sql).fetchone()

        # check there is a result and that it is congruent with any
        # provided taxon or rank information
        if tax_gbif is None:
            return {'status': 'unknown_id'}
        elif ((tax is not None and tax_gbif['canonicalName'] != tax) or
              (rnk is not None and tax_gbif['taxonRank'] != rnk.lower())):
            return {'status': 'id_mismatch'}

        #  store the names of the backbone hierarchy in the row
        hier = [(hi, tax_gbif[hi]) for hi in backbone_types
                if hi in tax_gbif.keys() and tax_gbif[hi] != u'']

        if tax_gbif['taxonomicStatus'] in ['accepted', 'doubtful']:
            # Accepted taxon
            return {'status': 'found', 'canon': _row_to_index_tuple(tax_gbif), 'hier': hier}
        else:
            # synonym or misapplied, so find the accepted usage
            acc_id = tax_gbif['acceptedNameUsageID']
            acc_sql = 'select * from backbone where taxonID = {};'.format(acc_id)
            acc_gbif = conn.execute(acc_sql).fetchone()
            return {'status': 'found', 'canon': _row_to_index_tuple(acc_gbif),
                    'user': _row_to_index_tuple(tax_gbif), 'hier': hier}
    else:
        # get the set of records associated with the taxon or rank
        tax_sql = ("select * from backbone where canonicalName='{}' and "
                   "taxonRank='{}';".format(tax, rnk.lower()))
        tax_gbif = conn.execute(tax_sql).fetchall()

        if len(tax_gbif) == 0:
            # No matching rows
            return {'status': 'no_match'}
        elif len(tax_gbif) == 1:
            # one matching row - extract it from the list
            tax_gbif = tax_gbif[0]
        else:
            # More than one row - try to mimic the preferred hits reported
            # by the GBIF API to select a single hit by looking at the counts
            # of the different statuses.

            # First, get the taxon statuses
            tx_status = [tx['taxonomicStatus'] for tx in tax_gbif]
            tx_counts = Counter(tx_status)

            if 'accepted' in tx_counts.keys():
                # Accepted hits are first preference
                if tx_counts['accepted'] == 1:
                    # only one accepted match alongside other usages so extract that hit
                    tax_gbif = tax_gbif[tx_status.index('accepted')]
                else:
                    # more than one accepted match, so return no match and a note, as
                    # the API interface does.
                    return {'status': 'no_match',
                            'note': 'Multiple equal matches for {}'.format(tax)}

            elif 'doubtful' in tx_counts.keys():
                # Doubtful hits get next preference - not quite sure about this!
                if tx_counts['doubtful'] == 1:
                    # only one doubtful match alongside other usages so extract that hit
                    tax_gbif = tax_gbif[tx_status.index('doubtful')]
                else:
                    # more than one doubtful match, so return no match and a note, as
                    # the API interface does.
                    return {'status': 'no_match',
                            'note': 'Multiple equal matches for {}'.format(tax)}

            else:
                # Rows can now contain synonyms (of varying kinds) and misapplied. Both of
                # these types have accepted usage values, so look for a unique accepted usage.
                tx_acc = {tx['acceptedNameUsageID'] for tx in tax_gbif
                          if tx['acceptedNameUsageID'] != u''}

                if len(tx_acc) > 1:
                    # More than one accepted usage
                    return {'status': 'no_match',
                            'note': 'Multiple equal matches for {}'.format(tax)}
                else:
                    # A single accepted usage - pick the first row to index
                    tax_gbif = tax_gbif[0]

        # Should now have a single row for the preferred hit
        # i) store the names of the backbone hierarchy in the row
        hier = [(hi, tax_gbif[hi]) for hi in backbone_types
                if hi in tax_gbif.keys() and tax_gbif[hi] != u'']

        if tax_gbif['taxonomicStatus'] in ['accepted', 'doubtful']:
            return {'status': 'found',  'canon': _row_to_index_tuple(tax_gbif), 'hier': hier}
        else:
            # synonym or misapplied, so find the accepted usage
            acc_id = tax_gbif['acceptedNameUsageID']
            acc_sql = 'select * from backbone where taxonID = {};'.format(acc_id)
            acc_gbif = conn.execute(acc_sql).fetchone()
            return {'status': 'found', 'canon': _row_to_index_tuple(acc_gbif),
                    'user': _row_to_index_tuple(tax_gbif), 'hier': hier}


"""
Two classes to handle datasets and the data worksheets they contain
"""


class DataWorksheet(object):

    """
    This is just a container for the metadata on a data worksheet.
    It basically only exists to set defaults on a data worksheet and
    to provide dot notation rather than using dictionary labels.

    Args:
        meta: The dictionary of worksheet name, title and description
        loaded from the summary worksheet.
    """

    def __init__(self, meta):

        # set defaults
        self.name = meta['name']
        self.description = meta['description']
        self.title = meta['title']
        self.max_row = 0
        self.max_col = 0
        self.descriptors = None
        self.taxa_fields = None
        self.field_name_row = None
        self.fields = []


class Dataset(object):

    """
    This class provides methods to load and store the metadata associated
    with a dataset stored in a SAFE project formatted Excel file.

    Args:
        filename: The Excel file from which to read the Dataset
        verbose: Should the reporting print to the screen during runtime
        gbif_database: A local path to an sqlite database containing the GBIF database
        to be used instead of the web API.
    """

    def __init__(self, filename, verbose=True, gbif_database=None):

        try:
            self.workbook = openpyxl.load_workbook(filename=filename, data_only=True,
                                                   read_only=True)
        except IOError:
            raise IOError('Could not open file {}'.format(filename))

        self.filename = os.path.basename(filename)

        # report tracking - clear out previous content in the logger. When the function runs
        # on a webserver, the same logger instance is used for all runs, so the log contents
        # persist unless they are tidied up.
        CH.reset()
        LOG.truncate(0)

        if verbose:
            # Add a stream handler writing to stdout, using the common formatter instance,
            # as well as the logging to the global LOG StringIO, otherwise command line 
            # usage doesn't see any output.
            stdout = logging.StreamHandler()
            stdout.setFormatter(FORMATTER)
            LOGGER.addHandler(stdout)

        LOGGER.info("Checking file '{}'".format(filename),
                    extra={'indent_before': 0, 'indent_after': 1})

        # basic info
        self.sheet_names = set(self.workbook.get_sheet_names())

        # initialise data contents
        self.project_id = None
        self.authors = []
        self.title = None
        self.description = None
        self.dataworksheet_summaries = []
        self.access = 'Open'
        self.embargo_date = None
        self.publication_doi = []
        self.dataworksheets = []
        self.keywords = []
        self.temporal_extent = None
        self.latitudinal_extent = None
        self.longitudinal_extent = None
        self.locations = set()
        self.locations_used = set()
        self.taxon_names = set()
        self.taxon_names_used = set()
        self.taxon_index = set()
        self.passed = False
        
        # Setup the taxonomy validation mechanism.
        if gbif_database is None:
            LOGGER.info('Using GBIF online API to validate taxonomy')
            self.use_local_gbif = False
            self.gbif_conn = None
        else:
            # If a GBIF file is provided, does the file exist?
            if not os.path.exists(gbif_database):
                LOGGER.info('Local GBIF database not found, defaulting to online API')
                self.use_local_gbif = False
                self.gbif_conn = None

            # And does the file behave like i) a sqlite database ii) containing a table 'backbone'
            try:
                # can connect sqlite to any existing path, but only attempts to query it
                # throw up exceptions
                conn = sqlite3.connect(gbif_database)
                _ = conn.execute('select count(*) from backbone;')
            except sqlite3.OperationalError:
                LOGGER.info('Local GBIF database does not contain the backbone table, defaulting '
                            'to online API')
                self.use_local_gbif = False
                self.gbif_conn = None
            except sqlite3.DatabaseError:
                LOGGER.info('Local SQLite database not valid, defaulting to online API')
                self.use_local_gbif = False
                self.gbif_conn = None
            else:
                LOGGER.info('Using local GBIF database to validate taxonomy')
                self.use_local_gbif = True
                self.gbif_conn = conn
                self.gbif_conn.row_factory = sqlite3.Row

    @staticmethod
    def report():
        """
        Static method to return the logging report stored in the global LOG StringIO
        Returns:
            A StringIO containing the logging report.
        """
        return LOG

    # Utility method to update extents
    def update_extent(self, extent, val_type, which):
        """
        Updates one of the three dataset extents
        Args:
            extent: A two tuple (min, max) for one of the extents
            val_type: The type that the values in the extent should be
            which: Which extent to update
        """

        if (not isinstance(extent, tuple) or
                len(extent) != 2 or
                not all([isinstance(vl, val_type) for vl in extent]) or
                extent[0] > extent[1]):
            raise ValueError('extent must be an increasing two-tuple of values of '
                             'type {}'.format(val_type))

        current_vals = getattr(self, which)
        if current_vals is None:
            setattr(self, which, extent)
        else:
            extent = zip(current_vals, extent)
            setattr(self, which, (min(extent[0]), max(extent[1])))

    # Main methods to populate class attributes
    def load_summary(self, validate_doi=False):

        """
        Checks the information in the summary worksheet and looks for the metadata and
        dataset worksheets. The function is intended to try and get as much information
        as possible from the worksheet: the dictionary of metadata returned will have
        None for any missing data, which should be handled by downstream code.

        Args:
            validate_doi: Check any publication DOIs, requiring a web connection.
        """

        # try and get the summary worksheet
        LOGGER.info("Checking Summary worksheet", extra={'indent_f': 0, 'indent_l': 'r'})
        start_errors = CH.counters['ERROR']

        try:
            worksheet = self.workbook['Summary']
        except KeyError:
            LOGGER.error("Summary worksheet not found, moving on.")
            return None

        # load dictionary of summary information block, allowing for multiple
        # columns for fields (data compilers, dataset sheets).
        # Note that worksheet.columns can load some odd invisible empty cells:
        # it doesn't guarantee an equal number of cells in each column. So
        # use the squared range.
        mxcl = worksheet.max_column
        mxrw = worksheet.max_row

        # populate block of rows of cell objects
        data = list(worksheet.get_squared_range(1, 1, mxcl, mxrw))

        # transpose and get values
        data = zip(*data)
        cols = tuple(tuple(cell.value for cell in col) for col in data)

        # convert into dictionary
        hdrs = cols[0]
        values = zip(*cols[1:])
        summary = {ky: vl for ky, vl in zip(hdrs, values)}

        # Check the minimal keys are expected
        required = {"SAFE Project ID", "Access status", "Title", "Description",
                    "Author name", "Author email", "Author affiliation", "Author ORCID",
                    "Worksheet name", "Worksheet title", "Worksheet description", 'Keywords'}

        found = set(summary.keys())

        # don't bail here - try and get as far as possible
        if not found.issuperset(required):
            LOGGER.warn('Missing metadata fields: ', extra={'join': required - found})

        # # check for contents
        # if any([set(x) == {None} for x in summary.values()]):
        #     LOGGER.error('Metadata fields with no information.', 1)

        # CHECK PROJECT ID
        if 'SAFE Project ID' not in summary:
            LOGGER.warn('SAFE Project ID missing')
        elif not isinstance(summary['SAFE Project ID'][0], long):
            LOGGER.warn('SAFE Project ID is not an integer')
        else:
            self.project_id = summary['SAFE Project ID'][0]

        # CHECK DATASET TITLE
        if 'Title' not in summary:
            LOGGER.error('Dataset title row missing')
        elif is_blank(summary['Title'][0]):
            LOGGER.error('Dataset title is blank')
        else:
            self.title = summary['Title'][0]

        # CHECK DATASET DESCRIPTION
        if 'Description' not in summary:
            LOGGER.error('Dataset description row missing')
        elif is_blank(summary['Description'][0]):
            LOGGER.error('Dataset description is blank')
        else:
            self.description = summary['Description'][0]

        # CHECK ACCESS STATUS AND EMBARGO DETAILS
        if 'Access status' not in summary:
            LOGGER.error('Access status missing')
        elif summary['Access status'][0] not in ['Open', 'Embargo']:
            LOGGER.error('Access status must be Open or Embargo '
                         'not {}'.format(summary['Access status'][0]))
        elif summary['Access status'] == 'Embargo':
            self.access = 'Embargo'
            if 'Embargo date' not in summary:
                LOGGER.error('Dataset embargoed but embargo date row missing.')
            elif is_blank(summary['Embargo'][0]):
                LOGGER.error('Dataset embargo date  is blank')
            else:
                embargo_date = summary['Embargo date'][0]
                now = datetime.datetime.now()
                if not isinstance(embargo_date, datetime.datetime):
                    LOGGER.error('Embargo date not formatted as date.')
                elif embargo_date < now:
                    LOGGER.error('Embargo date is in the past.')
                elif embargo_date > now + datetime.timedelta(days=2 * 365):
                    LOGGER.error('Embargo date more than two years in the future.')
                else:
                    self.embargo_date = embargo_date.date().isoformat()

        # CHECK KEYWORDS
        if 'Keywords' not in summary:
            LOGGER.error('Dataset keywords row missing')
        elif all([is_blank(kywd) for kywd in summary['Keywords']]):
            LOGGER.error('No keywords provided')
        else:
            # drop any blanks
            self.keywords = [vl for vl in summary['Keywords'] if not is_blank(vl)]
            if len(self.keywords) == 1 and ',' in self.keywords[0]:
                LOGGER.error('Put keywords in separate cells, not comma delimited in one cell')

        # CHECK FOR PUBLICATION DOIs if any are provided
        if 'Publication DOI' in summary:
            pub_doi = [vl for vl in summary['Publication DOI'] if not is_blank(vl)]
            # check formatting - basically make sure they have a proxy URL
            doi_is_url = [vl.startswith('https://doi.org/') for vl in pub_doi]
            if not all(doi_is_url):
                LOGGER.error('Please provide publication DOIs as a URL: https://doi.org/...')

            if validate_doi:
                for doi, is_doi in zip(pub_doi, doi_is_url):
                    if is_doi:
                        api_call = 'https://doi.org/api/handles/{}'.format(doi[16:])
                        r = requests.get(api_call)
                        if r.json()['responseCode'] != 1:
                            LOGGER.error('DOI not found: {}'.format(doi))

            self.publication_doi = pub_doi

        # CHECK AUTHORS
        # Get the set of author fields and create blank entries for any missing fields
        # to simplify handling the error checking.
        author_keys = ['Author name', 'Author affiliation', 'Author email', 'Author ORCID']
        authors = {k: summary[k] if k in summary else [None] * (len(cols) - 1)
                   for k in author_keys}

        # - switch in zenodo keys
        zenodo_keys = ['name', 'affiliation', 'email', 'orcid']
        for old, new in zip(author_keys, zenodo_keys):
            authors[new] = authors[old]
            authors.pop(old)

        # rotate and remove completely blank columns
        authors_list = zip(*authors.values())
        authors_list = [x for x in authors_list if x != tuple([None] * 4)]
        # return to the original orientation
        authors = {k: v for k, v in zip(authors, zip(*authors_list))}

        # now check the contents
        # i) Names
        author_names = [unicode(vl) for vl in authors['name'] if vl is not None]
        if len(author_names) < len(authors['name']):
            LOGGER.error('Missing author names')
        if author_names:
            bad_names = [vl for vl in author_names if not RE_NAME.match(vl)]
            if bad_names:
                LOGGER.error('Author name not formatted as last_name, first_names: ',
                             extra={'join': bad_names, 'quote': True})

        # ii) Affiliations (no regex checking)
        blank_affil = [is_blank(vl) for vl in authors['affiliation']]
        if any(blank_affil):
            LOGGER.warn('Missing affiliations - please provide if available')

        # iii) Email
        author_emails = [unicode(vl) for vl in authors['email'] if vl is not None]
        if len(author_emails) < len(authors['email']):
            LOGGER.warn('Missing author emails - please provide if available')
        if author_emails:
            bad_emails = [vl for vl in author_emails if not RE_EMAIL.match(vl)]
            if bad_emails:
                LOGGER.error('Email not properly formatted: ',
                             extra={'join': bad_emails, 'quote': True})

        # iii) ORCiD (not mandatory)
        author_orcid = [unicode(vl) for vl in authors['orcid'] if vl is not None]
        if len(author_orcid) < len(authors['orcid']):
            LOGGER.warn('Missing ORCiDs, consider adding them!')
        if author_orcid:
            bad_orcid = [vl for vl in author_orcid if not RE_ORCID.match(vl)]
            if bad_orcid:
                LOGGER.error('ORCID not properly formatted: ',
                             extra={'join': bad_orcid, 'quote': True})

        # and finally store as a dictionary per author
        self.authors = [dict(zip(authors.keys(), vals)) for vals in zip(*authors.values())]

        # CHECK DATA WORKSHEETS
        ws_keys = ['Worksheet name', 'Worksheet title', 'Worksheet description']
        data_worksheets = {k: summary[k] if k in summary else [None] * (len(cols) - 1)
                           for k in ws_keys}

        # - switch in short keys
        short_ws_keys = ['name', 'title', 'description']
        for old, new in zip(ws_keys, short_ws_keys):
            data_worksheets[new] = data_worksheets[old]
            data_worksheets.pop(old)

        # rotate and remove completely blank columns
        data_worksheets_list = zip(*data_worksheets.values())
        data_worksheets_list = [x for x in data_worksheets_list if x != tuple([None] * 3)]
        # return to the original orientation
        data_worksheets = {k: v for k, v in zip(data_worksheets, zip(*data_worksheets_list))}

        # now check the contents
        # i) Names
        ws_names = [vl for vl in data_worksheets['name'] if vl is not None]
        if len(ws_names) < len(data_worksheets['name']):
            LOGGER.error('Missing worksheet names')
        if ws_names:
            bad_names = [vl for vl in ws_names if vl not in self.sheet_names]
            if bad_names:
                LOGGER.error('Worksheet names not found in workbook: ',
                             extra={'join': bad_names, 'quote': True})

        # ii) Titles
        blank_names = [is_blank(vl) for vl in data_worksheets['title']]
        if any(blank_names):
            LOGGER.error('Missing worksheet title')

        # ii) Descriptions
        blank_desc = [is_blank(vl) for vl in data_worksheets['description']]
        if any(blank_desc):
            LOGGER.error('Missing worksheet description')

        # and finally store a list of dictionaries of data worksheet summary details
        self.dataworksheet_summaries = [dict(zip(data_worksheets.keys(), vals))
                                        for vals in zip(*data_worksheets.values())]

        # check for extra undocumented spreadsheets
        if 'Worksheet name' in summary:
            expected_sheets = set(data_worksheets['name']) | {'Summary', 'Taxa', 'Locations'}
            if not self.sheet_names.issubset(expected_sheets):
                LOGGER.error('Undocumented sheets found in  workbook: ',
                             extra={'join': self.sheet_names - expected_sheets, 'quote': True})

        # summary of processing
        n_errors = CH.counters['ERROR'] - start_errors
        if n_errors > 0:
            LOGGER.info('Summary contains {} errors'.format(n_errors))
        else:
            LOGGER.info('Summary formatted correctly')

    def load_locations(self, locations_json=None):

        """
        Attempts to load and check the contents of the Locations worksheet and
        compile the geographic extent of the locations used.

        Args:
            locations_json: A path to a JSON file containing a valid set of location names.
                With the default value of None, the function tries to get this from
                a SAFE project website service.
        Returns:
            Populates the locations field of the Dataset object
        """

        # try and get the locations worksheet
        LOGGER.info("Checking Locations worksheet",
                    extra={'indent_before': 0, 'indent_after': 1})
        start_errors = CH.counters['ERROR']

        try:
            locs_wb = self.workbook['Locations']
        except KeyError:
            # No locations is pretty implausible, but still persevere as if
            # they aren't going to be required
            LOGGER.warn("No locations worksheet found - moving on")
            return

        # GET THE GAZETTEER VALIDATION INFORMATION
        loc_payload = None
        if locations_json is None:
            # If no file is provided then try and get locations from the website service
            loc_get = requests.get('https://www.safeproject.net/call/json/get_locations_bbox')
            if loc_get.status_code != 200:
                LOGGER.error('Could not download locations. Use a local json file.')
            else:
                loc_payload = loc_get.json()
        else:
            # try and load the file
            try:
                loc_payload = simplejson.load(file(locations_json))
            except IOError:
                LOGGER.error('Could not load location names from file.')

        # process the payload
        if loc_payload is not None:
            valid_locations = loc_payload['locations']
            aliases = loc_payload['aliases']
        else:
            # create empty dictionaries to continue with validation
            valid_locations = {}
            aliases = {}

        # ITERATE OVER THE WORKSHEET ROWS
        loc_rows = locs_wb.rows

        # Get the field headers:
        # Duplicated headers are a problem because the values in the locations dictionaries get
        # overwritten. Depending on what gets overwritten, this can produce really unpredictable
        # bugs, so just stop here.
        hdrs = [cl.value for cl in loc_rows.next()]
        dupes = duplication([h for h in hdrs if not is_blank(h)])
        if dupes:
            LOGGER.error('Duplicated location sheet headers: ',
                         extra={'join': dupes})
            return

        # Check location names are available
        if 'Location name' not in hdrs:
            LOGGER.error('Location name column not found')
            return

        # Convert remaining rows into a list of location dictionaries
        locs = [{ky: cl.value for ky, cl in zip(hdrs, rw)} for rw in loc_rows]

        # strip out any rows that consist of nothing but empty cells
        locs = [row for row in locs if not all([is_blank(vl) for vl in row.values()])]

        # Location name cleaning - get names as strings
        for rw in locs:
            rw['Location name'] = unicode(rw['Location name'])

        # check for rogue whitespace
        ws_padded = [rw['Location name'] for rw in locs if is_padded(rw['Location name'])]
        if ws_padded:
            LOGGER.error('Locations names with whitespace padding: ',
                         extra={'join': ws_padded, 'quote': True})
            # clean whitespace padding
            for row in locs:
                row['Location name'] = row['Location name'].strip()

        # look for duplicates
        dupes = duplication([rw['Location name'] for rw in locs])
        if dupes:
            LOGGER.error('Duplicated location names: ',
                         extra={'join': dupes, 'quote': True})

        # VALIDATE LOCATIONS
        # Get existing location names and aliases -new location names must not appear
        # in it and existing locations must.
        existing_loc_names = set(valid_locations.keys() + aliases.keys())

        # Split up old and new if there are there any new ones?
        if 'New' in hdrs:

            # Check the New column is just yes, no
            is_new = {rw['New'] for rw in locs}
            valid_new = {'yes', 'no', 'Yes', 'No'}
            if not is_new.issubset(valid_new):
                LOGGER.error('New field contains values other than Yes and No: ',
                             extra={'join': is_new - valid_new, 'quote': True})

            # extract the new and old locations
            new_locs = [rw for rw in locs if rw['New'].lower() == 'yes']
            locs = [rw for rw in locs if rw['New'].lower() == 'no']
        else:
            new_locs = None

        # Process new locations if there are any
        if new_locs:
            LOGGER.info('{} new locations reported'.format(len(new_locs)))

            # check Lat Long and Type, which automatically updates the extents.
            # Unlike a data worksheet field, here we don't have any metadata or want
            # to keep it, so field checker gets passed an empty dictionary, which is discarded.
            if 'Latitude' in hdrs:
                lats = [vl['Latitude'] for vl in new_locs if vl['Latitude'] != u'NA']
                non_blank_lats = [vl for vl in lats if not is_blank(vl)]
                if len(non_blank_lats) < len(lats):
                    LOGGER.error('Blank latitude values for new locations: use NA.')
                self.check_field_geo({}, non_blank_lats, which='latitude')
            else:
                LOGGER.error('New locations reported but Latitude field missing')

            if 'Longitude' in hdrs:
                longs = [vl['Longitude'] for vl in new_locs if vl['Longitude'] != u'NA']
                non_blank_longs = [vl for vl in longs if not is_blank(vl)]
                if len(non_blank_longs) < len(longs):
                    LOGGER.error('Blank longitude values for new locations: use NA.')
                self.check_field_geo({}, non_blank_longs, which='longitude')
            else:
                LOGGER.error('New locations reported but Longitude field missing')

            if 'Type' in hdrs:
                geo_types = {vl['Type'] for vl in new_locs}
                bad_geo_types = geo_types - {'POINT', 'LINESTRING', 'POLYGON'}
                if bad_geo_types:
                    LOGGER.error('Unknown location types: ',
                                 extra={'join': bad_geo_types, 'quote': True})
            else:
                LOGGER.error('New locations reported but Type field missing')

            duplicates_existing = [rw['Location name'] for rw in new_locs
                                   if rw['Location name'] in existing_loc_names]

            if duplicates_existing:
                LOGGER.error('New location names duplicate existing names and aliases: ',
                             extra={'join': duplicates_existing})

            # new location names
            new_loc_names = {rw['Location name'] for rw in new_locs}
        else:
            new_loc_names = set()

        # Process existing locations if there are any
        if locs:
            LOGGER.info('{} existing locations reported'.format(len(locs)))

            # check names exist
            loc_names = {rw['Location name'] for rw in locs}
            unknown = loc_names - existing_loc_names
            if unknown:
                LOGGER.error('Unknown locations found: ',
                             extra={'join': unknown, 'quote': True})

            # are aliases being used?
            aliased = loc_names & set(aliases.keys())
            if aliased:
                LOGGER.warn('Locations aliases used. Maybe change to primary location names: ',
                            extra={'join': aliased})

            # Get the bounding box of known locations and aliased locations
            bbox_keys = (loc_names - (unknown | aliased)) | {aliases[ky] for ky in aliased}

            # get the extents of known unaliased locations
            if bbox_keys:
                bbox = [vl for ky, vl in valid_locations.iteritems() if ky in bbox_keys]
                bbox = zip(*bbox)
                self.update_extent((min(bbox[0]), max(bbox[1])), float, 'longitudinal_extent')
                self.update_extent((min(bbox[2]), max(bbox[3])), float, 'latitudinal_extent')

        else:
            loc_names = set()

        # combine locations into set
        self.locations = loc_names | new_loc_names

        # summary of processing
        n_errors = CH.counters['ERROR'] - start_errors
        if n_errors > 0:
            LOGGER.info('Locations contains {} errors'.format(n_errors))
        else:
            LOGGER.info('{} locations loaded correctly'.format(len(self.locations)))

    def load_taxa(self):

        """
        Attempts to load and check the content of the Taxa worksheet. The
        method checks that all taxa have a local name and then validates
        taxon names and parent names against the GBIF backbone. It populates
        two things:

            i)  the taxon_names attribute of the dataset, which is just a set of
                names used as a validation list for taxon names used in data worksheets.
            ii) the taxon_index attribute of the dataset, which contains a set
                of tuples recording the full hierarchy of the taxa in the dataset
                for use in dataset searching. Each tuple consists of:

                (gbif_id, gbif_parent_id, canonical_name, taxonomic_rank, status, as_name, as_rank)

            The taxon_index is doing dual-duty here:

            a) It is used to populate the taxonomic coverage section of the dataset
               description. In this case, synonymous entries are dropped and the as_name
               and as_rank are used to qualify the accepted name.
            b) It is used as the rows for a dataset_taxon table to index the taxonomic
               coverage of datasets. The as_name and as_rank are not intended to be
               included in this table.

        Returns:
            Updates the taxon_names and taxon_index attributes of the class instance.
        """

        # try and get the taxon worksheet
        LOGGER.info("Checking Taxa worksheet",
                    extra={'indent_before': 0, 'indent_after': 1})
        start_errors = CH.counters['ERROR']

        try:
            sheet = self.workbook['Taxa']
        except KeyError:
            # This might mean that the study doesn't have any taxa, so return an empty
            # set. If the datasets then contain taxonomic names, it'll fail gracefully.
            LOGGER.warn("No taxa worksheet found - assuming no taxa in data for now!")
            return

        # A) SANITIZE INPUTS
        # get and check the headers
        tx_rows = sheet.rows
        hdrs = [cl.value for cl in tx_rows.next()]

        # duplicated headers are a problem in that it will cause values in the taxon
        # dictionaries to be overwritten. Depending on what gets overwritten, this can
        # produce really unpredictable bugs, so just stop here.
        dupes = duplication([h for h in hdrs if not is_blank(h)])
        if dupes:
            LOGGER.error('Duplicated column headers in Taxa worksheet: ',
                         extra={'join': dupes})
            return

        # Load dictionaries of the taxa and strip out any rows that consist of nothing
        # but empty cells
        taxa = [{ky: cl.value for ky, cl in zip(hdrs, rw)} for rw in tx_rows]
        taxa = [row for row in taxa if not all([is_blank(vl) for vl in row.values()])]

        # check number of taxa found and standardise names
        if len(taxa) == 0:
            LOGGER.info('No taxon rows found'.format(len(taxa)))
            return
        else:
            # Remove values keyed to None (blank columns),
            # convert keys to lower case and update the list of headers
            _ = [tx.pop(None) for tx in taxa if None in tx]
            taxa = [{ky.lower(): vl for ky, vl in tx.iteritems()} for tx in taxa]
            hdrs = taxa[0].keys()

        # clean the contents of whitespace padding
        for idx, tx in enumerate(taxa):
            for ky, vl in tx.iteritems():
                if is_padded(vl):
                    LOGGER.error('Whitespace padding for {} in row {}:'
                                 ' {}'.format(ky, idx + 2, repr(vl)))
                    tx[ky] = vl.strip()

        # check which fields are found
        if 'name' not in hdrs:
            # dataset names are not found so can't continue
            LOGGER.error('No name column found - no further checking')
            self.taxon_names = set()
            return
        else:
            # Validate the names used within the data
            taxon_names = [(tx['name'], is_blank(tx['name'])) for tx in taxa]

            # Any missing names?
            if any(zip(*taxon_names)[1]):
                LOGGER.error('Blank entries in name column')

            taxon_names = [tx[0] for tx in taxon_names if not tx[1]]

            # Any duplication in cleaned names
            dupes = duplication(taxon_names)
            if dupes:
                LOGGER.error('Duplicated names found: ', extra={'join': dupes})

            # set the unique taxon names for the dataset instance
            self.taxon_names = set(taxon_names)

            # check to see if the headers are present to validate against GBIF
            if 'taxon name' not in hdrs or 'taxon type' not in hdrs:
                LOGGER.error('At least the taxon type and taxon name columns are required'
                             'for taxon verification')
                return

            # Standardize the taxon representation into a three tuple:
            # (name, (taxon name, taxon type, taxon id), (parent name, parent type, parent id))
            # The dict.get() method automatically fills in None for missing keys.
            taxa = [(tx['name'],
                     (tx.get('taxon name'), tx.get('taxon type'), tx.get('taxon id')),
                     (tx.get('parent name'), tx.get('parent type'), tx.get('parent id')))
                    for tx in taxa]

        # B) VALIDATE TAXONOMY
        # Initialise a set to record failed validation and the set of taxonomic levels
        # that are available to validate in the backbone. See README.md for details of
        # the validation process

        # taxon and parent types that can be checked against gbif and a set of
        # alternative types that will not have a taxon name
        backbone_types = BACKBONE_TYPES
        alt_types = ['morphospecies', 'functional group']

        # Keep a set of taxonomic hierarchy entries to validate after taxon processing
        taxon_hierarchy = set()

        # Check parents first, populating a dictionary keyed on unique parent tuple to
        # record the status of checked parents along with parent gbif information if valid
        parents = set([tx[2] for tx in taxa if tx[2] != (None, None, None)])
        parent_status = dict()

        if len(parents):
            LOGGER.info('Checking {} parent taxa'.format(len(parents)),
                        extra={'indent_after': 2})

        for prnt in parents:

            # Sanitize inputs. Only two patterns of taxon provision are valid:
            #   name + type + id and name + type
            provided = [not is_blank(p) for p in prnt]
            prnt_string = ', '.join([str(p) for p in prnt if p is not None])

            # initialise the status with invalid and no information
            parent_status[prnt] = ('invalid', None)

            if provided not in [[True, True, False], [True, True, True]]:
                LOGGER.error('{}: incomplete information'.format(prnt_string))
                continue

            # Standard taxon type?
            if prnt[1].lower() not in backbone_types:
                LOGGER.error('{}: type is not recognized'.format(prnt_string))
                continue

            # Now validate against GBIF
            if self.gbif_conn is None:
                prnt_info = web_gbif_validate(prnt[0], prnt[1], prnt[2])
            else:
                prnt_info = local_gbif_validate(self.gbif_conn, prnt[0], prnt[1])

            # handle lookup errors
            if prnt_info['status'] == 'validation_fail':
                LOGGER.error('{}: validation failed'.format(prnt_string))
            elif prnt_info['status'] == 'unknown_id':
                LOGGER.error('{}: GBIF ID not known'.format(prnt_string))
            elif prnt_info['status'] == 'id_mismatch':
                LOGGER.error('{}: ID does not match name and rank'.format(prnt_string))
            elif prnt_info['status'] == 'no_match':
                if 'note' in prnt_info:
                    LOGGER.error('{}: {}'.format(prnt_string, prnt_info['note']))
                else:
                    LOGGER.error('{}: name and rank combination not found'.format(prnt_string))
            else:
                LOGGER.info('{}: taxon found'.format(prnt_string))
                parent_status[prnt] = ('valid', prnt_info)

        # Now check main taxa
        LOGGER.info('Validating {} taxa'.format(len(taxa)),
                    extra={'indent_before': 1, 'indent_after': 2})

        for idx, taxon in enumerate(taxa):

            # row index
            rw_num = idx + 2

            # Sanitize inputs. Only two patterns of taxon provision are valid:
            #   name + type + id and name + type
            provided = [not is_blank(tx) for tx in taxon[1]]

            if provided not in [[True, True, False], [True, True, True]]:
                LOGGER.error('Row {} ({}): incomplete information'.format(rw_num, taxon[0]))
                continue

            # The taxon input is now sanitized of obvious problems so can be validated against
            # the GBIF backbone database. Validate _all_ backbone type entries: this is to make
            # sure that valid taxa that have had parents provided get indexed correctly. It is
            # only an issue when taxa are not found and also have an invalid parent

            if taxon[1][1].lower() not in backbone_types:
                gbif_info = {'status': 'non-backbone'}
            elif self.gbif_conn is None:
                gbif_info = web_gbif_validate(taxon[1][0], taxon[1][1],
                                              taxon[1][2], backbone_types)
            else:
                gbif_info = local_gbif_validate(self.gbif_conn,
                                                taxon[1][0], taxon[1][1],
                                                taxon[1][2], backbone_types)

            # Handle immediately problematic lookup errors, leaving found and no_match
            # for cross checking against parent information
            if gbif_info['status'] == 'validation_fail':
                LOGGER.error('Row {} ({}): validation failed'.format(rw_num, taxon[0]))
                continue
            elif gbif_info['status'] == 'unknown_id':
                LOGGER.error('Row {} ({}): GBIF ID not known'.format(rw_num, taxon[0]))
                continue
            elif gbif_info['status'] == 'id_mismatch':
                LOGGER.error('Row {} ({}): ID does not match name and '
                             'rank'.format(rw_num, taxon[0]))
                continue
            else:
                tax_status = gbif_info['status']

            # Lookup the parent status and information
            if taxon[2] == (None, None, None):
                par_status, parent_info = (None, None)
            else:
                par_status, parent_info = parent_status[taxon[2]]

            # Check the combinations of taxon and parent status. At this point
            # parents are one of found, no_match or None and taxa are one of
            # non-backbone, found or no_match.

            # The combinations are shown below. The taxon row is valid (O) for: a found
            # taxon (with or without a valid parent); a non-matching taxon with a valid
            # parent; and a non-backbone taxon type with a valid parent.

            # Everything else is invalid (X), including a found taxon with a valid parent
            # that isn't actually a parent of the child taxon - hence the question mark
            # in the table.

            #         | None  | pr_inv | pr_val |
            # tx_fnd  |  O a) |  X b)  |  ? c)  |
            # tx_nom  | [    X d)    ] |  O e)  |
            # tx_nnb  | [    X f)    ] |  O g)  |

            if tax_status == 'found' and par_status is None:
                # a) Good backbone with no parent, provide info on taxon status
                if 'user' in gbif_info:
                    LOGGER.warn('Row {} ({}): considered a {} of {} in GBIF '
                                'backbone'.format(rw_num, taxon[0], gbif_info['user'][4],
                                                  gbif_info['canon'][2]))
                    self.taxon_index.add(gbif_info['user'] + (None, None))
                    self.taxon_index.add(gbif_info['canon'] + gbif_info['user'][2:4])

                else:
                    LOGGER.info('Row {} ({}): in GBIF backbone '
                                '({})'.format(rw_num, taxon[0], gbif_info['canon'][4]))
                    self.taxon_index.add(gbif_info['canon'] + (None, None))

                # update hierarchy
                taxon_hierarchy.update(gbif_info['hier'])

            elif tax_status == 'found' and par_status == 'invalid':
                # b) Good backbone with bad parent
                LOGGER.error('Row {} ({}): found but provided parent information is '
                             'not valid.'.format(rw_num, taxon[0]))

            elif tax_status == 'found' and par_status == 'valid':
                # c) Good backbone with good parent - are they compatible? Check if all
                #    entries in the parent hierarchy appear in the taxon hierarchy
                parent_hier = set(parent_info['hier'])
                taxon_hier = set(gbif_info['hier'])
                if not set(parent_hier).issubset(taxon_hier):
                    LOGGER.error('Row {} ({}): found in GBIF backbone, but additional parent '
                                 'information is incompatible'.format(rw_num, taxon[0]))
                else:
                    if 'user' in gbif_info:
                        LOGGER.warn('Row {} ({}): considered a {} of {} in GBIF '
                                    'backbone'.format(rw_num, taxon[0], gbif_info['user'][4],
                                                      gbif_info['canon'][2]))

                        # Add user and canonical usage, including as_name and as_rank
                        self.taxon_index.add(gbif_info['user'] + (None, None))
                        self.taxon_index.add(gbif_info['canon'] + gbif_info['user'][2:4])
                    else:
                        LOGGER.info('Row {} ({}): in GBIF backbone '
                                    '({})'.format(rw_num, taxon[0], gbif_info['canon'][4]))
                        self.taxon_index.add(gbif_info['canon'] + (None, None))

                    taxon_hierarchy.update(gbif_info['hier'])

            elif tax_status == 'no_match' and par_status is None:
                # d) Taxon is a backbone type but is not found in GBIF.
                if 'note' in gbif_info:
                    LOGGER.error('Row {} ({}): {}'.format(rw_num, taxon[0], gbif_info['note']))
                else:
                    LOGGER.error('Row {} ({}): name and rank combination '
                                 'not found'.format(rw_num, taxon[0]))

            elif tax_status == 'no_match' and par_status == 'invalid':
                # d) Taxon is a backbone type not found in GBIF and the provided parent isn't valid
                LOGGER.error('Row {} ({}): not found in GBIF and has invalid parent '
                             'information.'.format(rw_num, taxon[0]))

            elif tax_status == 'no_match' and par_status == 'valid':
                # e) Taxon is a backbone type that is not present in GBIF but the user has provided
                #    a valid set of parent taxon information.
                LOGGER.info('Row {} ({}): not found in GBIF but has valid parent '
                            'information'.format(rw_num, taxon[0]))

                # construct a taxon index entry and add the parent hierarchy
                self.taxon_index.add((-1, parent_info['canon'][1], taxon[1][0],
                                      taxon[1][1].lower(), 'user', None, None))
                taxon_hierarchy.update(parent_info['hier'])

            elif tax_status == 'non-backbone' and (par_status is None or par_status == 'invalid'):
                # f) Taxon is a non-backbone type - must have a valid parent to be accepted.
                LOGGER.error('Row {} ({}): taxa of type {} must have valid parent '
                             'information.'.format(rw_num, taxon[0], taxon[1][1]))

            elif tax_status == 'non-backbone' and par_status == 'valid':
                # g) Taxon is a non backbone type with good parent info
                LOGGER.info("Row {} ({}): {} with valid parent "
                            "information ".format(rw_num, taxon[0], taxon[1][1]))

                # construct the taxon index - use the name for alternative types,
                # otherwise use the taxon name
                if taxon[1][1].lower() in alt_types:
                    taxon_entry = (-1, parent_info['canon'][1], taxon[0],
                                   taxon[1][1].lower(), 'user', None, None)
                else:
                    taxon_entry = (-1, parent_info['canon'][1], taxon[1][0],
                                   taxon[1][1].lower(), 'user', None, None)

                self.taxon_index.add(taxon_entry)
                taxon_hierarchy.update(parent_info['hier'])

            else:
                # Think all the combinations are covered but exit with information if not
                raise AttributeError('Bad taxon processing')

        # Look up the unique taxon hierarchy entries
        # - drop taxa already in the index
        indexed = [tx[0] for tx in self.taxon_index]
        taxon_hierarchy = {tx for tx in taxon_hierarchy if tx[1] not in indexed}

        # - sort into ascending taxonomic order
        taxon_hierarchy = list(taxon_hierarchy)
        taxon_hierarchy.sort(key=lambda val: backbone_types.index(val[0]))
        LOGGER.info('Indexing taxonomic hierarchy', extra={'indent_before': 1, 'indent_after': 2})

        # The two versions differ in what is used to look up hierarchy. The web API
        # has (rank, GBIF ID) and the local API has (rank, canonical name)
        for tx_lev, tx_id in taxon_hierarchy:
            if self.gbif_conn is not None:
                canon = local_gbif_validate(self.gbif_conn, tx_id, tx_lev)['canon']
            else:
                canon = web_gbif_validate(None, None, gbif_id=tx_id)['canon']

            self.taxon_index.add(canon + (None, None))
            LOGGER.info('Added {}'.format(canon[2]))

        # summary of processing
        n_errors = CH.counters['ERROR'] - start_errors
        if n_errors > 0:
            LOGGER.info('Taxa contains {} errors'.format(n_errors),
                        extra={'indent_before': 1})
        else:
            LOGGER.info('{} taxa loaded correctly'.format(len(self.taxon_names)),
                        extra={'indent_before': 1})

    def load_data_worksheet(self, meta):

        """
        A method to load and check the formatting and content of a
        data worksheet, turning the dictionary for this worksheet
        in the dataset object into a DataWorksheet object.

        Args:
            meta: A reference to the metadata dictionary for this
            worksheet in the datasets attribute, containing the
            worksheet name, title and description from the summary.
        """

        # now start populating with data
        if meta['name'] not in self.sheet_names:
            LOGGER.error('Data worksheet {} not found'.format(meta['name']),
                         extra={'indent_before': 0, 'indent_after': 1})
            return

        LOGGER.info('Checking data worksheet {}'.format(meta['name']),
                    extra={'indent_before': 0, 'indent_after': 1})

        # Create a dataworksheet to store details: basically
        # just a dictionary with dot notation and defaults.
        dwsh = DataWorksheet(meta)
        start_errors = CH.counters['ERROR']

        # get the worksheet and data dimensions
        worksheet = self.workbook[dwsh.name]
        dwsh.max_col = worksheet.max_column
        dwsh.max_row = worksheet.max_row

        # trap completely empty worksheets
        if dwsh.max_row == 1:
            LOGGER.error('Worksheet is empty')
            return

        # get the metadata field names
        # - first search at most the first 10 rows for the 'field_name' descriptor
        #   which shows the end of the metadata and the start of the data
        descriptors = [worksheet.cell(column=1, row=rw).value
                       for rw in range(1, min(10, dwsh.max_row) + 1)]
        if 'field_name' in descriptors:
            dwsh.field_name_row = descriptors.index('field_name') + 1
            dwsh.descriptors = descriptors[:dwsh.field_name_row]
        else:
            LOGGER.error('Cannot parse data: field_name row not found')
            return

        # Neither of the row and col maxima are particularly reliable as Excel can hang on to
        # cell references for previously used cells. We can ignore blank columns easily but
        # we do need to know where to actually stop for finding blank data in rows.
        # So, we explicitly check the row numbers, making them a mandatory part of the setup.

        # - get the values
        row_number_cells = worksheet.get_squared_range(1, dwsh.field_name_row + 1, 1, dwsh.max_row)
        row_numbers = [cl[0].value for cl in row_number_cells]

        # - trim blank or whitespace values from the end
        while is_blank(row_numbers[-1]):
            row_numbers.pop()

        # now check the row numbers are numbers and if they are
        # do they start at one and go up by one
        if not all([isinstance(vl, numbers.Number) for vl in row_numbers]):
            LOGGER.error('Non-numeric data found in row numbering')
        else:
            if row_numbers[0] != 1:
                LOGGER.error('Row numbering does not start at 1')

            one_increment = [(vl1 - vl2) == 1 for vl1, vl2 in
                             zip(row_numbers[1:], row_numbers[:-1])]
            if not all(one_increment):
                LOGGER.error('Row numbering does not consistently increment by 1')

        # Test for data with no row number
        data_end = len(row_numbers) + dwsh.field_name_row
        if data_end < dwsh.max_row:
            allegedly_blank = worksheet.get_squared_range(1, data_end + 1,
                                                          dwsh.max_col, dwsh.max_row)
            not_blank = [not is_blank(cell.value) for row in allegedly_blank for cell in row]
            if any(not_blank):
                LOGGER.error('Data extends beyond row numbering')

        # report on detected size
        dwsh.max_row = data_end
        dwsh.n_data_row = len(row_numbers)
        LOGGER.info('Worksheet contains {} rows and {} columns'.format(dwsh.max_row, dwsh.max_col),
                    extra={'indent_after': 2})

        # get the rows of metadata
        metadata = worksheet.get_squared_range(2, 1, dwsh.max_col, dwsh.field_name_row)
        # extract data
        metadata = [[cell.value for cell in row] for row in metadata]
        # turn into columns for each field
        metadata = zip(*metadata)
        # and hence dictionaries for each field
        metadata = [dict(zip(dwsh.descriptors, fld)) for fld in metadata]

        # insert column index and letter
        for idx, fld in enumerate(metadata):
            fld[u'col_idx'] = idx + 2
            fld[u'column'] = utils.get_column_letter(idx + 2)

        # check field names unique (drop None). This doesn't cause as many problems
        # as duplications in Taxa and Locations, which expect certain fields, so warn
        # and continue.
        field_names = [fld['field_name'] for fld in metadata if fld['field_name'] is not None]
        dupes = duplication(field_names)
        if dupes:
            LOGGER.error('Field names duplicated: ', extra={'join': dupes})

        # get taxa field names for cross checking observation and trait data
        dwsh.taxa_fields = [fld['field_name'] for fld in metadata if fld['field_type'] == 'Taxa']

        # check the data in each field against the metadata
        for meta in metadata:

            # read the values and check them against the metadata
            data = worksheet.get_squared_range(meta['col_idx'], dwsh.field_name_row + 1,
                                               meta['col_idx'], dwsh.max_row)
            data = [cl[0].value for cl in data]
            self.check_field(dwsh, meta, data)

        # add the new DataWorksheet into the Dataset
        self.dataworksheets.append(dwsh)

        # reporting
        n_errors = start_errors - CH.counters['ERROR']
        if n_errors > 0:
            LOGGER.info('Dataframe contains {} errors'.format(n_errors),
                        extra={'indent_before': 1})
        else:
            LOGGER.info('Dataframe formatted correctly',
                        extra={'indent_before': 1})

    def check_field(self, dwsh, meta, data):
        """
        Method to test whether the contents of a field are compatible
        with the metadata for the field. Largely used for checking
        data worksheets, but also by other methods. Acts as a wrapper
        around data type specific subfunctions.

        Args:
            dwsh: A DataWorksheet, used to pass worksheet level information,
             which at the moment is just taxa fields.
            meta: A dictionary of metadata
            data: A list of values
        """

        # Print out column checking header
        if is_blank(meta['field_name']):
            LOGGER.info('Checking Column {}'.format(meta['column']),
                        extra={'indent_before': 2, 'indent_after': 3})
        else:
            LOGGER.info('Checking field {field_name}'.format(**meta),
                        extra={'indent_before': 2, 'indent_after': 3})

        # Skip any field with no user provided metadata or data
        blank_data = [is_blank(vl) for vl in data]
        blank_meta = [is_blank(vl) for ky, vl in meta.iteritems()
                      if ky not in ['col_idx', 'column']]
        if all(blank_data) and all(blank_meta):
            LOGGER.info('Blank column loaded - safe to ignore')
            return
        elif all(blank_meta):
            LOGGER.error('Field contains no descriptor information but does contain values')
            return

        # try and figure out what else is available
        if is_blank(meta['field_name']):
            LOGGER.error('Field name is blank')

        # check the description
        if is_blank(meta['description']):
            LOGGER.error('Description is missing')

        # filter out missing and blank data, except for comments fields, where
        # blanks are not an error
        if meta['field_type'] != 'Comments':

            # Only NA is acceptable
            na_vals = [vl == u'NA' for vl in data]
            if any(na_vals):
                LOGGER.warn('{} / {} values missing'.format(sum(na_vals), len(na_vals)))

            # We won't tolerate blank data
            if any(blank_data):
                LOGGER.error('{} cells are blank or contain only whitespace '
                             'text'.format(sum(blank_data)))

            # Return the values that aren't NA, blank or whitespace only
            na_or_blank = [any(tst) for tst in zip(na_vals, blank_data)]
            data = [dt for dt, nb in zip(data, na_or_blank) if not nb]

        # run consistency checks where needed and trap unknown field types
        if meta['field_type'] == 'Date':
            self.check_field_datetime(meta, data, which='date')
        elif meta['field_type'] == 'Datetime':
            self.check_field_datetime(meta, data, which='datetime')
        elif meta['field_type'] == 'Time':
            self.check_field_time(meta, data)
        elif meta['field_type'] == 'Taxa':
            self.check_field_taxa(data)
        elif meta['field_type'] == 'Location':
            self.check_field_locations(data)
        elif meta['field_type'] in ['Categorical', 'Ordered Categorical']:
            self.check_field_categorical(meta, data)
        elif meta['field_type'] == 'Numeric':
            self.check_field_numeric(meta, data)
        elif meta['field_type'] == 'Abundance':
            self.check_field_abundance(meta, data, dwsh.taxa_fields)
        elif meta['field_type'] in ['Categorical Trait', 'Ordered Categorical Trait']:
            self.check_field_trait(meta, data, dwsh.taxa_fields, which='categorical')
        elif meta['field_type'] == 'Numeric Trait':
            self.check_field_trait(meta, data, dwsh.taxa_fields, which='numeric')
        elif meta['field_type'] in ['Categorical Interaction', 'Ordered Categorical Interaction']:
            self.check_field_interaction(meta, data, dwsh.taxa_fields, which='categorical')
        elif meta['field_type'] == 'Numeric Interaction':
            self.check_field_interaction(meta, data, dwsh.taxa_fields, which='numeric')
        elif meta['field_type'] == 'Latitude':
            self.check_field_geo(meta, data, which='latitude')
        elif meta['field_type'] == 'Longitude':
            self.check_field_geo(meta, data, which='longitude')
        elif meta['field_type'] in ['Replicate', 'ID']:
            # We've looked for missing data, no other constraints.
            pass
        elif meta['field_type'] == 'Comments':
            pass
        elif meta['field_type'] is None:
            LOGGER.error('Field type is empty')
        else:
            LOGGER.error('Unknown field type {field_type}'.format(**meta))

        # extend the dataworksheet fields
        dwsh.fields.append(meta)

    # Helper functions for checking data fields
    @staticmethod
    def _check_meta(meta, descriptor):
        """
        A standardised check to see if a required descriptor is present for
        a field and that it isn't simply empty or whitespace. The function
        reports problems to the Messages instance and returns a boolean
        showing if the checks passed successfully.

        Args:
            meta: A dictionary of field metadata descriptors
            descriptor: The name of the descriptor to check.

        Returns:
            A boolean, with True showing no problems and False showing
            that warnings occurred.
        """

        if descriptor not in meta:
            LOGGER.error('{} descriptor missing'.format(descriptor))
            return False
        elif is_blank(meta[descriptor]):
            LOGGER.error('{} descriptor is blank'.format(descriptor))
            return False
        else:
            return True

    def _check_taxon_meta(self, meta, taxa_fields):

        """
        Checks the taxonomic metadata of abundance and trait fields.
        This is more involved that the simple _check_meta(), because
        of the option to provide taxon_name or taxon_field descriptors

        Args:
            meta: A dictionary of metadata descriptors for the field
            taxa_fields: A list of Taxa fields in this worksheet.

        Returns:
            A boolean, with True showing no problems and False showing
            that warnings occurred.
        """

        # Are taxon_name and taxon_field provided and not blank: note use of
        # 'and' rather than '&' to allow missing descriptors to short cut

        tx_nm_prov = ('taxon_name' in meta) and (not is_blank(meta['taxon_name']))
        tx_fd_prov = ('taxon_field' in meta) and (not is_blank(meta['taxon_field']))

        if tx_nm_prov and tx_fd_prov:
            LOGGER.error('Taxon name and taxon field both provided, use one only')
            return False
        elif tx_nm_prov and meta['taxon_name'] not in self.taxon_names:
            LOGGER.error('Taxon name not found in the Taxa worksheet')
            return False
        elif tx_fd_prov and meta['taxon_field'] not in taxa_fields:
            LOGGER.error("Taxon field not found in this worksheet")
            return False
        elif not tx_nm_prov and not tx_fd_prov:
            LOGGER.error("One of taxon name or taxon field must be provided")
            return False
        else:
            if tx_nm_prov:
                self.taxon_names_used.update([meta['taxon_name']])
            return True

    def _check_interaction_meta(self, meta, taxa_fields):

        """
        Checks the taxonomic metadata of interaction fields.
        This is more involved that the simple _check_meta(), because
        of the option to provide taxon_name or taxon_field descriptors
        describing at least two taxonomic identifiers.

        Args:
            meta: A dictionary of metadata descriptors for the field
            taxa_fields: A list of Taxa fields in this worksheet.

        Returns:
            A boolean, with True showing no problems and False showing
            that warnings occurred.
        """

        # Are interaction_name and/or interaction_field provided and not blank:
        #  note use of 'and' rather than '&' to allow missing descriptors to short cut

        iact_nm_prov = ('interaction_name' in meta) and (not is_blank(meta['interaction_name']))
        iact_fd_prov = ('interaction_field' in meta) and (not is_blank(meta['interaction_field']))

        if not iact_nm_prov and not iact_fd_prov:
            LOGGER.error("At least one of interaction name or interaction field must be provided")
            return False
        else:
            if iact_nm_prov:
                # get the taxon names and descriptions from interaction name providers
                int_nm_lab, int_nm_desc = self._parse_levels(meta['interaction_name'])
                # add names to used taxa
                self.taxon_names_used.update(int_nm_lab)
                # check they are found
                if not all([lab in self.taxon_names for lab in int_nm_lab]):
                    LOGGER.error('Unknown taxa in interaction_name descriptor')
                    nm_check = False
                else:

                    nm_check = True
            else:
                int_nm_lab, int_nm_desc = [(), ()]
                nm_check = True

            if iact_fd_prov:
                # check any field labels match to known taxon fields
                int_fd_lab, int_fd_desc = self._parse_levels(meta['interaction_field'])
                if not all([lab in taxa_fields for lab in int_fd_lab]):
                    LOGGER.error('Unknown taxon fields in interaction_field descriptor')
                    fd_check = False
                else:
                    fd_check = True
            else:
                int_fd_lab, int_fd_desc = [(), ()]
                fd_check = True

            if len(int_nm_lab + int_fd_lab) < 2:
                LOGGER.error('At least two interacting taxon labels or fields must be identified')
                num_check = False
            else:
                num_check = True

            if nm_check and fd_check and num_check:
                return True
            else:
                return False

    @staticmethod
    def _parse_levels(txt):
        """
        Splits up category information formatted as label:desc;label:desc, which
        is used in both levels for categorical data and interaction descriptors.
        Args:
            txt: The text string to parse

        Returns:
            A list of two tuples of label and descriptions.
        """

        # remove terminal semi-colon, if used.
        if txt.endswith(';'):
            txt = txt[:-1]
        # - split the text up by semi-colon
        parts = txt.split(';')
        # - split descriptions
        parts = [pt.split(':') for pt in parts]
        n_parts = [len(pt) for pt in parts]

        # simple formatting checks
        if any([pt > 2 for pt in n_parts]):
            LOGGER.error('Extra colons in level description.')

        # standardise descriptions
        if all([pt == 1 for pt in n_parts]):
            parts = [[pt[0], None] for pt in parts]
        elif all([pt == 2 for pt in n_parts]):
            pass
        else:
            LOGGER.error('Provide descriptions for either all or none of the categories')
            parts = [pt[0:2] if len(pt) >= 2 else [pt[0], None] for pt in parts]

        return zip(*parts)

    def check_field_datetime(self, meta, data, which='datetime'):

        """
        Checks for data consistency in date and datetime fields. Excel
        doesn't distinguish, both are loaded as datetime.datetime objects
        so we check that the values are compatible with the user provided
        field type.

        Args:
            meta: The field metadata, to be updated with the range
            data: A list of data values, allegedly containing datetimes or dates
            which: The datetime type to be checked for.
        """

        # Check type (excluding NA values)
        bad = [dt for dt in data if not isinstance(dt, (datetime.datetime, datetime.time))]
        if len(bad):
            LOGGER.error('Data in field not formatted as date. Note that text can look'
                         '_exactly_ like a date: ', extra={'join': bad})

        is_time = [isinstance(dt, datetime.time) for dt in data]
        if any(is_time):
            LOGGER.warn('Some values _only_  contain time components')

        data = [dt for dt in data if isinstance(dt, datetime.datetime)]

        # Check no time component in actual dates
        if which == 'date':
            no_time = [vl.time() == datetime.time(0, 0) for vl in data]
            if not all(no_time):
                LOGGER.error('Some values also contain time components')

        # update the field metadata and the dataset extent
        extent = (min(data), max(data))
        meta['range'] = extent
        self.update_extent(extent, datetime.datetime, 'temporal_extent')

    def check_field_time(self, meta, data):

        """
        Checks for data consistency in time fields and reports to the
        Messages instance.

        Args:
            meta: The field metadata, to be updated with the range
            data: A list of data values, allegedly of type datetime.time
        """

        # Check type (excluding NA values)
        type_check = set([type(vl) for vl in data])
        if type_check != {datetime.time}:
            LOGGER.error('Non-time formatted data found.')

        # update the field metadata
        real_times = [dt for dt in data if isinstance(dt, datetime.time)]
        meta['range'] = (min(real_times), max(real_times))

    def check_field_taxa(self, data):

        """
        Checks if all the values provided in a Taxon field are found
        in the Taxa worksheet.

        Args:
            data: A list of data values, allegedly taxon names
        """

        found = set(data)
        if self.taxon_names == set():
            LOGGER.error('No taxa loaded', 2)
        if not found.issubset(self.taxon_names):
            LOGGER.error('Includes taxa missing from Taxa worksheet: ',
                         extra={'join': found - self.taxon_names, 'quote': True})

        # add the found taxa to the list of taxa used
        self.taxon_names_used.update(found)

    def check_field_locations(self, data):

        """
        Checks if all the values provided in a Locations field are
        found in the Locations worksheet, reporting to the Messages instance.

        Args:
            data: A list of data values, allegedly taxon names
        """

        # location names should be strings
        data = [unicode(dt) for dt in data]

        # check if locations are all provided
        found = set(data)
        if self.locations == set():
            LOGGER.error('No locations loaded')
        elif not found.issubset(self.locations):
            LOGGER.error('Includes locations missing from Locations worksheet:',
                         extra={'join': found - self.taxon_names, 'quote': True})

        # add the locations to the set of locations used
        self.locations_used.update(found)

    def check_field_abundance(self, meta, data, taxa_fields):

        """
        Checks abundance type data, reporting to the Messages instance.

        Args:
            meta: A dictionary of metadata descriptors for the field
            data: A list of data values, allegedly numeric
            taxa_fields: A list of Taxa fields in this worksheet.
        """

        # check the required descriptors
        self._check_meta(meta, 'method')
        self._check_taxon_meta(meta, taxa_fields)

        # Can still check values are numeric, whatever happens above.
        # We're not going to insist on integers here - could be mean counts.
        all_nums, nums = all_numeric(data)
        if not all_nums:
            LOGGER.error('Field contains non-numeric data')

        # update the field metadata
        meta['range'] = (min(nums), max(nums))

    def check_field_categorical(self, meta, data):

        """
        Checks factor data, reporting to the Messages instance.

        Args:
            meta: A dictionary of metadata descriptors for the field
            data: A list of data values, allegedly strings from a provided set of levels
        """

        # Are levels found?
        ct_ok = self._check_meta(meta, 'levels')

        if not ct_ok:
            # Can't check further if no levels descriptor
            pass
        elif ct_ok and not isinstance(meta['levels'], unicode):
            # Can't really check anything here either
            LOGGER.error('Category description does not seem to be text')
        else:
            # Now we can test if the labels match up
            level_labels, level_desc = self._parse_levels(meta['levels'])

            # - repeated labels?
            if len(set(level_labels)) < len(level_labels):
                LOGGER.error('Repeated level labels')

            # - check for integer level names
            integer_codes = [is_integer_string(vl) for vl in level_labels]
            if any(integer_codes):
                LOGGER.error('Integer level names not permitted')

            # Now look for consistency: get the unique values reported in the
            # data, convert to unicode to handle checking of integer labels and
            # then check the reported levels are a subset of the descriptors.
            reported = set(data)
            reported = {unicode(lv) for lv in reported}

            if not reported.issubset(level_labels):
                LOGGER.error('Categories found in data missing from description: ',
                             extra={'join': reported - set(level_labels), 'quote': True})

    def check_field_numeric(self, meta, data):

        """
        Checks numeric type data, reporting to the Messages instance.

        Args:
            meta: A dictionary of metadata descriptors for the field
            data: A list of data values, allegedly numeric
        """

        # Check required descriptors
        self._check_meta(meta, 'units')
        self._check_meta(meta, 'method')

        # Regardless of the outcome of the meta checks, can still check the
        # data is all numeric, as it claims to be.
        all_nums, nums = all_numeric(data)
        if not all_nums:
            LOGGER.error('Field contains non-numeric data')

        # update the field metadata
        meta['range'] = (min(nums), max(nums))

    def check_field_trait(self, meta, data, taxa_fields, which='categorical'):

        """
        Checks trait type data and reports to the Messages instance. Just a wrapper
        to check that a valid taxon has been provided before handing off to
        check_field_categorical or check_field_numeric

        Args:
            meta: A dictionary of metadata descriptors for the field
            data: A list of data values, allegedly numeric
            taxa_fields: A list of Taxa fields in this worksheet.
            which: The type of trait data in the field
        """

        # Check required descriptors
        self._check_taxon_meta(meta, taxa_fields)

        # Regardless of the outcome of the meta checks, check the contents:
        if which == 'categorical':
            self.check_field_categorical(meta, data)
        elif which == 'numeric':
            self.check_field_numeric(meta, data)

    def check_field_interaction(self, meta, data, taxa_fields, which='categorical'):

        """
        Checks interaction type data and reports to the Messages instance.
        Just a wrapper to check that intercating taxa have been identified
        before handing off to check_field_categorical or check_field_numeric

        Args:
            meta: A dictionary of metadata descriptors for the field
            data: A list of data values, allegedly numeric
            taxa_fields: A list of Taxa fields in this worksheet.
            which: The type of trait data in the field
        """

        # Check required descriptors
        self._check_interaction_meta(meta, taxa_fields)

        # Regardless of the outcome of the meta checks, check the contents:
        if which == 'categorical':
            self.check_field_categorical(meta, data)
        elif which == 'numeric':
            self.check_field_numeric(meta, data)

    def check_field_geo(self, meta, data, which='latitude'):

        """
        Checks geographic coordinates. It also automatically updates
        the geographic extent of the dataset.

        Args:
            meta: A dictionary of metadata descriptors for the field
            data: A list of data values
            which: One of latitude or longitude
        """

        # Are the values represented as decimal degrees - numeric.
        all_nums, nums = all_numeric(data)
        if not all_nums:
            LOGGER.error('Non numeric data found')
            if any([RE_DMS.search(unicode(vl)) for vl in data]):
                LOGGER.warn('Possible degrees minutes and seconds formatting? Use decimal degrees')

        # Check the locations
        if len(nums):
            min_geo = float(min(nums))
            max_geo = float(max(nums))
            extent = (min_geo, max_geo)

            if which == 'latitude':
                bnds = [-90, -4, 8, 90]
            elif which == 'longitude':
                bnds = [-180, 108, 120, 180]

            out_of_bounds = min_geo < bnds[0] or max_geo > bnds[3]
            out_of_borneo = bnds[0] <= min_geo < bnds[1] or bnds[2] < max_geo <= bnds[3]

            if out_of_bounds:
                LOGGER.error('{0} values not in valid range[{1[0]}, {1[3]}]: '
                             '{2}'.format(which.capitalize(), bnds, extent))
            elif out_of_borneo:
                LOGGER.warn('{0} values not in Borneo [{1[1]}, {1[2]}]: '
                            '{2}'.format(which.capitalize(), bnds, extent))

            # update the field metadata and the dataset extent
            meta['range'] = extent
            # Look up the extent name to update and then update it
            which_extent = {'latitude': 'latitudinal_extent', 'longitude': 'longitudinal_extent'}
            self.update_extent(extent, float, which_extent[which])

    def final_checks(self):
        """
        A method to run final checks:
        i)  The locations and taxa provided have been used in the data worksheets scanned.
        ii) Report the total number of errors and warnings
        """
        LOGGER.info('Checking provided locations and taxa all used in data worksheets',
                    extra={'indent_before': 0, 'indent_after': 1})
        # check locations
        if self.locations_used != self.locations:
            LOGGER.error('Provided locations not used: ',
                         extra={'join': self.locations - self.locations_used})
        else:
            LOGGER.info('Provided locations all used in datasets')

        # check taxa
        if self.taxon_names_used != self.taxon_names:
            LOGGER.error('Provided taxa  not used: ',
                         extra={'join': self.taxon_names - self.taxon_names_used})
        else:
            LOGGER.info('Provided taxa all used in datasets')

        if CH.counters['ERROR'] > 0:
            if CH.counters['WARNING'] > 0:
                LOGGER.info('FAIL: file contained {} errors and {} '
                            'warnings'.format(CH.counters['ERROR'], CH.counters['WARNING']),
                            extra={'indent_before': 0})
            else:
                LOGGER.info('FAIL: file contained {} errors'.format(CH.counters['ERROR']),
                            extra={'indent_before': 0})
        else:
            if CH.counters['WARNING'] > 0:
                LOGGER.info('PASS: file formatted correctly but with {} '
                            'warnings.'.format(CH.counters['WARNING']),
                            extra={'indent_before': 0})
                self.passed = True
            else:
                LOGGER.info('PASS: file formatted correctly with no warnings',
                            extra={'indent_before': 0})
                self.passed = True

    def export_metadata_dict(self):
        """
        Function to return a simple dictionary of the metadata content, combining
        dataworksheets, into a single representation of the contents

        Returns:
            A dictionary of metadata
        """

        # get the required components
        component_keys = ['access', 'authors', 'description', 'embargo_date', 'filename',
                          'keywords', 'latitudinal_extent', 'longitudinal_extent',
                          'project_id', 'temporal_extent', 'title']

        components = {ky: vl for ky, vl in self.__dict__.iteritems() if ky in component_keys}

        # add in the data worksheets
        components['dataworksheets'] = [dwsh.__dict__ for dwsh in self.dataworksheets]

        return components


"""
Higher level functions
"""


def check_file(fname, verbose=True, gbif_database=None, locations_json=None, validate_doi=False,
               check='sltwf'):

    """
    Runs the format checking across an Excel workbook.

    Parameters:
        fname: Path to an Excel file
        verbose: Boolean to indicate whether to print messages to stdout as the program runs?
        gbif_database: Path to a local copy of the GBIF taxonomy backbone if that is to be used
          instead of the GBIF web API.
        locations_json: The path to a json file of valid location names
        validate_doi: Check any publication DOIs resolve, requiring a web connection.
        check: String indicating which checking to run. Note that running worksheet checking
          without loading summary taxa and locations is going to be problematic!

    Returns:
        A Dataset object
    """

    # initialise the dataset object
    dataset = Dataset(fname, verbose=verbose, gbif_database=gbif_database)

    # load the metadata sheets
    if 's' in check:
        dataset.load_summary(validate_doi=validate_doi)
    if 't' in check:
        dataset.load_taxa()
    if 'l' in check:
        dataset.load_locations(locations_json=locations_json)

    # check the datasets
    if 'w' in check:
        if dataset.dataworksheet_summaries:
            for ws in dataset.dataworksheet_summaries:
                dataset.load_data_worksheet(ws)
        else:
            LOGGER.error('No data worksheets found')

    if 'f' in check:
        dataset.final_checks()

    return dataset


def main():

    """
    This program validates an Excel file formatted as a SAFE dataset. As it runs, it outputs
    a report that highlights any problems with the formatting.

    Much of the validation is to check that the data meets our metadata standards and is
    internally consistent. However, it uses external sources to perform validation in three
    areas.

    1. Taxon validation. The program validates taxonomic names against the GBIF taxonomy
    backbone. By default, it uses the GBIF web API to validate names, but can also use a
    local copy of the backbone provided in a sqlite database: this will work offline and
    is much faster but requires some simple setup.

    2. Location names. The program also validate sampling location names against the SAFE gazeteer.
    By default, this is loaded automatically from the SAFE website so requires an internet
    connection, but a local copy can be provided for offline use.

    3. DOI checking. Optionally, the program will validate any DOIs provided as having used
    the database. This requires a web connection and cannot be performed offline.
    """

    parser = argparse.ArgumentParser(description=main.__doc__)
    parser.add_argument('fname', help="Path to the Excel file to be validated.")
    parser.add_argument('-c', '--check', default='sltwf',
                        help='Which of the summary, locations, taxa, worksheets and '
                             'finalisation should be checked.')
    parser.add_argument('-l', '--locations_json', default=None,
                        help='Path to a locally stored json file of valid location names')
    parser.add_argument('-g', '--gbif_database', default=None,
                        help=('The path to a local sqlite database containing the GBIF '
                              'taxonomy backbone.'))
    parser.add_argument('--validate_doi', action="store_true", default=False,
                        help=('Check the validity of any publication DOIs, '
                              'provided by the user. Requires a web connection.'))

    args = parser.parse_args()

    check_file(fname=args.fname, verbose=True, locations_json=args.locations_json,
               gbif_database=args.gbif_database, validate_doi=args.validate_doi,
               check=args.check)


if __name__ == "__main__":
    main()
