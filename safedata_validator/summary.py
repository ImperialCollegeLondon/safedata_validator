"""This module handles the parsing and validation of the summary table for a Dataset. The
table consists of rows of information, with row labels in the first column. The module
defines the single [Summary object][safedata_validator.summary.Summary] which provides
methods for loading the summary data from file.
"""  # noqa D415

import datetime
import re
from dataclasses import dataclass

import requests  # type: ignore
from openpyxl.worksheet.worksheet import Worksheet

from safedata_validator.extent import Extent
from safedata_validator.logger import LOGGER, get_handler, loggerinfo_push_pop
from safedata_validator.resources import Resources
from safedata_validator.validators import (
    IsNotPadded,
    IsNotSpace,
    IsString,
    NoPunctuation,
    blank_value,
)

# Compile some regex expressions used in Summary checking
RE_DOI = re.compile(r"https?://(dx.)?doi.org/")
RE_ORCID = re.compile(r"[0-9]{4}-[0-9]{4}-[0-9]{4}-[0-9]{3}[0-9X]")
RE_EMAIL = re.compile(r"\S+@\S+\.\S+")
RE_NAME = re.compile(r"[^,]+,[ ]?[^,]+")
RE_CONTAINS_WSPACE = re.compile(r"\s")


@dataclass
class SummaryField:
    """Summary field details.

    This dataclass records details of summary fields within a block:

    * the expected key in the first column of the summary table,
    * whether a field is mandatory within the block,
    * an optional internal name to be used for the key,
    * the acceptable types that values in the field can have.
    """

    key: str
    mandatory: bool
    internal: str | None
    types: type | tuple[type, ...]


@dataclass
class SummaryBlock:
    """Summary block details.

    This dataclass records details of a block of related summary fields:

    * The list of fields in the block
    * Whether the block is mandatory
    * The title for the block, used in reporting
    * If the block should only accept a single record.
    """

    fields: list[SummaryField]
    mandatory: bool
    title: str
    singular: bool


class Summary:
    """Interface for dataset summary metadata.

    This class provides an interface to the summary metadata for a dataset. The loading
    methods check the information provided in the summary worksheet and populates the
    attributes of the class instance to pass that information to other components.

    The methods are intended to try and get as much information as possible from the
    Summary table: the instance attributes may therefore be set to None for missing
    metadata, so classes using Summary should handle None values.

    Args:
        resources: An instance of Resources providing the safedata_validator
            configuration.
    """

    def __init__(self, resources: Resources) -> None:
        self.title: str
        """A string giving the dataset title."""
        self.description: str
        """A string giving a description of the dataset."""
        self.access: dict
        """A dictionary giving access metadata."""
        self.authors: list[dict]
        "A list of dictionaries of author metadata."
        self.permits: list[dict]
        """A list of dictionaries of research permit metadata."""
        self.publication_doi = None
        """A list of DOIs associated with the dataset."""
        self.funders = None
        """A list of dictionaries of funder metadata."""
        self.keywords: list[str]
        """A list of keyword strings."""
        self.temporal_extent: Extent = Extent(
            "temporal extent",
            (datetime.date,),
            hard_bounds=resources.extents.temporal_hard_extent,
            soft_bounds=resources.extents.temporal_soft_extent,
        )
        """Extent instance for the temporal extent of the Dataset."""
        self.latitudinal_extent: Extent = Extent(
            "latitudinal extent",
            (float, int),
            hard_bounds=resources.extents.latitudinal_hard_extent,
            soft_bounds=resources.extents.latitudinal_soft_extent,
        )
        """Extent instance for the latitudinal extent of the Dataset."""
        self.longitudinal_extent: Extent = Extent(
            "longitudinal extent",
            (float, int),
            hard_bounds=resources.extents.longitudinal_hard_extent,
            soft_bounds=resources.extents.longitudinal_soft_extent,
        )
        """Extent instance for the longitudinal extent of the Dataset."""
        self.external_files: list[dict] | None = None
        """A list of dictionaries of external file metadata."""
        self.data_worksheets: list[Worksheet] = []
        """A list of dictionaries of data tables in the Dataset."""

        self._rows: dict = {}
        """A private attribute holding the row data for the summary."""
        self._ncols: int
        """A private attribute holding the total number of columns in the summary."""
        self.n_errors: int = 0
        """The number of validation errors found in the summary."""
        self.projects: dict[int, str] = resources.projects
        """A dictionary of valid project data."""
        self.project_id: list[int] | None = None
        """A list of project ID codes, if project IDs are configured."""

        self.validate_doi = False
        """A boolean flag indicating whether DOI values should be validated."""

        # Define the blocks and fields in the summary - note that the project ids block
        # is mandatory if a project database has been populated.
        self.fields: dict[str, SummaryBlock] = dict(
            core=SummaryBlock(
                fields=[
                    SummaryField("title", True, None, str),
                    SummaryField("description", True, None, str),
                ],
                mandatory=True,
                title="Core fields",
                singular=True,
            ),
            project_ids=SummaryBlock(
                fields=[
                    SummaryField("safe project id", False, None, int),
                    SummaryField("project id", False, None, int),
                ],
                mandatory=True if self.projects else False,
                title="Project IDs",
                singular=False,
            ),
            access=SummaryBlock(
                fields=[
                    SummaryField("access status", True, "access", str),
                    SummaryField(
                        "embargo date", False, "embargo_date", datetime.datetime
                    ),
                    SummaryField("access conditions", False, "access_conditions", str),
                ],
                mandatory=True,
                title="Access details",
                singular=True,
            ),
            keywords=SummaryBlock(
                fields=[SummaryField("keywords", True, None, str)],
                mandatory=True,
                title="Keywords",
                singular=False,
            ),
            doi=SummaryBlock(
                fields=[SummaryField("publication doi", True, None, str)],
                mandatory=False,
                title="DOI",
                singular=False,
            ),
            date=SummaryBlock(
                fields=[
                    SummaryField("start date", True, None, datetime.datetime),
                    SummaryField("end date", True, None, datetime.datetime),
                ],
                mandatory=False,
                title="Date Extents",
                singular=True,
            ),
            geo=SummaryBlock(
                fields=[
                    SummaryField("west", True, None, float),
                    SummaryField("east", True, None, float),
                    SummaryField("south", True, None, float),
                    SummaryField("north", True, None, float),
                ],
                mandatory=False,
                title="Geographic Extents",
                singular=True,
            ),
            authors=SummaryBlock(
                fields=[
                    SummaryField("author name", True, "name", str),
                    SummaryField("author affiliation", False, "affiliation", str),
                    SummaryField("author email", False, "email", str),
                    SummaryField("author orcid", False, "orcid", str),
                ],
                mandatory=True,
                title="Authors",
                singular=False,
            ),
            funding=SummaryBlock(
                fields=[
                    SummaryField("funding body", True, "body", str),
                    SummaryField("funding type", True, "type", str),
                    SummaryField("funding reference", False, "ref", (str, int, float)),
                    SummaryField("funding link", False, "url", str),
                ],
                mandatory=False,
                title="Funding Bodies",
                singular=False,
            ),
            external=SummaryBlock(
                fields=[
                    SummaryField("external file", True, "file", str),
                    SummaryField("external file description", True, "description", str),
                ],
                mandatory=False,
                title="External Files",
                singular=False,
            ),
            worksheet=SummaryBlock(
                fields=[
                    SummaryField("worksheet name", True, "name", str),
                    SummaryField("worksheet title", True, "title", str),
                    SummaryField("worksheet description", True, "description", str),
                    SummaryField("worksheet external file", False, "external", str),
                ],
                mandatory=False,
                title="Worksheets",
                singular=False,
            ),
            permits=SummaryBlock(
                fields=[
                    SummaryField("permit type", True, "type", str),
                    SummaryField("permit authority", True, "authority", str),
                    SummaryField("permit number", True, "number", (str, int, float)),
                ],
                mandatory=False,
                title="Permits",
                singular=False,
            ),
        )
        """A dictionary setting the summary blocks that can be present."""

    @loggerinfo_push_pop("Checking Summary worksheet")
    def load(
        self,
        worksheet: Worksheet,
        sheetnames: set,
        validate_doi: bool = False,
    ) -> None:
        """Populate a Summary instance from an Excel Worksheet.

        Args:
            worksheet: An openpyxl worksheet instance.
            sheetnames: A set of sheet names found in the workbook.
            validate_doi: Should publication DOIs be validated (needs web connection).
        """
        handler = get_handler()
        start_errors = handler.counters["ERROR"]

        self.validate_doi = validate_doi

        rows = load_rows_from_worksheet(worksheet)

        self._ncols = worksheet.max_column

        # convert into dictionary using the lower cased first entry as the key after
        # checking for empty values (None) and non-string values.
        row_headers = IsString([r[0] for r in rows])
        if not row_headers:
            # Check for None separately because seeing 'None' as a field key in the
            # report is very confusing for end users.
            if None in row_headers.failed:
                LOGGER.error("Summary metadata fields column contains empty cells")
                row_headers

            # Get other non-string headers.
            non_none_failed_headers = [
                hdr for hdr in row_headers.failed if hdr is not None
            ]
            if non_none_failed_headers:
                LOGGER.error(
                    "Summary metadata fields column contains non text values: ",
                    extra={"join": non_none_failed_headers},
                )

        # Now we have warned about bad values, explicitly cast row keys to string to
        # continue processing.
        self._rows = {str(rw[0]).lower(): rw[1:] for rw in rows}

        # Check if metadata keys have white space padding
        self._check_for_whitespace()

        # Validate the keys found in the summary table
        self._validate_keys()

        # Now process the field blocks
        self._load_core()
        self._load_project_ids()
        self._load_access_details()
        self._load_authors()
        self._load_keywords()
        self._load_doi()
        self._load_temporal_extent()
        self._load_geographic_extent()
        self._load_funders()
        self._load_permits()
        self._load_external_files()
        self._load_data_worksheets(sheetnames)

        # summary of processing
        self.n_errors = handler.counters["ERROR"] - start_errors
        if self.n_errors > 0:
            LOGGER.info(f"Summary contains {self.n_errors} errors")
        else:
            LOGGER.info("Summary formatted correctly")

    def _check_for_whitespace(self) -> None:
        """Check that the summary keys do not have whitespace padding.

        This function checks that the keys in a summary table do not have white space
        padding, if they do the white space padding is removed and an error is logged.
        """

        clean_metadata_keys = IsNotPadded(self._rows.keys())
        if not clean_metadata_keys:
            # Report whitespace padding and clean up tuples
            LOGGER.error(
                "Whitespace padding in summary field names: ",
                extra={"join": clean_metadata_keys.failed},
            )

            # Order preserved in dict and validator
            cleaned_entries = [
                (ky, val) for ky, val in zip(clean_metadata_keys, self._rows.values())
            ]
            self._rows = dict(cleaned_entries)

    def _validate_keys(self) -> None:
        """Validate the summary keys recovered.

        This function checks that the keys in a summary table include the minimum set of
        mandatory fields in mandatory blocks and that all found keys are known.
        """

        # Populate found, required and known field keys
        found: set[str] = set(self._rows.keys())
        required: set[str] = set()
        known: set[str] = set()

        for block in self.fields.values():
            # Required keys
            if block.mandatory:
                required = required.union(
                    fld.key for fld in block.fields if fld.mandatory
                )
            # Known keys
            known = known.union(fld.key for fld in block.fields)

        # Look for and report on issue
        missing = required - found
        unknown = found - known
        if missing:
            LOGGER.error("Missing mandatory metadata fields: ", extra={"join": missing})

        if unknown:
            LOGGER.error("Unknown metadata fields: ", extra={"join": unknown})

    def _read_block(self, block: SummaryBlock) -> list | None:
        """Read a block of fields from a summary table.

        This internal method takes a given block definition from the Summary class
        [fields][safedata_validator.summary.Summary.fields] attribute and returns a list
        of dictionary records for that block. This function automatically does some
        common checking for missing data, bad input types etc, leaving the block
        specific functions to handle unique tests.

        Args:
            block: A SummaryBlock instance describing the block
        """

        mandatory_fields = [fld.key for fld in block.fields if fld.mandatory]
        optional_fields = [fld.key for fld in block.fields if not fld.mandatory]
        field_map = [
            (fld.key, fld.internal) for fld in block.fields if fld.internal is not None
        ]
        field_types = {fld.key: fld.types for fld in block.fields}

        # Get the full list of field names
        all_fields = mandatory_fields + optional_fields

        # Get the data, filling in completely missing rows
        data = {
            k: self._rows[k] if k in self._rows else [None] * (self._ncols - 1)
            for k in all_fields
        }

        # Empty cells are already None, but also filter values to catch
        # pure whitespace content and replace with None
        for ky, vals in data.items():
            vals = IsNotSpace(vals)
            if not vals:
                LOGGER.error(f"Whitespace only cells in field {ky}")

            data[ky] = vals.values

        # Pivot to dictionary of records
        block_list = [dict(zip(data.keys(), vals)) for vals in zip(*data.values())]

        # Drop empty records
        block_list = [bl for bl in block_list if any(bl.values())]

        # Contine if data are present
        if not block_list:
            if block.mandatory:
                LOGGER.error(f"No {block.title} metadata found")
            else:
                LOGGER.info(f"No {block.title} metadata found")
            return None
        else:
            LOGGER.info(f"Metadata for {block.title} found: {len(block_list)} records")

            if len(block_list) > 1 and block.singular:
                LOGGER.error("Only a single record should be present")

            # report on block fields
            for fld in mandatory_fields:
                fld_values = [rec[fld] for rec in block_list]
                if not all(fld_values):
                    LOGGER.error(f"Missing metadata in mandatory field {fld}")

            # report on actual data that is of the wrong type
            for fld in all_fields:
                bad_values = [
                    rec[fld]
                    for rec in block_list
                    if rec[fld] is not None
                    and not isinstance(rec[fld], field_types[fld])
                ]

                if bad_values:
                    LOGGER.error(
                        f"Field {fld} contains values of wrong type: ",
                        extra={"join": bad_values},
                    )

            # remap names if provided
            for old, new in field_map:
                for rec in block_list:
                    rec[new] = rec[old]
                    rec.pop(old)

            return block_list

    @loggerinfo_push_pop("Loading author metadata")
    def _load_authors(self):
        """Load the author block.

        Provides summary validation specific to the author block.
        """
        authors = self._read_block(self.fields["authors"])

        # Author specific validation
        if authors is not None:
            # Badly formatted names
            bad_names = [
                rec["name"]
                for rec in authors
                if isinstance(rec["name"], str) and not RE_NAME.match(rec["name"])
            ]
            if bad_names:
                LOGGER.error(
                    "Author names not formatted as last_name, first_names: ",
                    extra={"join": bad_names},
                )

            # Emails not formatted properly
            bad_emails = [
                rec["email"]
                for rec in authors
                if isinstance(rec["email"], str) and not RE_EMAIL.match(rec["email"])
            ]
            if bad_emails:
                LOGGER.error(
                    "Author emails not properly formatted: ", extra={"join": bad_emails}
                )

            # ORCIDs not strings
            bad_orcid = [
                rec["orcid"]
                for rec in authors
                if isinstance(rec["orcid"], str) and not RE_ORCID.match(rec["orcid"])
            ]
            if bad_orcid:
                LOGGER.error(
                    "Author ORCIDs not properly formatted: ", extra={"join": bad_orcid}
                )

        self.authors = authors

    @loggerinfo_push_pop("Loading keywords metadata")
    def _load_keywords(self):
        """Load the keywords block.

        Provides summary validation specific to the keywords block.
        """
        keywords = self._read_block(self.fields["keywords"])

        # extra data validation for keywords
        if keywords:
            keywords = [rec["keywords"] for rec in keywords]
            keywords = NoPunctuation(keywords)
            if not keywords:
                LOGGER.error(
                    "Put each keyword in a separate cell, do not separate "
                    "keywords using commas or semi-colons"
                )

            self.keywords = keywords.values

    @loggerinfo_push_pop("Loading permit metadata")
    def _load_permits(self):
        """Load the permits block.

        Provides summary validation specific to the permits block - users provide a
        permit authority, number and permit type.
        """

        permits = self._read_block(self.fields["permits"])

        # Permit specific checking for allowed permit types
        if permits:
            permit_types = [
                rec["type"].lower() for rec in permits if isinstance(rec["type"], str)
            ]
            valid_permit_types = {"research", "export", "ethics"}
            if not set(permit_types).issubset(valid_permit_types):
                LOGGER.error(
                    "Unknown permit types: ",
                    extra={"join": permit_types - valid_permit_types},
                )

        self.permits = permits

    @loggerinfo_push_pop("Loading DOI metadata")
    def _load_doi(self):
        """Load the DOI block.

        Provides summary validation specific to the DOI block.
        """
        # CHECK FOR PUBLICATION DOIs
        pub_doi = self._read_block(self.fields["doi"])

        # Extra data validation for DOIs
        if pub_doi is not None:
            # Check DOI URLS _are_ urls
            pub_doi_re = [
                RE_DOI.search(v["publication doi"])
                for v in pub_doi
                if isinstance(v["publication doi"], str)
            ]
            if not all(pub_doi_re):
                LOGGER.error("Publication DOIs not all in format: https://doi.org/...")

            if self.validate_doi:
                for is_doi in pub_doi_re:
                    if is_doi:
                        api_call = (
                            f"https://doi.org/api/handles/"
                            f"{is_doi.string[is_doi.end():]}"
                        )
                        r = requests.get(api_call)
                        if r.json()["responseCode"] != 1:
                            LOGGER.error(f"DOI not found: {is_doi.string}")

        self.publication_doi = pub_doi

    @loggerinfo_push_pop("Loading funding metadata")
    def _load_funders(self):
        """Load the funders block.

        Provides summary validation specific to the permits block - users provide a
        permit authority, number and permit type.
        """

        # LOOK FOR FUNDING DETAILS - users provide a funding body and a description
        # of the funding type and then optionally a reference number and a URL

        funders = self._read_block(self.fields["funding"])

        # TODO - currently no check beyond _read_block but maybe actually check
        #        the URL is a URL and maybe even opens? Could use urllib.parse

        self.funders = funders

    @loggerinfo_push_pop("Loading temporal extent metadata")
    def _load_temporal_extent(self):
        """Load the temporal extent block.

        Provides summary validation specific to temporal extents.
        """
        temp_extent = self._read_block(self.fields["date"])

        # temporal extent validation and updating
        if temp_extent is not None:
            start_date = temp_extent[0]["start date"]
            end_date = temp_extent[0]["end date"]

            if not (
                isinstance(start_date, datetime.datetime)
                and isinstance(end_date, datetime.datetime)
            ):
                LOGGER.error("Temporal extents are not date values")
                return

            if not (
                start_date.time() == datetime.time(0, 0)
                and end_date.time() == datetime.time(0, 0)
            ):
                LOGGER.error("Temporal extents should be date not datetime values")
                return

            if start_date > end_date:
                LOGGER.error("Start date is after end date")
                return

            self.temporal_extent.update([start_date.date(), end_date.date()])

    @loggerinfo_push_pop("Loading geographic extent metadata")
    def _load_geographic_extent(self):
        """Load the geographic extents block.

        Provides summary validation specific to geographic extents block.
        """
        geo_extent = self._read_block(self.fields["geo"])

        if geo_extent is not None:
            bbox = geo_extent[0]

            if all([isinstance(v, float) for v in bbox.values()]):
                if bbox["south"] > bbox["north"]:
                    LOGGER.error("South limit is greater than north limit")
                else:
                    self.latitudinal_extent.update([bbox["south"], bbox["north"]])

                if bbox["west"] > bbox["east"]:
                    LOGGER.error("West limit is greater than east limit")
                else:
                    self.longitudinal_extent.update([bbox["west"], bbox["east"]])

    @loggerinfo_push_pop("Loading external file metadata")
    def _load_external_files(self):
        """Load the external files block.

        Provides summary validation specific to the external files block. Small datasets
        will usually be contained entirely in a single Excel file, but where formatting
        or size issues require external files, then names and descriptions are included
        in the summary information.
        """

        external_files = self._read_block(self.fields["external"])

        # external file specific validation - no internal spaces.
        if external_files is not None:
            bad_names = [
                exf["file"]
                for exf in external_files
                if isinstance(exf["file"], str)
                and RE_CONTAINS_WSPACE.search(exf["file"])
            ]
            if any(bad_names):
                LOGGER.error(
                    "External file names must not contain whitespace: ",
                    extra={"join": bad_names},
                )

        self.external_files = external_files

    @loggerinfo_push_pop("Loading data worksheet metadata")
    def _load_data_worksheets(self, sheetnames):
        """Load the worksheets block.

        Provides summary validation specific to the worksheets block. The main things to
        be checked here are:

        1. Are there any standard worksheets incorrectly included in the data worksheets
           block?
        2. Are all the data worksheets present in the workbook documented?
        3. Do any worksheets linked to external files used documented external files?
        """

        # Load data worksheet data and convert an empty block from None to an empty list
        data_worksheets = self._read_block(self.fields["worksheet"])
        data_worksheets = [] if data_worksheets is None else data_worksheets

        # 1. Strip out faulty inclusion of standard worksheets
        cited_sheets = {ws["name"] for ws in data_worksheets}
        standard_sheets = {"Summary", "GBIFTaxa", "NCBITaxa", "Taxa", "Locations"}
        cited_standard_sheets = cited_sheets.intersection(standard_sheets)

        if cited_standard_sheets:
            LOGGER.error(
                "Do not include standard metadata sheets in data worksheet details: ",
                extra={"join": cited_standard_sheets},
            )

            data_worksheets = [
                ws for ws in data_worksheets if ws["name"] in standard_sheets
            ]

        # 2. Check for existing sheets without description
        extra_names = set(sheetnames) - standard_sheets - cited_sheets
        if extra_names:
            LOGGER.error(
                "Undocumented sheets found in workbook: ", extra={"join": extra_names}
            )

        # 3. Look to see what data is available:
        #    - No worksheets or external files: no data to document is an error.
        #    - Only external files: no tabular description of external files, just
        #      descriptions of the files themselves.
        #    - Named worksheets must exist and any external files linked must also
        #      exist.

        if not data_worksheets and self.external_files is None:
            LOGGER.error("No data worksheets or external files provided - no data.")
            return
        elif not data_worksheets:
            LOGGER.info("Only external file descriptions provided")
            return

        # Get external file names
        if self.external_files is not None:
            external_names = {ex["file"] for ex in self.external_files}
        else:
            external_names = set()

        # Check provided data worksheets
        for each_ws in data_worksheets:
            if each_ws["name"] not in sheetnames:
                # Unknown worksheet
                LOGGER.error(f"Data worksheet {each_ws['name']} not found")
            elif (
                each_ws["external"] is not None
                and each_ws["external"] not in external_names
            ):
                # Worksheet points to unknown external file
                LOGGER.error(
                    f"Data worksheet {each_ws['name']} linked to unknown "
                    f"external files: {each_ws['external']}",
                )
            else:
                LOGGER.info(f"Data worksheet  {each_ws['name']} found.")

        self.data_worksheets = data_worksheets

    @loggerinfo_push_pop("Loading access metadata")
    def _load_access_details(self):
        """Load the access block.

        Provides summary validation specific to the access block.
        """

        access = self._read_block(self.fields["access"])
        access = access[0]

        # Access specific validation - bad types handled by _read_block
        # - status must be in list of three accepted values
        if isinstance(access["access"], str):
            status = access["access"].lower()
            embargo_date = access["embargo_date"]

            if status not in ["open", "embargo", "restricted"]:
                LOGGER.error(
                    f"Access status must be Open, Embargo or Restricted not "
                    f"{access['access']}"
                )

            if status == "embargo":
                if embargo_date is None:
                    LOGGER.error("Dataset embargoed but no embargo date provided")
                elif isinstance(embargo_date, datetime.datetime):
                    now = datetime.datetime.now()

                    if embargo_date < now:
                        LOGGER.error("Embargo date is in the past.")
                    elif embargo_date > now + datetime.timedelta(days=2 * 365):
                        LOGGER.error("Embargo date more than two years in the future.")
                    else:
                        LOGGER.info(f"Dataset access: embargoed until {embargo_date}")

                if access["access_conditions"] is not None:
                    LOGGER.error("Access conditions cannot be set on embargoed data.")

            elif status == "restricted":
                access_conditions = access["access_conditions"]

                if embargo_date is not None:
                    LOGGER.error("Do not set an embargo date with restricted datasets")

                if access_conditions is None:
                    LOGGER.error(
                        "Dataset restricted but no access conditions specified"
                    )
                else:
                    LOGGER.info(
                        f"Dataset access: restricted with conditions "
                        f"{access_conditions}"
                    )
            else:
                LOGGER.info(f"Dataset access: {status}")

        self.access = access

    @loggerinfo_push_pop("Loading core metadata")
    def _load_core(self):
        """Load the core block.

        Provides summary validation specific to the core block.
        """

        core = self._read_block(self.fields["core"])

        # Guard against all rows being absent.
        if core is None:
            return

        self.title = core[0]["title"]
        self.description = core[0]["description"]

    @loggerinfo_push_pop("Loading project id metadata")
    def _load_project_ids(self):
        """Load the project ids block.

        Provides summary validation specific to the project ids block.
        """

        project_data = self._read_block(self.fields["project_ids"])

        if not self.projects and project_data is None:
            LOGGER.info("No project id data required or provided.")
            return

        # Bail if no validation should be done, warning if data provided.
        if not self.projects and project_data is not None:
            LOGGER.error("Project ids are not required but are provided.")
            return

        # Bail if validation should be done but no data provided
        if self.projects and project_data is None:
            LOGGER.error("Project ids are required but not provided.")
            return

        # Now try and validate what is found in the data
        bare_pids = [
            p["project id"] for p in project_data if p["project id"] is not None
        ]
        safe_pids = [
            p["safe project id"]
            for p in project_data
            if p["safe project id"] is not None
        ]

        if bare_pids and safe_pids:
            LOGGER.error(
                "Both 'project id' and 'safe project id' provided: "
                "use only 'project id'."
            )
            proj_ids = bare_pids + safe_pids
        elif safe_pids:
            LOGGER.warning(
                "Use 'project id' rather than the legacy 'safe project id' key."
            )
            proj_ids = safe_pids
        elif bare_pids:
            proj_ids = bare_pids

        # Check any provided values are valid
        invalid_proj_ids = [p for p in proj_ids if p not in self.projects]
        valid_proj_ids = [p for p in proj_ids if p in self.projects]
        if invalid_proj_ids:
            LOGGER.error(
                "Unknown project ids provided: ", extra={"join": invalid_proj_ids}
            )

        self.project_id = valid_proj_ids
        LOGGER.info("Valid project ids provided: ", extra={"join": valid_proj_ids})


def load_rows_from_worksheet(worksheet: Worksheet) -> list[tuple]:
    """Load worksheet rows, removing blank rows.

    Args:
        worksheet: An openpyxl worksheet instance.
    """
    # TODO - make 'internal' blank rows an error.
    rows = []
    for this_row in worksheet.iter_rows(values_only=True):
        if not all([blank_value(vl) for vl in this_row]):
            rows.append(this_row)

    return rows
