"""This module contains the Dataset, Dataworksheet and BaseField classes, along with
subclasses of BaseField for different data types. These are the core functions of
safedata_validator, responsible for opening a file containing formatted data,
and loading and validating the field metadata and data rows in the main data tables.
"""  # noqa D415

from __future__ import annotations

import datetime
import os
from itertools import islice
from logging import CRITICAL, ERROR, WARNING

import simplejson
from dateutil import parser
from openpyxl import load_workbook, worksheet
from openpyxl.utils import get_column_letter

from safedata_validator.extent import Extent
from safedata_validator.locations import Locations
from safedata_validator.logger import (
    FORMATTER,
    LOGGER,
    get_handler,
    loggerinfo_push_pop,
)
from safedata_validator.resources import Resources
from safedata_validator.summary import Summary
from safedata_validator.taxa import Taxa
from safedata_validator.validators import (
    RE_DMS,
    HasDuplicates,
    IsInSet,
    IsLocName,
    IsNotBlank,
    IsNotExcelError,
    IsNotNA,
    IsNotNumericString,
    IsNotPadded,
    IsNumber,
    IsString,
    blank_value,
    valid_r_name,
)

# These are lists, not sets because lists preserve order for preserving logging
# message order in unit testing.
MANDATORY_DESCRIPTORS: tuple[str, ...] = ("field_type", "description", "field_name")
OPTIONAL_DESCRIPTORS = [
    "levels",
    "method",
    "units",
    "taxon_name",
    "taxon_field",
    "interaction_name",
    "interaction_field",
    "file_container",
]


class Dataset:
    """Process the components of a safedata_validator formatted dataset.

    The Dataset class links a dataset with a particular configuration of the
    safedata_validator resources, and then loads the components of the dataset.

    Args:
        resources: An Resources instance configuring safedata_validator to process a
            dataset.
    """

    def __init__(self, resources: Resources | None = None) -> None:
        # Try and load the default resources if None provided
        if resources is None:
            resources = Resources()

        self.filename: str | None = None
        self.resources = resources
        self.summary = Summary(resources)
        self.taxa = Taxa(resources)
        self.dataworksheets: list[DataWorksheet] = []
        self.n_errors = 0
        self.passed = False

        # Extents - these can be loaded from the Summary or compiled from the
        # data, so the Summary and dataset extents are held separately so that
        # they can be validated against one another once all data is checked.

        self.temporal_extent = Extent(
            "temporal extent",
            (datetime.date,),
            hard_bounds=resources.extents.temporal_hard_extent,
            soft_bounds=resources.extents.temporal_soft_extent,
        )
        self.latitudinal_extent = Extent(
            "latitudinal extent",
            (float, int),
            hard_bounds=resources.extents.latitudinal_hard_extent,
            soft_bounds=resources.extents.latitudinal_soft_extent,
        )
        self.longitudinal_extent = Extent(
            "longitudinal extent",
            (float, int),
            hard_bounds=resources.extents.longitudinal_hard_extent,
            soft_bounds=resources.extents.longitudinal_soft_extent,
        )

        # Create Locations and pass in dataset extents
        self.locations = Locations(
            resources,
            latitudinal_extent=self.latitudinal_extent,
            longitudinal_extent=self.longitudinal_extent,
        )

    def load_from_workbook(
        self,
        filename: str,
        validate_doi: bool = False,
        chunk_size: int = 1000,
        console_log: bool = True,
    ) -> None:
        """Populate a Dataset instance from a safedata_validator formatted Excel file.

        This method populates a Dataset using the safedata_validator format for Excel
        workbooks in .xlsx format.

        Logging messages are used to report the progress of the validation process.
        These are always captured in an internal log (`logging.LOG`) but are also
        printed by default to the command line. The `console_log` argument can be used
        to suppress console logging, when this command is being run programmatically
        rather than via the `safedata_validate` command line script.

        Args:
            filename: A path to the workbook containing the dataset
            validate_doi: Should DOIs in the dataset summary be validated
            chunk_size: Data is read from worksheets in chunks of rows - this
                argument sets the size of that chunk.
            console_log: Suppress command line logging.
        """

        # Handle logging details - flush and reset from previous runs.
        handler = get_handler()
        handler.reset()

        if console_log:
            handler.setLevel("DEBUG")
        else:
            handler.setLevel("CRITICAL")

        # Open the workbook with:
        #  - read_only to use the memory optimised read_only implementation.
        #    This is a bit restricted as it only exposes row by row iteration
        #    to avoid expensive XML traversal but has a lower memory footprint.
        #  - data_only to load values not formulae for equations.
        self.filename = os.path.basename(filename)
        wb = load_workbook(filename, read_only=True, data_only=True)

        # Populate summary
        if "Summary" in wb.sheetnames:
            self.summary.load(
                wb["Summary"],
                wb.sheetnames,
                validate_doi=validate_doi,
            )
        else:
            # No summary is impossible - so an error and no dataworksheets
            # will be loaded, but taxon and locations could be checked if present
            LOGGER.error("No summary worksheet found - moving on")

        # Populate locations
        if "Locations" in wb.sheetnames:
            self.locations.load(wb["Locations"])
        else:
            # No locations is pretty implausible - lab experiments?
            LOGGER.warning("No locations worksheet found - moving on")

        # Throw an error if both Taxa and GBIFTaxa have been given as worksheet names
        gbif_sheets = {"GBIFTaxa", "Taxa"}.intersection(wb.sheetnames)

        if len(gbif_sheets) == 2:
            LOGGER.error(
                "Both Taxa and GBIFTaxa provided as sheet names, this "
                "is not allowed! Only checking GBIFTaxa sheet"
            )
            self.taxa.gbif_taxa.load(wb["GBIFTaxa"])
        # Otherwise populate gbif_taxa from the one that has been provided
        elif len(gbif_sheets) == 1:
            self.taxa.gbif_taxa.load(wb[next(iter(gbif_sheets))])

        # Populate ncbi taxa
        ncbi_sheet = "NCBITaxa" in wb.sheetnames
        if ncbi_sheet:
            self.taxa.ncbi_taxa.load(wb["NCBITaxa"])

        if not gbif_sheets and not ncbi_sheet:
            # Leave the default empty Taxa object
            LOGGER.warning("Neither Taxa worksheet found - moving on")

        # Load data worksheets
        for sheet_meta in self.summary.data_worksheets:
            sheet_name = sheet_meta["name"]

            if sheet_name in wb.sheetnames:
                dwsh = DataWorksheet(sheet_meta, self)
                dwsh.load_from_worksheet(wb[sheet_name], row_chunk_size=chunk_size)
                self.dataworksheets.append(dwsh)

        # Run final checks
        self.final_checks()

    @loggerinfo_push_pop("Running final checking")
    def final_checks(self) -> None:
        """Run final validation checks on a Dataset instance.

        This method checks that:
        1. all locations and taxa provided have been used in the data worksheets scanned
        2. no worksheet taxon names are duplicated between GBIFTaxa and NCBITaxa.
        3. extents in the data are congruent with the summary extents

        Finally, the method reports the total number of errors and warnings from
        processing the dataset.
        """

        LOGGER.info("Checking provided locations and taxa all used in data worksheets")
        FORMATTER.push()

        # check locations and taxa
        # - Can't validate when there are external files which might use any unused ones
        # - If the sets are not the same, then the worksheet reports will catch
        #   undocumented ones so here only look for ones that are provided but not used
        #   in the data.
        if not self.locations.is_empty:
            if self.summary.external_files:
                LOGGER.warning(
                    "Location list cannot be validated when external data files are"
                    " used"
                )
            elif self.locations.locations_used == self.locations.locations:
                LOGGER.info("Provided locations all used in datasets")
            elif self.locations.locations_used == (
                self.locations.locations & self.locations.locations_used
            ):
                LOGGER.error(
                    "Provided locations not used: ",
                    extra={
                        "join": self.locations.locations - self.locations.locations_used
                    },
                )

        # check taxa
        if not self.taxa.is_empty:
            if self.summary.external_files:
                LOGGER.warning(
                    "Taxon list cannot be validated when external data files are used"
                )
            else:
                if self.taxa.taxon_names_used == self.taxa.taxon_names:
                    LOGGER.info("Provided taxa all used in datasets")
                elif self.taxa.taxon_names_used != self.taxa.taxon_names.union(
                    self.taxa.taxon_names_used
                ):
                    LOGGER.error(
                        "Provided taxa not used: ",
                        extra={
                            "join": self.taxa.taxon_names - self.taxa.taxon_names_used
                        },
                    )

                if self.taxa.repeat_names != set():
                    LOGGER.error(
                        "The following taxa are defined in both GBIFTaxa and "
                        "NCBITaxa: ",
                        extra={"join": self.taxa.repeat_names},
                    )

        # Check the extents - there are both summary and dataset extents so
        # check at least one is populated for each extent and that they are
        # compatible
        FORMATTER.pop()
        LOGGER.info("Checking temporal and geographic extents")
        FORMATTER.push()

        extents_to_check = (
            ("Temporal", "temporal_extent"),
            ("Latitudinal", "latitudinal_extent"),
            ("Longitudinal", "longitudinal_extent"),
        )

        for label, this_extent in extents_to_check:
            dataset_extent = getattr(self, this_extent)
            summary_extent = getattr(self.summary, this_extent)

            # If neither: need to provide in summary. If both: consistent.
            if not (dataset_extent.populated or summary_extent.populated):
                LOGGER.error(
                    f"{label} extent not set from data or provided in summary: "
                    "add extents to dataset Summary"
                )
            elif (dataset_extent.populated and summary_extent.populated) and (
                (dataset_extent.extent[0] < summary_extent.extent[0])
                or (dataset_extent.extent[1] > summary_extent.extent[1])
            ):
                LOGGER.error(
                    f"{label} extent values from the data fall outside the extents "
                    "set in the Summary sheet "
                    f"({[str(x) for x in dataset_extent.extent]})"
                )
            elif dataset_extent.populated and summary_extent.populated:
                LOGGER.warning(
                    f"The {label} extent is set in Summary but also "
                    "is populated from the data - this may be deliberate!"
                )
            elif dataset_extent.populated and not summary_extent.populated:
                summary_extent = dataset_extent
                LOGGER.info(
                    f"The {label} extent in Summary is not set and has "
                    "been populated from the data"
                )

        # Dedent for final result
        FORMATTER.pop(n=2)
        handler = get_handler()

        if handler.counters["ERROR"] > 0:
            self.n_errors = handler.counters["ERROR"]
            if handler.counters["WARNING"] > 0:
                LOGGER.info(
                    f"FAIL: file contained {handler.counters['ERROR']} errors "
                    f"and {handler.counters['WARNING']} warnings"
                )
            else:
                LOGGER.info(f"FAIL: file contained {handler.counters['ERROR']} errors")
        else:
            self.passed = True

            if handler.counters["WARNING"] > 0:
                LOGGER.info(
                    "PASS: file formatted correctly but with "
                    f"{handler.counters['WARNING']} warnings"
                )
            else:
                LOGGER.info("PASS: file formatted correctly with no warnings")

    def to_json(self):
        """Export a Dataset to JSON.

        This method exports key data about the dataset in JSON format. This
        method is used to export a description of a dataset that can be used
        to populate a metadata server and publish datasets to Zenodo.
        """

        # TODO - continue with class.to_dict() methods?

        json_dict = dict(
            # Summary information
            title=self.summary.title,
            description=self.summary.description,
            authors=self.summary.authors,
            filename=self.filename,
            external_files=self.summary.external_files,
            access=self.summary.access["access"],
            embargo_date=self.summary.access["embargo_date"],
            access_conditions=self.summary.access["access_conditions"],
            funders=self.summary.funders,
            permits=self.summary.permits,
            keywords=self.summary.keywords,
            dataworksheets=[dwsh.to_dict() for dwsh in self.dataworksheets],
            # Taxa
            # Gbif taxa if they exist
            gbif_timestamp=self.resources.gbif_timestamp,
            gbif_taxa=[
                dict(
                    zip(
                        (
                            "worksheet_name",
                            "taxon_id",
                            "parent_id",
                            "taxon_name",
                            "taxon_rank",
                            "taxon_status",
                        ),
                        tx,
                    )
                )
                for tx in self.taxa.gbif_taxa.taxon_index
            ],
            # NCBI taxa if they exist
            ncbi_timestamp=self.resources.ncbi_timestamp,
            ncbi_taxa=[
                dict(
                    zip(
                        (
                            "worksheet_name",
                            "taxon_id",
                            "parent_id",
                            "taxon_name",
                            "taxon_rank",
                            "taxon_status",
                        ),
                        tx,
                    )
                )
                for tx in self.taxa.ncbi_taxa.taxon_index
            ],
            # Locations
            locations=[
                dict(zip(("name", "new_location", "wkt_wgs84"), lc))
                for lc in self.locations.location_index
            ],
            # Publication details - these are populated by the
            # Zenodo publication mechanism.
            zenodo_concept_id=None,
            zenodo_record_id=None,
            zenodo_publication_date=None,
        )

        # Extents - summary take priority over dataset.
        for ext in ("temporal_extent", "latitudinal_extent", "longitudinal_extent"):
            summary_extent = getattr(self.summary, ext)
            data_extent = getattr(self, ext)
            if summary_extent.populated:
                json_dict[ext] = summary_extent.extent
            else:
                json_dict[ext] = data_extent.extent

        return simplejson.dumps(json_dict, default=str, indent=2)

    @property
    def error_breakdown(self):
        """Get error counts broken down by dataset component."""
        return {
            "total": self.n_errors,
            "summary": self.summary.n_errors,
            "locations": self.locations.n_errors,
            "gbif": self.taxa.gbif_taxa.n_errors,
            "ncbi": self.taxa.ncbi_taxa.n_errors,
            "data": [(d.name, d.n_errors) for d in self.dataworksheets],
        }


class DataWorksheet:
    """Process the contents of safedata_validator formatted data table.

    This class is used to load and check the formatting and content of a data table
    using the safedata_validator format. It requires the containing Dataset as an
    argument to get access to Summary, Locations, Taxa and Extents attributes and the
    specific data table metadata from the Summary instance.

    The workflow for methods validating a data table is:

    1. Create a DataWorksheet instance using the summary metadata.
    2. Parse the rows at the top of the file containing field metadata and use
       `validate_field_meta` to validate the contents.
    3. Parse the data rows below the field metadata using the `validate_data_rows`
       method. This can be done repeatedly to handle chunked inputs.
    4. Use the report method to obtain all of the logging associated with data
       validation and run any final checks.

    This workflow is currently implemented for Excel format data worksheets in the
    `load_from_worksheet` method.

    Args:
        sheet_meta: The metadata dictionary for this worksheet from
            the Summary instance, containing the worksheet name, title and description
            and the names of any external files to be associated with this worksheet.
        dataset: A safedata_validator Dataset object, providing Taxa, Locations and
            Summary information.
    """

    def __init__(
        self,
        sheet_meta: dict,
        dataset: Dataset | None = None,
    ) -> None:
        # Set initial values

        # TODO - checks on sheetmeta
        self.name = sheet_meta["name"]
        self.description = sheet_meta["description"]
        self.title = sheet_meta["title"]
        # For sheet meta with an external file, add in the external file name
        self.external = sheet_meta.get("external")

        # Links to dataset for dataset level information, and taxa and
        # location validation lists.
        self.dataset = dataset

        # Create the field meta attributes, populated using validate_field_meta
        self.fields_loaded = False
        self.field_meta: list[dict] = []
        self.field_types = None
        self.n_fields: int = 0
        self.n_trailing_empty_fields: int = 0
        self.n_descriptors: int
        self.descriptors: list = []
        self.fields: list[EmptyField | BaseField] = []
        self.taxa_fields: list = []

        # Keep track of row numbering
        self.n_row = 0
        self.current_row = 1
        self.row_numbers_sequential = True
        self.row_numbers_missing = False
        self.row_numbers_noninteger = False
        self.blank_rows = False
        self.trailing_blank_rows = False
        self.start_errors = get_handler().counters["ERROR"]
        self.n_errors = 0

    @loggerinfo_push_pop("Validating field metadata")
    def validate_field_meta(self, field_meta: dict) -> None:
        """Validate the field metadata for a data table.

        This method is used to add and validate field metadata. The field meta can
        include trailing empty fields - these will be checked when data rows are added.
        The method also then sets up a list of field validators, matching the order of
        the fields in the table, which are used for data validation.

        Note:
            This relies on the maintenance of the dictionary insertion order, which
            required OrderedDict in the past but is guaranteed for the standard dict in
            Python >= 3.7.
        """

        # Clean off descriptor whitespace padding?
        clean_descriptors = IsNotPadded(field_meta.keys())
        if not clean_descriptors:
            # Report whitespace padding and clean up tuples
            LOGGER.error(
                "Whitespace padding in descriptor names: ",
                extra={"join": clean_descriptors.failed},
            )

            # Order preserved in dict and validator
            cleaned_entries = [
                (ky, val) for ky, val in zip(clean_descriptors, field_meta.values())
            ]
            field_meta = dict(cleaned_entries)

        self.descriptors = list(field_meta.keys())
        self.n_descriptors = len(self.descriptors)

        # Checking field_meta structure - should be an equal length and non-zero tuple
        # of values for each descriptor. If not, return without setting fields_loaded.
        descriptor_tuple_lengths = {len(tp) for tp in field_meta.values()}
        if len(descriptor_tuple_lengths) > 1 or descriptor_tuple_lengths == {0}:
            LOGGER.error("Cannot load unequal length or empty field metadata")
            return

        # * Expected descriptors - do _not_ preclude user defined descriptors
        #   but warn about them to alert to typos. Missing descriptors vary
        #   with field type so are handled in BaseField.__init__()
        unknown_descriptors = set(field_meta.keys()).difference(
            set(MANDATORY_DESCRIPTORS).union(OPTIONAL_DESCRIPTORS)
        )
        if unknown_descriptors:
            LOGGER.warning(
                "Unknown field descriptors:", extra={"join": unknown_descriptors}
            )

        # * Check for field name as _last_ descriptor. This is primarily to
        #   make it easy to read the dataframe - just skip the other descriptors
        if list(field_meta.keys())[-1] != "field_name":
            LOGGER.error("Field_name row is not the last descriptor")

        # Repackage field metadata into a list of per field descriptor
        # dictionaries, _importantly_ preserving the column order.
        field_meta_list = [
            dict(zip(field_meta.keys(), val)) for val in zip(*field_meta.values())
        ]

        # Check provided field names are unique. This doesn't cause as many
        # problems as duplications in Taxa and Locations, which expect certain
        # fields, so warn and continue. Ignore empty mandatory values - these
        # are handled in field checking.
        field_names = [
            fld["field_name"]
            for fld in field_meta_list
            if fld["field_name"] is not None
        ]
        dupes = HasDuplicates(field_names)
        if dupes:
            LOGGER.error("Field names duplicated: ", extra={"join": dupes.duplicated})

        # Lowercase the field types
        for fld in field_meta_list:
            fld["field_type"] = (
                None if fld["field_type"] is None else fld["field_type"].lower()
            )

        # Populate the instance variables
        self.fields_loaded = True
        self.field_meta = field_meta_list
        self.n_fields = len(field_meta_list)
        self.n_trailing_empty_fields = 0

        # get taxa field names for cross checking observation and trait data
        self.taxa_fields = [
            fld["field_name"] for fld in self.field_meta if fld["field_type"] == "taxa"
        ]

        # Now initialise the Field objects using the mapping of field types
        # to the BaseField subclasses. This needs to handle trailing fields
        # containing _no_ metadata (e.g. from inaccurate Excel sheet bounds or
        # CSV comments) but also fields with missing field_type (but other
        # metadata).
        #
        # - trailing empty fields are assumed to be blank columns
        #   and validate_data_rows should test for content.
        # - otherwise, these are assumed to be fields and should
        #   trigger warnings about descriptors

        # Get logical flags for empty and trailing empty metadata
        field_meta_empty = [set(vl.values()) == {None} for vl in self.field_meta]
        trailing_empty = [
            all(field_meta_empty[-(n + 1) :]) for n in range(0, len(field_meta_empty))
        ]
        trailing_empty.reverse()
        self.n_trailing_empty_fields = sum(trailing_empty)

        # Get the type map and field list
        field_subclass_map = BaseField.field_type_map()
        unknown_field_types = set()

        for col_idx, (tr_empty, fd_empty, fmeta) in enumerate(
            zip(trailing_empty, field_meta_empty, self.field_meta)
        ):
            fmeta["col_idx"] = col_idx + 1

            # Consider cases
            if tr_empty:
                # Hopefully empty trailing field
                self.fields.append(EmptyField(fmeta))
            elif fd_empty:
                # Empty field within other fields
                self.fields.append(BaseField(fmeta, dwsh=self, dataset=self.dataset))
            elif fmeta["field_type"] is None:
                # Non-empty field of None type - use BaseField for basic checks
                self.fields.append(BaseField(fmeta, dwsh=self, dataset=self.dataset))
            elif fmeta["field_type"] not in field_subclass_map:
                #  Non-empty field of unknown type - use BaseField for basic checks
                unknown_field_types.add(fmeta["field_type"])
                self.fields.append(BaseField(fmeta, dwsh=self, dataset=self.dataset))
            else:
                # Known field type.
                fld_class = field_subclass_map[fmeta["field_type"]]
                self.fields.append(fld_class(fmeta, dwsh=self, dataset=self.dataset))

        if unknown_field_types:
            LOGGER.error("Unknown field types: ", extra={"join": unknown_field_types})

    def validate_data_rows(self, data_rows: list) -> None:
        """Validate the contents of data rows from a data table.

        This method is used to pass a list of rows of data from a data table into the
        appropriate field checkers created by the validate_field_meta method. Each row
        is represented as a list, with the row number as the first entry in the list.

        Args:
            data_rows: A list of rows containing data from a data table.
        """

        if not self.fields_loaded:
            LOGGER.critical("No fields defined - use validate_field_meta")
            return

        # Check the lengths of the rows - this should log critical
        # because it is likely to be a programming error, not a user error.
        if len(data_rows) == 0:
            LOGGER.critical("Empty data_rows passed to validate_data_rows")
            return

        row_lengths = {len(rw) for rw in data_rows}

        if len(row_lengths) != 1:
            LOGGER.critical("Data rows of unequal length - cannot validate")
            return
        elif row_lengths.pop() != (self.n_fields + 1):
            LOGGER.critical(
                "Data rows not of same length as field metadata - cannot validate"
            )
            return

        # Handle empty rows.
        blank_set = {None}
        blank_row = [set(vals) == blank_set for vals in data_rows]

        trailing_blank_row = [
            all(blank_row[-(n + 1) :]) for n in range(0, len(blank_row))
        ]
        trailing_blank_row.reverse()

        internal_blank_row = [
            blnk & ~trail for blnk, trail in zip(blank_row, trailing_blank_row)
        ]

        # Internals within a set of row are automatically bad, but trailing blank
        # rows are only bad if more data is loaded below them.
        if any(internal_blank_row):
            self.blank_rows = True

        if not self.trailing_blank_rows and any(trailing_blank_row):
            self.trailing_blank_rows = True
        elif self.trailing_blank_rows and not all(trailing_blank_row):
            self.blank_rows = True

        # Drop blank rows for further processing
        data_rows = [drw for drw, is_blank in zip(data_rows, blank_row) if not is_blank]

        # Check there is any remaining data, then convert the values into columns
        # and extract the row numbers
        n_row = len(data_rows)
        if n_row == 0:
            return
        else:
            self.n_row += n_row

        data_cols = list(zip(*data_rows))
        row_numbers = data_cols.pop(0)

        # Check for bad values (blanks, non integers) in row numbers
        valid_row_numbers = IsNotBlank(row_numbers, keep_failed=False)

        if not valid_row_numbers:
            self.row_numbers_missing = True

        row_numbers = valid_row_numbers.values

        if any([not isinstance(vl, int) for vl in row_numbers]):
            self.row_numbers_noninteger = True

        # Check the row numbering - skip if this has already failed or
        # the row numbers have already included None or non-integers
        if self.row_numbers_sequential and not (
            self.row_numbers_noninteger or self.row_numbers_missing
        ):
            expected_numbers = list(
                range(self.current_row, self.current_row + len(row_numbers))
            )
            self.current_row += len(row_numbers)

            if row_numbers != expected_numbers:
                self.row_numbers_sequential = False
        else:
            self.current_row += len(row_numbers)

        # Now feed the sets of values into the Field validation
        for data, field_inst in zip(data_cols, self.fields):
            field_inst.validate_data(list(data))

    def report(self) -> None:
        """Report data validation for a data table.

        As the validate_data_rows method works on chunks of rows, the DataWorksheet
        instance accumulates logging messages across individual fields. These are not
        reported until after all the data rows have been validated, when aggregated
        reports for each field can be made. This method causes these field logs to be
        emitted and then carries out final checks and reporting across all of the field
        data.
        """

        if not self.n_row:
            if self.external is None:
                LOGGER.error("No data passed for validation.")
            else:
                LOGGER.info(
                    "Data table description associated with "
                    f"external file {self.external}"
                )
        else:
            LOGGER.info("Validating field data")
            FORMATTER.push()

            for fld in self.fields:
                fld.report()
            FORMATTER.pop()

        # Report on row numbering
        if self.row_numbers_missing:
            LOGGER.error("Missing row numbers in data")

        if self.row_numbers_noninteger:
            LOGGER.error("Row numbers contain non-integer values")

        if (
            not (self.row_numbers_noninteger or self.row_numbers_missing)
            and not self.row_numbers_sequential
        ):
            LOGGER.error("Row numbers not consecutive or do not start with 1")

        # Internal blank rows?
        if self.blank_rows:
            LOGGER.error("Data contains empty rows")

        # report on detected size
        LOGGER.info(
            f"Worksheet '{self.name}' contains {self.n_descriptors} descriptors, "
            f"{self.n_row} data rows and "
            f"{self.n_fields - self.n_trailing_empty_fields} fields"
        )

        # reporting
        self.n_errors = get_handler().counters["ERROR"] - self.start_errors
        if self.n_errors > 0:
            LOGGER.info(f"Dataframe contains {self.n_errors} errors")
        else:
            LOGGER.info("Dataframe formatted correctly")

    def load_from_worksheet(
        self, worksheet: worksheet, row_chunk_size: int = 1000
    ) -> None:
        """Populate a Dataworksheet instance from an openpyxl worksheet.

        This method takes a newly initialised DataWorksheet instance and populates the
        details using the contents of a safedata_validator formatted data table in an
        Excel spreadsheet.

        To keep memory requirements low, data in the worksheet is validated by loading
        sets of rows containing at most `row_chunk_size` rows. Note that this will only
        actually reduce memory use for openpyxl ReadOnlyWorksheets, as these load data
        on demand, but will still work on standard Worksheets.

        Args:
            worksheet: An openpyxl worksheet instance
            row_chunk_size: The number of rows of data to load in each chunk
        """

        # Not using @loggerinfo_push_pop to provide access to WS name
        LOGGER.info(f"Checking worksheet '{self.name}'")
        FORMATTER.push()

        if self.fields_loaded:
            LOGGER.critical("Field metadata already loaded - use fresh instance.")
            FORMATTER.pop()
            return

        # get the data dimensions
        max_row = worksheet.max_row

        # trap completely empty worksheets
        if max_row == 0:
            LOGGER.error("Worksheet is empty")
            FORMATTER.pop()
            return

        # Read the field metadata first:
        # - There should be rows starting with a string and then rows starting with
        #   a number, or a StopIteration at the end of the row iterator.
        ws_rows = worksheet.iter_rows(values_only=True)
        field_meta = []

        # While there are rows and while they start with a string cell, collect
        # field metadata
        for row in ws_rows:
            if isinstance(row[0], str):
                field_meta.append(row)
            else:
                break

        # Convert field meta to dict and validate
        field_meta_dict = {rw[0]: rw[1:] for rw in field_meta}
        self.validate_field_meta(field_meta_dict)

        if self.fields_loaded:
            # Get an iterator on the data rows
            data_rows = worksheet.iter_rows(
                min_row=len(field_meta_dict) + 1, values_only=True
            )

            # Load and validate chunks of data
            n_chunks = ((max_row - self.n_descriptors) // row_chunk_size) + 1
            for chunk in range(n_chunks):
                # itertools.islice handles generators and StopIteration, and also
                # trap empty slices
                data = list(islice(data_rows, row_chunk_size))
                if data:
                    self.validate_data_rows(data)

        # Finish up
        self.report()
        FORMATTER.pop()

    def to_dict(self) -> dict:
        """Return a dictionary representation of a DataWorksheet instance."""

        # Update the fields to account for EmptyFields that have been validated: they
        # have no field metadata and there is no content in the rows below. EmptyFields
        # are always trailing empty fields.

        fields = [
            field_to_dict(fld, idx + 1)
            for idx, fld in enumerate(self.fields)
            if not isinstance(fld, EmptyField)
        ]

        output = dict(
            taxa_fields=self.taxa_fields,
            max_row=self.n_row + self.n_descriptors,
            max_col=self.n_fields - self.n_trailing_empty_fields,
            name=self.name,
            title=self.title,
            description=self.description,
            descriptors=self.descriptors,
            external=self.external,
            fields=fields,
            field_name_row=self.n_descriptors,
            n_data_row=self.n_row,
        )

        return output


class BaseField:
    """Validating metadata and data for a field.

    This class provides an implementation for checking field metadata for a field type
    and then testing whether the data contents are compatible with that metadata. The
    class is largely used for checking fields in data tables worksheets, but can also
    be used for checking other tabular data.

    The base class defines the core checking methods and then subclasses are used to
    extend those methods to implement specific checking for particular data types. The
    main workflow for the class is:

    * Instantiate a class object using the field metadata to check the metadata.
    * The base class provides `_check_meta`, `_check_taxon_meta` and
      `_check_interaction_meta` to provide common shared components of metadata
      validation in extended `__init__` methods in subclasses.
    * Data is then validated using the `validate_data` method - it ingests chunks of
      rows, to avoid having to read all the data in a table from file at once. The data
      validation process accumulates logging messages in a stack via the `_log` method.
    * The `report` method is then used to run final checks and emit the message stack,
      once all of the data has been ingested.

    The base class also has a much more complex signature than the base
    functionality requires. Various subclasses need access to:

    * Dataset level information - extents, taxa and locations.
    * Dataworksheet level information - taxon fields

    Rather than having complex inheritance with changing subclass signatures
    and kwargs, the BaseField class makes all information available to all
    subclasses.

    Args:
        meta: A dictionary of field metadata
        dwsh: A DataWorksheet instance, used to pass worksheet level
            information, which at the moment is just taxa fields.
        dataset: An Dataset instance, used to provide dataset level
            information, such as taxa, locations and summary, and update
            dataset level attributes such as data extents.

    Class Attributes:
        field_types: A list of the field types - as entered in data table metadata -
            that are handled by a class based on BaseField.
        required_descriptors: A list of field metadata names that are required by a
            class based on BaseField.
        check_taxon_meta: A boolean flag indicating whether a class based on BaseField
            should check taxonomic metadata.
        check_interaction_meta: A similar boolean flag for interaction metadata.

    Attributes:
        meta: The metadata dictionary for the field
        dwsh: The DataWorksheet object in which the field is contained.
        taxa: The Taxa object for the Dataset or None.
        locations: The Locations object for the Dataset or None.
        summary: The Summary object for the Dataset or None.
        log_stack: A list of tuples containing logging messages to be emitted.
        n_rows: The number of rows of data in the field.
        n_na: The number of NA values in the field data.
        n_blank: The number of blank cells in the field data.
        n_excel_error: The number of Excel error codes in the field data.
        bad_values: A list of invalid value given a class based on BaseField.
        bad_rows: A list of row numbers containing bad values.
        field_name: The provided field name.

    """

    # These class attributes describe the field_types handled by the class and sets the
    # required descriptors for those field_types. The no_validation attributes is used
    # to suppress data validation for use on e.g. comments fields.

    field_types: tuple[str, ...] | None = None
    required_descriptors = MANDATORY_DESCRIPTORS
    check_taxon_meta = False
    check_interaction_meta = False

    def __init__(
        self,
        meta: dict,
        dwsh: DataWorksheet | None = None,
        dataset: Dataset | None = None,
    ) -> None:
        self.meta = meta
        self.dwsh = dwsh

        self.taxa: Taxa | None = None
        self.locations = None
        self.summary = None

        # Shortcuts to main components
        self.dataset = dataset
        if self.dataset:
            self.taxa = getattr(self.dataset, "taxa")
            self.locations = getattr(self.dataset, "locations")
            self.summary = getattr(self.dataset, "summary")

        # Attributes
        self.log_stack: list = []
        self.n_rows = 0
        self.n_na = 0
        self.n_blank = 0
        self.n_excel_error = 0
        self.bad_values: list = []
        self.bad_rows: list = []  # TODO check on implementation of this

        # Get a field name - either from the field_meta, or a column letter
        # from col_idx or 'Unknown'
        self.field_name = self.meta.get("field_name")

        if self.field_name is None:
            idx = self.meta.get("col_idx")
            if idx is None:
                self.field_name = "Unknown"
            else:
                self.field_name = f"Column_{get_column_letter(idx)}"

        # Now check required descriptors - values are present, non-blank and
        # are unpadded strings (currently all descriptors are strings).
        for dsc in self.required_descriptors:
            self._check_meta(dsc)

        # Specific check that field name is valid - the column letter codes are
        # always valid, so missing names won't trigger this.
        # _check_meta has already tested and logged non-string errors, so
        # only test here for strings that fail.
        if isinstance(self.field_name, str) and not valid_r_name(self.field_name):
            self._log(
                f"Field name is not valid: {self.field_name!r}. "
                "Common errors are spaces and non-alphanumeric "
                "characters other than underscore and full stop"
            )

        # TODO - rethink implementation? Quite specific behaviour unique to
        #        some fields, so could implement as overloaded extra stub.
        if self.check_taxon_meta:
            self._check_taxon_meta()

        if self.check_interaction_meta:
            self._check_interaction_meta()

    def _log(self, msg: str, level: int = ERROR, extra: dict | None = None) -> None:
        """Adds messages to the field log stack.

        Rather than directly emitting a log message, field processing accumulates a
        stack of messages, which are all emitted when the
        [safedata_validator.field.BaseField.report][] method is called.

        Args:
            msg: The log message to add
            level: The logging level to use.
            extra: A dictionary of extra information to pass to the logging.log method.
        """
        self.log_stack.append((level, msg, extra))

    def _check_meta(self, descriptor: str) -> bool:
        """Check required field metadata.

        A standardised checker to see if a required descriptor is present for a field
        and that it isn't simply empty or whitespace. Messages are added to the
        log_stack if required. It returns a boolean showing whether checks pass or not -
        the actual causes of failure are added to the field message log.

        Args:
            descriptor: The name of the descriptor to check.
        """

        if descriptor not in self.meta:
            self._log(f"{descriptor} descriptor missing")
            return False

        val = self.meta[descriptor]

        if blank_value(val):
            self._log(f"{descriptor} descriptor is blank")
            self.meta[descriptor] = None  # standardise whitestring to None
            return False
        elif not isinstance(val, str):
            self._log(f"{descriptor} descriptor is not a string: {val!r}")
            return False
        elif val != val.strip():
            self._log(f"{descriptor} descriptor has whitespace padding: {val!r}")
            self.meta[descriptor] = val.strip()
            return False
        else:
            return True

    def _check_taxon_meta(self) -> bool:
        """Check taxon field metadata.

        Checks the taxonomic metadata of abundance and trait fields. This is more
        involved that the simple _check_meta(), because of the option to provide
        taxon_name or taxon_field descriptors. It returns a boolean showing whether
        checks pass or not - the actual causes of failure are added to the field message
        log.
        """

        # Are taxon_name and taxon_field provided and not blank: note use of
        # 'and' rather than '&' to allow missing descriptors to short cut

        tx_nm = self.meta.get("taxon_name")
        tx_fd = self.meta.get("taxon_field")
        tx_nm_prov = not blank_value(tx_nm)
        tx_fd_prov = not blank_value(tx_fd)

        if tx_nm_prov and tx_fd_prov:
            self._log("Taxon name and taxon field both provided, use one only")
            return False
        elif not tx_nm_prov and not tx_fd_prov:
            self._log("One of taxon name or taxon field must be provided")
            return False
        elif tx_nm_prov:
            if self.taxa is None:
                self._log(
                    "Taxon name provided but no Taxa instance available", CRITICAL
                )
                return False
            elif self.taxa.is_empty:
                self._log("Taxon name provided but no taxa loaded")
                return False
            elif tx_nm not in self.taxa.taxon_names:
                self._log("Taxon name not found in the Taxa worksheet")
                return False
            else:
                self.taxa.taxon_names_used.add(tx_nm)
                return True
        elif tx_fd_prov:
            if self.dwsh is None:
                self._log(
                    f"Taxon field provided but no dataworksheet provided for this "
                    f"field: {tx_fd}",
                    CRITICAL,
                )
                return False
            elif tx_fd not in self.dwsh.taxa_fields:
                self._log(f"Taxon field not found in this worksheet: {tx_fd}")
                return False
            else:
                return True
        else:
            return True

    def _check_interaction_meta(self) -> bool:
        """Check interaction metadata.

        Checks the taxonomic metadata of interaction fields. This is more involved that
        the simple _check_meta(), because of the option to provide taxon_name or
        taxon_field descriptors describing at least two taxonomic identifiers. It
        returns a boolean showing whether checks pass or not - the actual causes of
        failure are added to the field message log.
        """

        # TODO - currently no checking for descriptions being present in
        #        the interaction information. Not sure how many old datasets
        #        would be affected, but this would be good to include.

        # Are interaction_name and/or interaction_field provided and not blank
        iact_nm = self.meta.get("interaction_name")
        iact_fd = self.meta.get("interaction_field")
        iact_nm_prov = not blank_value(iact_nm)
        iact_fd_prov = not blank_value(iact_fd)

        if not iact_nm_prov and not iact_fd_prov:
            self._log(
                "At least one of interaction name or interaction field must be provided"
            )
            return False

        if iact_nm_prov:
            # Check that self.taxa has actually been populated
            if self.taxa is None:
                self._log(
                    "Interaction name provided but no Taxa instance available", CRITICAL
                )
                return False
            elif self.taxa.is_empty:
                self._log("Interaction name provided but no taxa loaded")
                return False

            # get the taxon names and descriptions from interaction name providers
            iact_nm_lab, iact_nm_desc = self._parse_levels(str(iact_nm))

            # add names to used taxa
            self.taxa.taxon_names_used.update(iact_nm_lab)

            # check they are found
            iact_nm_lab_in_set = IsInSet(iact_nm_lab, tuple(self.taxa.taxon_names))

            if not iact_nm_lab_in_set:
                self._log(
                    "Unknown taxa in interaction_name descriptor",
                    extra={"join": iact_nm_lab_in_set.failed},
                )
                nm_check = False
            else:
                nm_check = True

            iact_nm_lab = iact_nm_lab_in_set.values
        else:
            iact_nm_lab = []
            iact_nm_desc = ()
            nm_check = True

        if iact_fd_prov:
            # Check that self.dwsh has been populated properly
            if self.dwsh is None:
                self._log(
                    f"Interaction field provided but no dataworksheet provided for this"
                    f" field: {iact_fd}",
                    CRITICAL,
                )
                return False
            # check any field labels match to known taxon fields
            iact_fd_lab, iact_fd_desc = self._parse_levels(str(iact_fd))

            iact_fd_lab_in_set = IsInSet(iact_fd_lab, tuple(self.dwsh.taxa_fields))
            if not iact_fd_lab_in_set:
                self._log(
                    "Unknown taxon fields in interaction_field descriptor",
                    extra={"join": iact_fd_lab_in_set.failed},
                )
                fd_check = False
            else:
                fd_check = True

            iact_fd_lab = iact_fd_lab_in_set.values
        else:
            iact_fd_lab = []
            iact_fd_desc = ()
            fd_check = True

        if len(iact_nm_lab + iact_fd_lab) < 2:
            self._log(
                "At least two interacting taxon labels or fields must be identified"
            )
            num_check = False
        else:
            num_check = True

        all_desc = iact_nm_desc + iact_fd_desc
        if None in all_desc:
            self._log(
                "Label descriptions for interacting taxa incomplete or missing", WARNING
            )

        if nm_check and fd_check and num_check:
            return True
        else:
            return False

    def _parse_levels(self, txt: str) -> tuple[list[str], tuple[str | None, ...]]:
        """Parse categorical variable level descriptions.

        Splits up category information formatted as label:desc;label:desc, which is used
        in both levels for categorical data and interaction descriptors.

        Args:
            txt: The text string to parse

        Returns:
            A list of tuples containing pair of level labels and descriptions.
        """
        # remove terminal semi-colon, if used.
        if txt.endswith(";"):
            txt = txt[:-1]

        # - split the text up by semi-colon
        levels = txt.split(";")

        # - split descriptions and standardize
        parsed_n_parts: set[int] = set()
        levels_parsed: list[list] = []

        for each_level in levels:
            # Split off any description
            level_parts = each_level.split(":")

            # Check for descriptions and build levels_parsed
            level_n_parts = len(level_parts)
            if level_n_parts == 1:
                levels_parsed.append([level_parts[0], None])
            else:
                levels_parsed.append(level_parts[0:2])

            parsed_n_parts.add(level_n_parts)

        # simple formatting checks
        if any([np > 2 for np in parsed_n_parts]):
            self._log("Extra colons in level description.")

        if len(parsed_n_parts) > 1:
            self._log("Provide descriptions for either all or none of the categories")

        level_labels, level_desc = zip(*levels_parsed)

        # - repeated labels?
        if len(set(level_labels)) < len(level_labels):
            self._log("Repeated level labels")

        # - check for numeric level names: integers would be more common
        # but don't let floats sneak through either!
        level_labels_not_numeric = IsNotNumericString(level_labels)
        if not level_labels_not_numeric:
            self._log("Numeric level names not permitted")

        # Remove white space around the labels: simple spacing in the text
        # makes it easier to read and insisting on no space is unnecessary
        level_labels_clean = [vl.strip() for vl in level_labels_not_numeric]

        return level_labels_clean, level_desc

    @classmethod
    def field_type_map(cls):
        """Returns a map of field types to BaseField subclasses.

        This class method iterates over the existing subclasses of BaseField and returns
        a returns a dictionary mapping the field types handled by each subclass to the
        subclass type. This allows the field types listed in field metadata to be
        quickly matched to the correct BaseField subclass.
        """

        field_type_map = {}

        # Create a stack starting with the calling class
        classes = cls.__subclasses__()

        # Pop subclasses while the stack is not empty
        while classes:
            # Get a field class
            current_field_class = classes.pop(0)

            # Add any subclasses from that class
            classes.extend(current_field_class.__subclasses__())

            # Extend the field map
            for ftype in current_field_class.field_types:
                field_type_map.update({ftype: current_field_class})

        return field_type_map

    def validate_data(self, data: list) -> None:
        """Standard method to validate data.

        Most classes that inherit from BaseField will overwrite this to include more
        specific validation rules.
        """

        data = self.run_common_validation(data)

    def run_common_validation(self, data: list) -> list:
        """Validates a list of data provided for a field.

        This base class method runs the common shared validation steps for input data
        for a given field. It checks for:

        * Explicit missing data, using the NA value.
        * Empty cells, which is an error.
        * Excel cell error codes (such as `#VALUE!`)

        Subclasses should also carry out field specific testing. To ensure that only the
        data that passes the common checks is subjected to extra testing, within their
        data validation methods subclasses should use:

            data = self.run_common_validation(data)

        Args:
            data: a set of values from a data table for the field.

        Returns:
            A list of the cleaned input values, excluding blanks, NA and any Excel
            errors.
        """

        # Unless any input is ok (comments fields, basically) filter out
        # missing and blank data. Also check for Excel formula errors.
        self.n_rows += len(data)

        # Look for NAs
        data_no_na = IsNotNA(data, keep_failed=False)
        if not data_no_na:
            self.n_na += len(data_no_na.failed)

        # Look for blank data
        data_no_blanks = IsNotBlank(data_no_na, keep_failed=False)
        if not data_no_blanks:
            self.n_blank += len(data_no_blanks.failed)

        # Look for formula errors:

        # With openpyxl, using data_only=False loads function cells with
        # data_type = 'f', but using data_only=True returns cells with a
        # data_type for the resulting value _or_ an 'e' code for a function
        # error. Using value_only=True in iter_row for speed loses that cue,
        # so the values simply come back as a strings that have to be matched
        # to the of Excel error codes. See also:
        #
        # https://groups.google.com/g/openpyxl-users/c/iNi1MKSP-Bc/m/2q_7cHavBQAJ

        data_no_excel_errors = IsNotExcelError(data_no_blanks, keep_failed=False)
        if not data_no_excel_errors:
            self.n_excel_error += len(data_no_excel_errors.failed)

        # Return cleaned data for use in further checking.
        return data_no_excel_errors.values

    def report(self) -> None:
        """Report on field creation and data validation.

        This method emits any messages queued in the log stack during the creation
        of a field instance and data validation. The method also contains any final
        checking of data values across all validated data.
        """

        LOGGER.info(f"Checking field {self.field_name}")

        FORMATTER.push()

        # Emit the accumulated messages
        for log_level, msg, extra in self.log_stack:
            LOGGER.log(log_level, msg, extra=extra)

        # Run any final logging
        # - warn about NAs
        if self.n_na:
            LOGGER.warning(f"{self.n_na} / {self.n_rows} values missing")

        # We don't tolerate blank cells
        if self.n_blank:
            LOGGER.error(
                f"{self.n_blank} cells are blank or contain only whitespace text"
            )

        # Report on excel error codes
        if self.n_excel_error:
            LOGGER.error(f"{self.n_excel_error} cells contain Excel formula errors")

        FORMATTER.pop()


def field_to_dict(fld: EmptyField | BaseField, col_idx: int) -> dict:
    """Convert a object inheriting from BaseField into a dictionary.

    A function to return a dictionary representation of a field object. This would more
    naturally be a method of BaseField, but needs to include the extended attributes of
    subclasses of BaseField.

    TODO - the BaseField should define all these attributes and then this becomes
    a method

    Args:
        fld: An instance inheriting from BaseField
        col_idx: The column index of the field

    Returns:
        A dictionary representation of the field attributes.
    """

    output = dict(
        field_name=fld.field_name,
        description=fld.meta["description"],
        field_type=fld.meta["field_type"],
        units=None,
        method=None,
        levels=None,
        taxon_field=None,
        taxon_name=None,
        interaction_field=None,
        interaction_name=None,
        range=None,
        col_idx=col_idx,
    )

    if hasattr(fld, "units"):
        output["units"] = getattr(fld, "units")
        output["method"] = getattr(fld, "method")

    if hasattr(fld, "levels"):
        output["levels"] = getattr(fld, "levels")

    if hasattr(fld, "taxon_field"):
        output["levels"] = getattr(fld, "taxon_field")

    if hasattr(fld, "interaction_name"):
        output["interaction_name"] = getattr(fld, "interaction_name")

    if hasattr(fld, "interaction_field"):
        output["interaction_field"] = getattr(fld, "interaction_field")

    if hasattr(fld, "range"):
        output["range"] = getattr(fld, "range")

    return output


class CommentField(BaseField):
    """A BaseField subclass for comments fields.

    This subclass overrides the base data validation to allow any content in comments
    fields. See [safedata_validator.field.BaseField][] for details.
    """

    field_types = ("comments",)

    def validate_data(self, data: list) -> None:
        """Validate data in comment fields.

        This overrides the BaseField
        [validate_data][safedata_validator.field.BaseField.validate_data] method to
        remove any checking on comments fields.
        """

        return


class ReplicateField(BaseField):
    """A BaseField subclass for Replicate and ID fields.

    Applies the base data checking needed on Replicate and ID fields. See
    [safedata_validator.field.BaseField][] for details.
    """

    field_types = ("replicate", "id")


class NumericField(BaseField):
    """A BaseField subclass for numeric fields.

    Extends [BaseField][safedata_validator.field.BaseField] to validate numeric data
    fields.
    """

    field_types: tuple[str, ...] = ("numeric",)
    required_descriptors = (*MANDATORY_DESCRIPTORS, "method", "units")

    def validate_data(self, data: list) -> None:
        """Validate numeric field data.

        Runs the BaseField
        [run_common_validation][safedata_validator.field.BaseField.run_common_validation]
        method and also ensures that data values are numeric.
        """
        data = self.run_common_validation(data)

        numeric = IsNumber(data)

        if not numeric:
            self._log("Cells contain non-numeric values")


class CategoricalField(BaseField):
    """A BaseField subclass for categorical and ordered categorical fields.

    Extends the [BaseField `__init__` method][safedata_validator.field.BaseField]
    to store the level labels and track the reported levels during data
    validation.
    """

    field_types: tuple[str, ...] = ("categorical", "ordered categorical")
    required_descriptors = (*MANDATORY_DESCRIPTORS, "levels")

    def __init__(
        self,
        meta: dict,
        dwsh: DataWorksheet | None = None,
        dataset: Dataset | None = None,
    ) -> None:
        super().__init__(meta, dwsh=dwsh, dataset=dataset)

        # Additional code to validate the levels metadata and store a set of
        # values
        self.level_labels = set()
        self.reported_levels: set[str] = set()
        levels = meta.get("levels")

        if levels is not None and isinstance(levels, str) and not levels.isspace():
            level_labels, level_desc = self._parse_levels(levels)
            self.level_labels = set(level_labels)

    def validate_data(self, data: list) -> None:
        """Validate categorical field data.

        Runs the BaseField
        [run_common_validation][safedata_validator.field.BaseField.run_common_validation]
        method and also checks that string values are provided and populates the set of
        reported levels.
        """
        data = self.run_common_validation(data)

        # Now look for consistency: get the unique values reported in the
        # data, convert to unicode to handle checking of numeric labels and
        # then check the reported levels are a subset of the descriptors.
        # XLRD reads all numbers as floats, so coerce floats back to int
        data_as_string = IsString(data, keep_failed=False)
        if not data_as_string:
            self._log("Cells contain non-text values")

        self.reported_levels.update(data_as_string)

    def report(self) -> None:
        """Report on field creation and data validation for categorical fields.

        Extends the [BaseField report method][safedata_validator.field.BaseField.report]
        to handle undeclared or unused level labels.
        """
        super().report()
        FORMATTER.push()

        extra = self.reported_levels.difference(self.level_labels)
        unused = self.level_labels.difference(self.reported_levels)

        if extra:
            LOGGER.error(
                "Categories found in data missing from levels descriptor: ",
                extra={"join": extra},
            )
        if unused:
            LOGGER.error(
                "Categories found in levels descriptor not used in data: ",
                extra={"join": unused},
            )

        FORMATTER.pop()


class TaxaField(BaseField):
    """A BaseField subclass for taxa fields.

    Extends [BaseField][safedata_validator.field.BaseField] to to track the reported
    taxa in the field during data validation.
    """

    field_types = ("taxa",)

    def __init__(
        self,
        meta: dict,
        dwsh: DataWorksheet | None = None,
        dataset: Dataset | None = None,
    ) -> None:
        super().__init__(meta, dwsh=dwsh, dataset=dataset)

        if self.taxa is None:
            self._log("No taxon details provided for dataset")

        self.taxa_found: set[str] = set()

    def validate_data(self, data: list) -> None:
        """Validate taxa field data.

        Runs the BaseField
        [run_common_validation][safedata_validator.field.BaseField.run_common_validation]
        method and also looks for non-string values and tracks unused or unknown taxon
        names.
        """
        data = self.run_common_validation(data)

        data_as_string = IsString(data, keep_failed=False)

        if not data_as_string:
            self._log("Cells contain non-string values")

        if self.taxa is not None:
            self.taxa_found.update(data_as_string)
            self.taxa.taxon_names_used.update(data_as_string)

    def report(self) -> None:
        """Report on field creation and data validation for taxa fields.

        Extends the [BaseField report method][safedata_validator.field.BaseField.report]
        to emit unused or unknown taxon names
        """
        super().report()
        FORMATTER.push()

        # TODO - not sure about this - no other fields test for emptiness?
        if self.taxa_found == set():
            LOGGER.error("No taxa loaded")
            return

        if self.taxa is not None:
            extra_taxa = self.taxa_found.difference(self.taxa.taxon_names)

            if extra_taxa:
                LOGGER.error("Includes unreported taxa: ", extra={"join": extra_taxa})

            # add the found taxa to the list of taxa used
            self.taxa.taxon_names_used.update(self.taxa_found)

        FORMATTER.pop()


class LocationsField(BaseField):
    """A BaseField subclass for location fields.

    Extends the [BaseField `__init__` method][safedata_validator.field.BaseField]
    to check that all the values provided in a locations field are found in the
    locations data for the dataset.
    """

    field_types = ("locations", "location")

    def __init__(
        self,
        meta: dict,
        dwsh: DataWorksheet | None = None,
        dataset: Dataset | None = None,
    ) -> None:
        super().__init__(meta, dwsh=dwsh, dataset=dataset)

        if self.locations is None:
            self._log("No location details provided for dataset")

        self.locations_found: set[str] = set()

    def validate_data(self, data: list) -> None:
        """Validate location field data.

        Runs the BaseField
        [run_common_validation][safedata_validator.field.BaseField.run_common_validation]
        method and also checks that location names are all known.
        """
        data = self.run_common_validation(data)

        data_locs = IsLocName(data, keep_failed=False)

        if not data_locs:
            self._log("Cells contain invalid location values")

        # Now should be strings and integer codes - convert to string
        # representations as used in the locations
        data = [str(v) for v in data_locs]

        if self.locations is not None:
            self.locations_found.update(data)
            self.locations.locations_used.update(data)

    def report(self) -> None:
        """Report on field creation and data validation for categorical fields.

        Extends the [BaseField report method][safedata_validator.field.BaseField.report]
        to emit undeclared location names.
        """
        super().report()
        FORMATTER.push()

        # TODO - not sure about this - no other fields test for emptiness?
        if self.locations_found == set():
            LOGGER.error("No locations loaded")
            return

        if self.locations is not None:
            extra_locs = self.locations_found.difference(self.locations.locations)

            if extra_locs:
                LOGGER.error(
                    "Includes unreported locations: ", extra={"join": extra_locs}
                )

            # add the found taxa to the list of taxa used
            self.locations.locations_used.update(self.locations_found)

        FORMATTER.pop()


class GeoField(BaseField):
    """A BaseField subclass for latitude and longitude fields.

    Extends the [BaseField `__init__` method][safedata_validator.field.BaseField]
    to check latitude and longitude data.
    """

    field_types = ("latitude", "longitude")

    def __init__(
        self,
        meta: dict,
        dwsh: DataWorksheet | None = None,
        dataset: Dataset | None = None,
    ) -> None:
        super().__init__(meta, dwsh=dwsh, dataset=dataset)

        if self.dataset is None:
            self._log("No dataset object provided - cannot update extents")

        self.min = None
        self.max = None

    def validate_data(self, data: list) -> None:
        """Validate latitude and longitude data.

        Runs the BaseField
        [run_common_validation][safedata_validator.field.BaseField.run_common_validation]
        method, and also checks for non-decimal formatting (e.g. 12°24'32"W) and
        collates the data range.
        """
        data = self.run_common_validation(data)

        data_as_number = IsNumber(data, keep_failed=False)

        if not data_as_number:
            self._log("Field contains non-numeric data")

            if any([RE_DMS.search(str(dt)) for dt in data_as_number.failed]):
                self._log(
                    "Possible degrees minutes and seconds formatting? Use decimal "
                    "degrees",
                    WARNING,
                )

        if data_as_number.values:
            self.min = (
                min([*data_as_number.values, self.min])
                if self.min
                else min(data_as_number.values)
            )
            self.max = (
                max([*data_as_number.values, self.max])
                if self.max
                else max(data_as_number.values)
            )

    def report(self) -> None:
        """Report on field creation and data validation for latitude and longitude data.

        Extends the [BaseField report method][safedata_validator.field.BaseField.report]
        to check the coordinate range against the dataset geographic extents.
        """
        super().report()
        FORMATTER.push()

        if self.min is None or self.max is None:
            return

        # Note that self.min and self.max can be None here if all data is invalid, but
        # the update process handles None.
        if self.dataset is not None:
            if self.meta["field_type"] == "latitude":
                self.dataset.latitudinal_extent.update([self.min, self.max])
            elif self.meta["field_type"] == "longitude":
                self.dataset.longitudinal_extent.update([self.min, self.max])

        FORMATTER.pop()


class NumericTaxonField(NumericField):
    """A NumericField subclass for numeric trait and abundance fields.

    Further extends the [NumericField][safedata_validator.field.NumericField] subclass
    to also apply checking of the taxonomic metadata for numeric traits and abundance
    data.
    """

    # BREAKING CHANGE - abundance did not previously require 'units' as metadata,
    # which is stupid in retrospect. Individuals? Individuals per m2? Individuals
    # per hour? I do not know how many existing datasets will fall foul of this.

    field_types = ("abundance", "numeric trait")
    check_taxon_meta = True


class CategoricalTaxonField(CategoricalField):
    """A CategoricalField subclass for categorical trait fields.

    Further extends the [NumericField][safedata_validator.field.NumericField] subclass
    to also apply checking of the taxonomic metadata for numeric traits and abundance
    data.
    """

    field_types = ("categorical trait", "ordered categorical trait")
    check_taxon_meta = True


class NumericInteractionField(NumericField):
    """A NumericField subclass for numeric interaction fields.

    Further extends the [NumericField][safedata_validator.field.NumericField] subclass
    to also apply checking of the interaction metadata for numeric interaction data.
    """

    field_types = ("numeric interaction",)
    check_interaction_meta = True


class CategoricalInteractionField(CategoricalField):
    """A CategoricalField subclass for categorical trait fields.

    Further extends the [NumericField][safedata_validator.field.NumericField] subclass
    to also apply checking of the interaction metadata for numeric traits and abundance
    data.
    """

    field_types = ("categorical interaction",)
    check_interaction_meta = True


class TimeField(BaseField):
    """A BaseField subclass for time fields.

    Extends the [BaseField][safedata_validator.field.BaseField] subclass to validate
    data in time fields.  Time field validation is primarily about checking the
    consistency of the provided data. The function accepts either as ISO strings in a
    consistent time format or datetime.time objects. The parsing of timestrings follows
    ISO8601, which includes quite a range of valid inputs: '11:12:13' and '111213' are
    both valid.

    One problem here is that - when stored as Excel datetime cells - dates and times in
    Excel are poorly typed: essentially as a number with a meaning that depends in part
    on the cell format. The openpyxl package returns datetime.time objects for cells
    with time formats.
    """

    field_types = ("time",)

    def __init__(
        self,
        meta: dict,
        dwsh: DataWorksheet | None = None,
        dataset: Dataset | None = None,
    ) -> None:
        super().__init__(meta, dwsh=dwsh, dataset=dataset)

        # Defaults
        self.first_data_class_set: set | None = None
        self.consistent_class = True
        self.expected_class = True

        self.bad_strings: list[str] = []
        self.min = None
        self.max = None

    def validate_data(self, data: list) -> None:
        """Validate time field data.

        Runs the BaseField
        [run_common_validation][safedata_validator.field.BaseField.run_common_validation]
        methoda nd also checks that time data has consistent formatting and are valid
        time values.
        """
        data = self.run_common_validation(data)

        # TODO: report row number of issues - problem of mismatch between
        # self.n_rows (_all_ rows) and index in data (non-NA rows)

        # Check for consistent class formatting, using first row. Do not try and
        # validate further when data is not consistently formatted.
        if self.consistent_class and self.expected_class:
            cell_types = [type(dt) for dt in data]
            cell_type_set = set(cell_types)

            # Are all the values of expected types?
            if not {datetime.time, str}.issuperset(cell_type_set):
                self.expected_class = False
                return

            # Are the values internally consistent and consistent with
            # previously loaded data
            if len(cell_type_set) == 0:
                # If data is empty skip consistency checking
                return
            elif len(cell_type_set) > 1:
                self.consistent_class = False
                return
            elif self.first_data_class_set is None:
                self.first_data_class_set = cell_type_set
            elif self.first_data_class_set != cell_type_set:
                self.consistent_class = False
                return
        else:
            return

        # Value checking only happens while data are consistently formatted
        # There is no need to check time objects passed in, just time formatted
        # strings
        if self.first_data_class_set == {str}:
            for val in data:
                try:
                    _ = parser.isoparser().parse_isotime(val)
                except ValueError:
                    self.bad_strings.append(val)

    def report(self):
        """Report on field creation and data validation for time fields.

        Extends the [BaseField report method][safedata_validator.field.BaseField.report]
        to flag inconsistent time formatting and invalid data.
        """
        super().report()
        FORMATTER.push()

        if not self.expected_class:
            LOGGER.error(
                "Time data include values neither ISO string nor time formatted"
            )
            return

        if not self.consistent_class:
            LOGGER.error("Time data mixes ISO string and time formatted rows")
            return

        if self.bad_strings:
            LOGGER.error(
                "ISO time strings contain badly formatted values: e.g.",
                extra={"join": self.bad_strings[:5]},
            )

        FORMATTER.pop()


class DatetimeField(BaseField):
    """A BaseField subclass for date and datetime fields.

    Extends the [BaseField `__init__` method][safedata_validator.field.BaseField]
    to validate date and datetime. This is primarily about checking the
    consistency of the provided data. The function accepts either as ISO strings
    in a consistent datetime or date format or datetime.datetime objects. The
    parsing of timestrings follows ISO8601, which includes quite a range of
    valid inputs.

    One problem here is that - when stored as Excel datetime cells - dates and
    times in Excel are poorly typed: essentially as a number with a meaning that
    depends in part on the cell format. The openpyxl package returns
    datetime.datetime objects for cells with datetime or date formats.
    """

    field_types = ("datetime", "date")

    def __init__(
        self,
        meta: dict,
        dwsh: DataWorksheet | None = None,
        dataset: Dataset | None = None,
    ) -> None:
        super().__init__(meta, dwsh=dwsh, dataset=dataset)

        if self.dataset is None:
            self._log("No dataset object provided - cannot update extents")

        # Defaults
        self.first_data_class_set: set | None = None
        self.consistent_class = True
        self.expected_class = True
        self.all_midnight = True

        self.bad_strings: list[str] = []
        self.min = None
        self.max = None

    def validate_data(self, data: list) -> None:
        """Validate date and datetime field data.

        Runs the BaseField
        [run_common_validation][safedata_validator.field.BaseField.run_common_validation]
        method and also checks that time data has consistent formatting and are valid
        date or datetime values.
        """

        data = self.run_common_validation(data)

        # TODO: report row number of issues - problem of mismatch between
        # self.n_rows (_all_ rows) and index in data (non-NA rows)

        # Check for consistent class formatting, using first row. Do not try and
        # validate further when data is not consistently formatted.
        if self.consistent_class and self.expected_class:
            cell_types = [type(dt) for dt in data]
            cell_type_set = set(cell_types)

            # Are all the values of expected types?
            if not {datetime.datetime, str}.issuperset(cell_type_set):
                self.expected_class = False
                return

            # Are the values internally consistent and consistent with
            # previously loaded data
            if len(cell_type_set) == 0:
                # If data is empty skip consistency checking
                return
            elif len(cell_type_set) > 1:
                self.consistent_class = False
                return
            elif self.first_data_class_set is None:
                self.first_data_class_set = cell_type_set
            elif self.first_data_class_set != cell_type_set:
                self.consistent_class = False
                return
        else:
            return

        # Value checking only happens while data are consistently formatted
        # Look for string formatting, but also presence/absence of non-zero
        # time components.
        midnight = datetime.time(0, 0)

        if self.first_data_class_set == {str}:
            parsed_strings = []
            for val in data:
                try:
                    val_parse = parser.isoparse(val)
                    if val_parse.time() != midnight:
                        self.all_midnight = False
                    parsed_strings.append(val_parse)
                except ValueError:
                    self.bad_strings.append(val)

            # standardised to datetime objects
            data = parsed_strings

        elif self.first_data_class_set == {datetime.datetime}:
            for val in data:
                if val.time() != midnight:
                    self.all_midnight = False

        # Check that data isn't empty
        if data:
            # Update range for extents
            if self.min is None:
                self.min = min(data)
            else:
                self.min = min([self.min, *data])

            if self.max is None:
                self.max = min(data)
            else:
                self.max = max([self.max, *data])

    def report(self):
        """Report on field creation and data validation for date and datetime fields.

        Extends the [BaseField report method][safedata_validator.field.BaseField.report]
        to flag inconsistent  date and datetime formatting and invalid data.
        """
        super().report()
        FORMATTER.push()

        # INconsistent and bad data classes
        if not self.expected_class:
            LOGGER.error(
                "Date or datetime data include values neither ISO string nor time "
                "formatted"
            )
            return

        if not self.consistent_class:
            LOGGER.error(
                "Date or datetime data mixes ISO string and time formatted rows"
            )
            return

        # Bad string formats
        if self.bad_strings:
            LOGGER.error(
                "ISO datetime strings contain badly formatted values: e.g.",
                extra={"join": self.bad_strings[:5]},
            )

        # Datetime vs Date checking - Datetimes _could_ all be at midnight, so not an
        # error but does look odd.
        if self.meta["field_type"] == "datetime" and self.all_midnight:
            LOGGER.warning("Field is of type datetime, but only reports dates")
        if self.meta["field_type"] == "date" and not self.all_midnight:
            LOGGER.error("Field is of type date, but includes time data")

        # Update extent if possible - note that inheritance means that isinstance
        # in extent.Extent is not successfully testing for datetime.datetime rather
        # than set datatype of datetime.date
        if not (self.dataset is None or self.min is None or self.max is None):
            self.dataset.temporal_extent.update([self.min.date(), self.max.date()])

        FORMATTER.pop()


class FileField(BaseField):
    """A BaseField subclass for time fields.

    Extends the [BaseField][safedata_validator.field.BaseField] subclass to check file
    fields. The data values need to match to an external file or can be contained within
    an archive file provided in the 'file_container' descriptor.
    """

    field_types = ("file",)

    def __init__(
        self,
        meta: dict,
        dwsh: DataWorksheet | None = None,
        dataset: Dataset | None = None,
    ) -> None:
        super().__init__(meta, dwsh=dwsh, dataset=dataset)

        self.unknown_file_names: set[str] = set()

        # Check whether filename testing is possible
        if self.summary is None:
            self._log(
                "No Summary instance provided - cannot check file fields",
                level=CRITICAL,
            )
            return
        if self.summary.external_files is None:
            self._log("No external files listed in Summary")
            return

        self.external_names = {ex["file"] for ex in self.summary.external_files}

        # Look for a file container metadata value, pointing to a single file
        # containing all of the listed values (zip etc)
        self.file_container = self.meta.get("file_container")

        if self.file_container is not None:
            if self.file_container not in self.external_names:
                self._log(
                    f"Field file_container value not found in external files:"
                    f" {self.file_container}"
                )

    def validate_data(self, data: list):
        """Validate file field data.

        Runs the BaseField
        [run_common_validation][safedata_validator.field.BaseField.run_common_validation]
        method, and also checks for unknown files given in a file field.
        """

        data = self.run_common_validation(data)

        # If the files are listed in external and not provided in a file
        # container then check they are all present
        if not self.file_container:
            unknown_files = set(data) - self.external_names
            if unknown_files:
                self.unknown_file_names.update(unknown_files)

    def report(self):
        """Report on field creation and data validation for file fields.

        Extends the [BaseField report method][safedata_validator.field.BaseField.report]
        to flag unknown files.
        """
        super().report()
        FORMATTER.push()

        if self.unknown_file_names:
            LOGGER.error(
                "Field contains external files not provided in Summary: ",
                extra={"join": self.unknown_file_names},
            )

        FORMATTER.pop()


class EmptyField:
    """A BaseField-like interface for empty fields.

    This class mocks the interface of BaseField and is only used to check trailing empty
    fields in loaded field data. These are columns which have no field metadata, but
    which have are included in the data table. The data rows should be empty.

    This class does not inherit from BaseField, so is never included in the
    BaseField.field_type_map.

    Args:
        meta: A dictionary of field metadata. All values should be None with the
            exception of the column index number (`col_idx`).
    """

    def __init__(self, meta: dict) -> None:
        self.meta = meta
        self.empty = True
        # Get a field name - either a column letter from col_idx if set or 'Unknown'
        idx = self.meta.get("col_idx")
        if idx is None:
            self.field_name = "Unknown"
        else:
            self.field_name = f"Column_{get_column_letter(idx)}"

    def validate_data(self, data) -> None:
        """Validates that empty fields contain no data."""
        blank = [blank_value(val) for val in data]

        if not all(blank):
            self.empty = False

    def report(self):
        """Report on an empty field."""

        if not self.empty:
            LOGGER.info(f"Checking field {self.field_name}")
            FORMATTER.push()
            LOGGER.error("Trailing field with no descriptors contains data.")
            FORMATTER.pop()
