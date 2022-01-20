from openpyxl import load_workbook
import datetime
from safedata_validator.resources import Resources
from safedata_validator.locations import Locations
from safedata_validator.taxa import Taxa
from safedata_validator.summary import Summary
from safedata_validator.logger import LOGGER, FORMATTER, CH, log_and_raise
from safedata_validator.extent import Extent

class Dataset:

    def __init__(self, filename, resources=Resources()):
        """
        The Dataset class links an input file with a particular configuration
        of the safedata_validator validation sources, and then loads the
        components of the dataset.

        Args:
            filename: Path to an .xlsx file containing safedata formatted data.
            resources: An instance of class Resources providing a set of
                validation resources for taxa and locations. The default
                attempts to locate these resources from a user or site config
                file.
        """

        self.resources = resources
        self.filename = filename

        self.summary = Summary()
        self.locations = Locations()
        self.taxa = Taxa()

        # Extents - these can be loaded from the Summary or compiled from the
        # data, so the Summary and dataset extents are held separately so that
        # they can be validated against one another once all data is checked.
        self.temporal_extent = Extent('temporal extent', datetime.date)
        self.latitudinal_extent = Extent('latitudinal extent', float)
        self.longitudinal_extent = Extent('longitudinal extent', float)

        # Open the workbook with:
        #  - read_only to use the memory optimised read_only implementation.
        #    This is a bit restricted as it only exposes row by row iteration
        #    to avoid expensive XML traversal but has a low memory footprint.
        #  - data_only to load values not formulae for equations.

        self.wb = load_workbook(filename, read_only=True, data_only=True)

        # Populate summary
        if 'Summary' in self.wb.sheetnames:
            LOGGER.info("Checking Summary worksheet")
            FORMATTER.push()
            self.summary.load(self.wb['Summary'], self.wb.sheetnames, validate_doi=XXX, valid_pid=XXX, )
            FORMATTER.pop()
        else:
            # No summary is impossible - so report error TODO - not counted!
            LOGGER.error("No locations worksheet found - moving on")

        # Populate locations
        if 'Locations' in self.wb.sheetnames:
            LOGGER.info("Checking Locations worksheet")
            FORMATTER.push()
            self.locations.load(self.wb['Locations'], self.resources)
            FORMATTER.pop()
        else:
            # No locations is pretty implausible - lab experiments?
            LOGGER.warn("No locations worksheet found - moving on")

        # Populate taxa
        if 'Taxa' in self.wb.sheetnames:
            LOGGER.info("Checking Taxa worksheet")
            FORMATTER.push()
            self.taxa.load(self.wb['Taxa'], self.resources)
            FORMATTER.pop()
        else:
            # Leave the default empty Taxa object
            LOGGER.warn("No Taxa worksheet found - moving on")

