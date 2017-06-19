#!/usr/bin/python

from __future__ import print_function
import openpyxl
import sys
import datetime
import re
from collections import Counter

"""
Functions to check the contents of a SAFE project formatted Excel dataset.
Where possible, the functions will try to keep on running to give a list

"""

# Global definition of set of acceptable NA values: blank cells or 'NA'
NA = {None, u'NA'}
# define some regular expressions used to check validity
re_orcid = re.compile('[0-9]{4}-[0-9]{4}-[0-9]{4}-[0-9]{4}')
re_email = re.compile('\S+@\S+\.\S+')
re_name = re.compile('[^,]+,[ ]?[^,]+')
re_whitespace_only = re.compile('^\s+$')

class Messages(object):

    """
    This class keeps a record of messages to be conveyed to the
    user and keeps a tally of warnings.
    """

    def __init__(self, header):
        self.header = header
        self.msg = []
        self.type = []

    def info(self, message):
        self.msg.append(message)
        self.type.append('info')

    def warn(self, message):
        self.msg.append(message)
        self.type.append('warn')

    def subhead(self, message):
        self.msg.append(message)
        self.type.append('subhead')

    def abort(self, message):
        self.msg.append(message)
        self.type.append('abort')

    def count_warnings(self):

        return sum([msg == 'warn' for msg in self.type])

    def no_warnings(self):

        if self.count_warnings() > 0:
            return False
        else:
            return True

    def display(self):

        print(self.header)
        for m, t in zip(self.msg, self.type):
            if t == 'info':
                print(' - ' + m)
            elif t == 'warn':
                print(' ! ' + m)
            elif t == 'subhead':
                print(' * ' + m)
            elif t == 'abort':
                print(' !!! Aborting - ' + m)


def open_file(fname):

    try:
        wb = openpyxl.load_workbook(filename=fname, data_only=True, read_only=True)
    except:
        raise IOError('Could not open file')

    return wb


def get_summary(wb):

    """
    Checks the information in the summary worksheet and looks for the metadata and
    dataset worksheets.

    Parameters:
        wb: An openpyxl Workbook instance
    Returns:
        A dictionary of the summary metadata
    """

    # try and get the summary worksheet
    m = Messages("Checking Summary worksheet")
    try:
        ws = wb['Summary']
    except KeyError:
        m.abort("Summary worksheet not found")
        m.display()
        return None

    # load dictionary of summary information block, allowing for multiple
    # columns for fields (data compilers, dataset sheets).
    summary_dict = {}
    col_range = range(2, ws.max_column + 1)
    max_row = ws.max_row

    for rw in range(1, max_row + 1):
        vals = [ws.cell(row=rw, column=cl).value for cl in col_range]
        summary_dict[ws.cell(row=rw, column=1).value] = vals

    # build return dictionary
    ret_dict = {}

    # Check the minimal keys are expected
    required = {"SAFE Project ID", "Access status", "Title", "Description",
                "Author name", "Author email", "Author affiliation", "Author ORCID",
                "Worksheet name", "Worksheet title", "Worksheet description", "Worksheet type", }

    found = set(summary_dict.keys())

    # don't bail here - try and get as far as possible
    if not found.issuperset(required):
        m.warn('Missing metadata fields: {}'.format(','.join(required - found)))

    # check for contents
    if any([set(x) == {None} for x in summary_dict.values()]):
        m.warn('Metadata fields with no information.')

    #
    if 'SAFE Project ID' not in summary_dict:
        m.warn('SAFE Project ID missing')
    elif type(summary_dict['SAFE Project ID'][0]) != long:
        m.warn('SAFE Project ID is not an integer.')
    else:
        ret_dict['project_id'] = summary_dict['SAFE Project ID'][0]

    # CHECK ACCESS STATUS AND EMBARGO DETAILS
    if 'Access status' in summary_dict:
        access_status = summary_dict['Access status'][0]
        ret_dict['access_status'] = summary_dict['Access status'][0]
        ret_dict['embargo_date'] = None
        if access_status == 'Embargo':
            if 'Embargo date' not in summary_dict:
                m.warn('Dataset embargoed but no date provided.')
            embargo_date = summary_dict['Embargo date'][0]
            now = datetime.datetime.now()
            if type(embargo_date) != datetime.datetime:
                m.warn('Embargo date not formatted as date.')
            elif embargo_date < now:
                m.warn('Embargo date is in the past.')
            elif embargo_date > now + datetime.timedelta(days=2*365):
                m.warn('Embargo date more than two years in the future.')
            else:
                ret_dict['embargo_date'] = embargo_date
        elif access_status == 'Open':
            pass
        else:
            m.warn('Access status must be Open or Embargo not {}'.format(access_status))

    # CHECK AUTHORS
    author_keys  = ['Author name', 'Author affiliation', 'Author email', 'Author ORCID']
    if set(summary_dict.keys()).issuperset(author_keys):
        authors = zip(*[summary_dict[x] for x in author_keys])

        # remove any completely blank entries
        authors = [x for x in authors if x != tuple([None] * 4)]

        # convert to dict in Zenodo style and check completeness and validity
        for ind, au in enumerate(authors):
            au = {k: v for k, v in zip(['name', 'affiliation', 'email', 'orcid'], au)}
            # look for missing values
            for k, v in au.iteritems():
                if v is None:
                    m.warn('Author {} missing'.format(k))
            # check validity
            if not re_name.match(au['name']):
                m.warn('Author name not formated as last_name, first_names: {}'.format(au['name']))
            if not re_orcid.match(str(au['orcid'])) or au['orcid'] is None:
                m.warn('ORCID not properly formatted: {}'.format(au['orcid']))
            if not re_email.match(au['email']):
                m.warn('Email not properly formatted: {}'.format(au['email']))
            authors[ind] = au

        ret_dict['authors'] = authors
    else:
        m.warn('Author metadata block incomplete.')
        ret_dict['authors'] = None

    # CHECK DATASETS
    ds_keys = ['Worksheet name', 'Worksheet title', 'Worksheet description', 'Worksheet type']
    valid_sheets = set(wb.get_sheet_names())

    if set(summary_dict.keys()).issuperset(ds_keys):
        datasets = zip(*[summary_dict[x] for x in ds_keys])

        # remove any completely blank entries
        datasets = [x for x in datasets if x != tuple([None] * 4)]

        # validate
        for ind, dt in enumerate(datasets):
            dt = {k: v for k, v in zip(['worksheet', 'title', 'description', 'type'], dt)}
            # look for missing values
            for k, v in dt.iteritems():
                if v is None:
                    m.warn('Dataset {} missing'.format(k))
            # validate
            if dt['worksheet'] not in valid_sheets:
                m.warn('Dataset worksheet not found: {}.'.format(dt['worksheet']))
            if dt['type'] not in ['Dataframe', 'Matrix']:
                m.warn('Dataset type not valid: {}.'.format(dt['type']))
            datasets[ind] = dt

        ret_dict['datasets'] = datasets
    else:
        m.warn('Dataset metadata block incomplete.')
        ret_dict['datasets'] = None

    # check for extra undocumented spreadsheets
    if 'Worksheet name' in summary_dict:
        expected_sheets = set(summary_dict['Worksheet name'] +
                              ['Summary', 'Taxa', 'Locations'])
        if valid_sheets != expected_sheets:
            m.warn('Extra sheets found in workbook: '
                   '{}'.format(','.join(valid_sheets - expected_sheets)))

    if m.no_warnings():
        m.info('Summary formatted correctly')
    else:
        m.info('Summary contains {} errors'.format(m.count_warnings()))

    m.display()

    return ret_dict


def get_locations(wb):

    # try and get the locations worksheet
    m = Messages("Checking Locations worksheet")
    try:
        locs = wb['Locations']
    except KeyError:
        # No locations is pretty implausible
        m.abort("No lcoations worksheet found")
        m.display()
        return None

    # Check the headers
    fields = []
    max_col = locs.max_column
    max_row = locs.max_row

    for cl in range(1, max_col):
        fields.append(locs.cell(row=1, column=cl).value)

    # check the key fields are there
    if not set(fields).issuperset({'Location name'}):
        m.abort('Location name must be included')
        m.display()
        return None

    # get the location names
    names_col = fields.index('Location name') + 1
    rows = range(2, max_row + 1)
    loc_names = [locs.cell(row=rw, column=names_col).value for rw in rows]

    # check for duplicate names
    if len(set(loc_names)) != len(loc_names):
        m.warn('Duplicated location names in Locations worksheet')

    loc_names = set(loc_names)

    if m.no_warnings():
        m.info('{} locations loaded correctly'.format(len(loc_names)))
    else:
        m.info('Locations contains {} errors'.format(m.count_warnings()))

    m.display()

    return loc_names


def get_taxa(wb):

    """
    Checks and reads the content of the taxa worksheet.

    Parameters:
        wb: An openpyxl Workbook instance.
    Returns:
        A set of the taxon names to be used in the datasets.
    """

    # try and get the taxon worksheet
    m = Messages("Checking Taxa worksheet")
    try:
        taxa = wb['Taxa']
    except KeyError:
        # This might mean that the study doesn't have any taxa, so return an empty
        # set. If the datasets then contain taxonomic names, it'll fail gracefully.
        m.info("No taxa worksheet found - assuming no taxa in data for now!")
        return set()

    # Check the headers
    fields = []
    max_col = taxa.max_column
    max_row = taxa.max_row

    for cl in range(1, max_col):
        fields.append(taxa.cell(row=1, column=cl).value)

    # check the two key fields are there
    if not set(fields).issuperset({'Taxon name', 'Taxon type'}):
        m.abort('Taxon name and Taxon type columns must be included')
        return None

    # get the taxon names and types
    names_col = fields.index('Taxon name') + 1
    types_col = fields.index('Taxon type') + 1
    rows = range(2, max_row + 1)
    taxon_names = [taxa.cell(row=rw, column=names_col).value for rw in rows]
    taxon_types = [taxa.cell(row=rw, column=types_col).value for rw in rows]

    # check for duplicate names
    if len(set(taxon_names)) != len(taxon_names):
        m.warn('Duplicated taxon names in taxa worksheet')

    # Check for some common taxonomic levels. We want as much taxonomic
    # context as possible, but if all taxa are morphospecies / functional
    # groups to family level then genus and species are just empty columns.
    taxon_levels = ['Kingdom', 'Phylum', 'Class', 'Order', 'Family', 'Genus', 'Species']
    if len(set(taxon_types) & set(taxon_levels)) == 0:
        m.warn('Please provide fields giving at least some of the core taxonomic levels')

    # now cross check the given types against the field headers
    types_found = set(taxon_types)
    types_valid = set(fields) - {'Taxon name', 'Taxon type'}
    if not types_found.issubset(types_valid):
        m.warn('Some rows have taxon types that do not match a column name: '
               '{}'.format(','.join(types_found - types_valid)))

    if m.no_warnings():
        m.info('{} taxa loaded correctly'.format(len(taxon_names)))
    else:
        m.info('Taxa contains {} errors'.format(m.count_warnings()))

    m.display()

    return set(taxon_names)


def check_dataframe(wb, ws_name, taxa, locations):

    """
    This function carries out basic checks on a 'dataframe' worksheet -
    that is, tabular data with fields of different types of variable

    Parameters:
        wb: An openpyxl Workbook instance
        ws: The worksheet name
        taxa: A list of valid taxa
        locations: A list of valid locations

    :return:
    """

    m = Messages('Checking dataframe {}'.format(ws_name))
    # get the worksheet and data dimensions
    ws = wb[ws_name]
    max_col = ws.max_column
    max_row = ws.max_row

    # get the metadata field names
    # - first search the first 20 rows for the 'field_name' descriptor which
    #   shows the end of the metadata and the start of the data
    descriptors = [ws.cell(column=1, row=rw).value for rw in range(1, 21)]
    if 'field_name' in descriptors:
        field_name_row = descriptors.index('field_name') + 1
        descriptors = descriptors[:field_name_row]
    else:
        m.abort('Cannot parse data: field_name row not found')
        m.display()
        return None

    # get the metadata for each field
    field_metadata = []
    for cl in range(2, max_col + 1):
        this_field = {k: ws.cell(column=cl, row=i + 1).value for i, k in enumerate(descriptors)}
        field_metadata.append(this_field)

    # check field types - allow for empty field metadata, could be:
    # i) spacer column within fields
    # ii) the user has some comments or noise out to the right (or did,
    #     and Excel just can't let it go).
    # We don't much care about this.
    ft_found = set(fld['field_type'] for fld in field_metadata)
    ft_known = {'Location', 'Date', 'Time', 'Datetime', 'Categorical',
                'Continuous', 'Abundance', 'Comments', 'Taxa', None}

    if not ft_found.issubset(ft_known):
        m.warn('Unknown field types: {}'.format(','.join(ft_found - ft_known)))

    # check extra descriptors
    if 'Categorical' in ft_found and 'categories' not in descriptors:
        m.warn('Categorical data fields found but no category descriptions provided.')

    if 'Continuous' in ft_found and 'units' not in descriptors:
        m.warn('Continuous data fields found but no units provided.')

    # check field names unique (drop None)
    fn_found = Counter([fld['field_name'] for fld in field_metadata
                        if fld['field_name'] is not None])
    duplicated = [k for k, v in fn_found.iteritems() if v > 1]
    if len(duplicated) > 0:
        m.warn('Field names duplicated: {}'.format(', '.join(duplicated)))

    # check mandatory fields

    # check each field
    data_rows = range(field_name_row + 1, max_row + 1)
    for idx, meta in enumerate(field_metadata):

        # skip any field with no metadata
        if set(meta.values()) == {None}:
            break

        # prep the messages instance to pass to functions
        m.subhead('Checking field {}'.format(meta['field_name']))

        # read the values
        data_block = ws.get_squared_range(idx + 2, field_name_row + 1, idx + 2, max_row)
        data = [cl[0].value for cl in data_block]

        # check for missing data
        check_missing_data(meta, data, m)

        # run consistency checks where needed
        if meta['field_type'] == 'Categorical':
            check_field_categorical(meta, data, m)
        elif meta['field_type'] in ['Date', 'Datetime']:
            check_field_datelike(meta, data, m)
        elif meta['field_type'] == 'Time':
            check_field_time(meta, data, m)
        elif meta['field_type'] == 'Taxa':
            check_field_taxa(meta, data, m, taxa)
        else:
            pass

    if m.no_warnings():
        m.info('Dataframe formatted correctly')
    else:
        m.info('Dataframe contains {} errors'.format(m.count_warnings()))

    m.display()

    return None


def check_missing_data(meta, data, m):

    # two acceptable options for missing data: None (blank cell) or u'NA'
    # - what about blank strings?

    na_vals = [vl in NA for vl in data]
    if any(na_vals):
        m.info('{} / {} values missing'.format(sum(na_vals), len(na_vals)))

    # and then there are non-empty cells containing only whitespace strings
    ws_only = [re_whitespace_only.match(vl) for vl in data if type(vl) in [str, unicode]]
    if any(ws_only):
        m.warn('{} cells contain whitespace only text'.format(sum(ws_only)))


"""
Field checker functions below. Must be capable of handling the NA values elegantly
"""


def check_field_datelike(meta, data, m):

    """
    Checks for consistency in date and datetime fields

    :param meta:
    :param data:
    :return:
    """

    # Check type (excluding NA values)
    type_check = set([type(vl) for vl in data if vl not in NA])
    if type_check != {datetime.datetime}:
        if datetime.time in type_check:
            m.warn('Time formatted data found in date field.')
        else:
            m.warn('Non-date formatted data found.')

    # Check no time component in date fields
    if meta['field_type'] == 'Date':
        no_time = [vl.time() == datetime.time(0, 0) for vl in data if vl not in NA]
        if not all(no_time):
            m.warn('Datetimes found in field of type Date.')


def check_field_time(meta, data, m):

    """
    Checks for consistency in time fields

    :param meta:
    :param data:
    :return:
    """

    # Check type (excluding NA values)
    type_check = set([type(vl) for vl in data if vl not in NA])
    if type_check != {datetime.time}:
        m.warn('Non-time formatted data found.')


def check_field_categorical(meta, data, m):

    # this has already been tested but make it robust
    if 'categories' not in meta:
        m.abort('Category levels metadata not provided')
        return m

    # check and parse the category labels
    if meta['categories'] is None:
        m.abort('Category descriptions empty')
        return m
    elif type(meta['categories']) is not unicode:
        m.abort('Category descriptions are not text')
        return m
    else:
        # split the text up by semi-colon
        categories = meta['categories'].split(';')
        # strip off any description after a colon
        category_labels = set([ct.split(':')[0] for ct in categories])

    # get the unique values reported in the data, removing missing values and
    # then convert to unicode in case anyone has used numeric level names.
    reported = set(data) - NA
    reported = {unicode(lv) for lv in reported}

    # check for completeness
    if not reported.issubset(category_labels):
        m.warn('Categories found in data missing from description'
               '{}: '.format(','.join(reported - category_labels)))


def check_field_continuous(meta, data):

    """
    Checks for consistency in time fields

    :param meta:
    :param data:
    :return:
    """

    # Check type (excluding NA values)
    type_check = set([type(vl) for vl in data if vl not in NA])
    if not type_check.issubset({long, float}):
        raise RuntimeError('Non numeric data found in {}.'.format(meta['field_name']))


def check_field_taxa(meta, data, m, taxa):

    # check if taxa are all provided
    found = set(data)
    if not found.issubset(taxa):
        m.warn('Includes taxa missing from Taxa worksheet: {}'.format(', '.join(found - taxa)))


def main():

    fn = sys.argv[1]
    wb = open_file(fname=fn)
    summary = get_summary(wb)
    locations = get_locations(wb)
    taxa = get_taxa(wb)

    if summary is not None:
        for ds in summary['datasets']:
            if ds['type'] == 'Dataframe':
                check_dataframe(wb, ds['worksheet'], taxa, locations)
            elif ds['type'] == 'Matrix':
                # check_matrix(wb, ds['name'], taxa, locations)
                pass
    else:
        pass


if __name__ == "__main__":
    main()
